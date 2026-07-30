"""
BOM 파싱 모듈 — 취소선 / 음영 셀 인식 포함
"""
import re
import pandas as pd
from openpyxl import load_workbook


def _is_strike(cell):
    try:
        return bool(cell.font and cell.font.strike)
    except Exception:
        return False


def _highlight_color(cell):
    """음영이 있으면 RGB 문자열 반환, 없으면 None"""
    try:
        fill = cell.fill
        if fill and fill.fill_type == 'solid':
            c = fill.fgColor
            if c.type == 'rgb':
                rgb = c.rgb  # 'FFRRGGBB' 형식
                if rgb not in ('FFFFFFFF', 'FF000000', '00000000', 'FFffffff', '000000'):
                    return rgb[-6:]  # RRGGBB 반환
    except Exception:
        pass
    return None


def parse_bom(filepath: str, with_vc_specs: bool = False):
    """
    Returns:
        rows           - 전체 BOM 행 리스트
        variant_cols   - {col_idx(0-based): vc_code} 딕셔너리
        struck_parts   - 취소선 부품 리스트
        highlighted_parts - 음영 부품 리스트
        vc_specs       - {vc_code: {사양 헤더 문자열}} (with_vc_specs=True 일 때만 추가 반환)

    with_vc_specs 는 기존 호출부(4개 값 언패킹)를 깨지 않기 위한 선택 인자다.
    """
    df  = pd.read_excel(filepath, header=None, sheet_name=0)
    wb2 = load_workbook(filepath, data_only=True)
    ws2 = wb2.active

    # ── Variant 컬럼 위치 파악 (col ≥ 21, 3자리 숫자) ────────────────────
    # VC 번호가 적힌 행은 양식마다 다르다 — DYA 표준양식(bom_generator)은 4행
    # (MATRIX_VC_ROW=4)이지만 HKMC 수령 양식은 8행에 있는 등 고정할 수 없다.
    # 예전에는 4행으로 고정해서 HKMC 양식을 올리면 VC가 0개로 잡혀 수량 기반
    # 검증이 전부 무동작하고 «오사양 누락»만 전건 오검출됐다(실측 145건).
    # → 상단 일부 행을 훑어 3자리 숫자가 가장 많은 행을 VC 헤더로 자동 판정한다.
    #   (매트릭스 영역의 데이터 셀은 수량 1~2자리 또는 품번이라 3자리와 겹치지 않음)
    def _count_vc_codes(ri):
        return sum(1 for ci, v in enumerate(df.iloc[ri].tolist())
                   if ci >= 20 and pd.notna(v) and re.match(r'^\d{3}$', str(v).strip()))

    variant_cols = {}
    scan_rows = range(min(12, len(df)))
    vc_row = max(scan_rows, key=_count_vc_codes, default=None)
    if vc_row is not None and _count_vc_codes(vc_row) > 0:
        for ci, val in enumerate(df.iloc[vc_row].tolist()):
            if pd.notna(val) and re.match(r'^\d{3}$', str(val).strip()) and ci >= 20:
                variant_cols[ci] = str(val).strip()

    # ── VC별 사양 헤더 수집 ──────────────────────────────────────────────
    # HKMC 양식은 VC 번호 행 위쪽에 그 VC의 사양이 여러 단으로 적혀 있다
    # (예: T&P / 4W STD / 인조 / MNL / 히터 / USB). 이 값들이 «부품 N열 사양이
    # 배정된 VC와 맞는가» 검증(사양 배정 모순)의 기준이 된다.
    # DYA 표준양식은 VC 행이 4행이고 위쪽에 사양 단이 없어 빈 집합이 되며,
    # 그 경우 해당 검증은 자동으로 건너뛴다.
    vc_specs = {vc: set() for vc in variant_cols.values()}
    if vc_row:
        for ri in range(0, vc_row):
            row = df.iloc[ri].tolist()
            for ci, vc_code in variant_cols.items():
                if ci >= len(row):
                    continue
                val = row[ci]
                if pd.isna(val):
                    continue
                token = str(val).replace('\n', ' ').strip()
                if token and token.lower() != 'nan':
                    vc_specs[vc_code].add(token)

    # ── 취소선 행·셀 / 음영 셀 수집 ─────────────────────────────────────
    strike_pno_rows = set()
    strike_qty      = {}         # {(xl_row, vc_code): True}
    highlight_map   = {}         # {xl_row: rgb_str}

    for ri_xl in range(6, ws2.max_row + 1):
        pno_cell = ws2.cell(ri_xl, 10)
        if _is_strike(pno_cell) and pno_cell.value:
            strike_pno_rows.add(ri_xl)

        # 주요 컬럼(P/NO, P/NAME, DESC) 음영 체크
        for ci_xl in (10, 12, 14):
            rgb = _highlight_color(ws2.cell(ri_xl, ci_xl))
            if rgb:
                highlight_map[ri_xl] = rgb
                break

        for ci_df, vc_code in variant_cols.items():
            qty_cell = ws2.cell(ri_xl, ci_df + 1)
            if _is_strike(qty_cell) and qty_cell.value:
                try:
                    float(qty_cell.value)
                    strike_qty[(ri_xl, vc_code)] = True
                except Exception:
                    pass
    wb2.close()

    # ── 행 파싱 ──────────────────────────────────────────────────────────
    rows, struck_parts, highlighted_parts = [], [], []

    for ri in range(5, len(df)):
        ri_xl = ri + 1
        row   = df.iloc[ri].tolist()

        vc     = str(row[0]).strip()  if pd.notna(row[0])  else ''
        pno    = str(row[9]).strip()  if pd.notna(row[9])  else ''
        pname  = str(row[11]).strip() if pd.notna(row[11]) else ''
        desc   = str(row[13]).strip() if pd.notna(row[13]) else ''
        mat    = str(row[21]).strip() if pd.notna(row[21]) else ''
        region = str(row[20]).strip() if pd.notna(row[20]) else ''

        level = None
        for lc in range(1, 9):
            if pd.notna(row[lc]) and str(row[lc]) not in ('nan', ''):
                try:
                    level = int(float(row[lc]))
                    break
                except Exception:
                    pass

        is_pno_struck = ri_xl in strike_pno_rows
        if is_pno_struck and pno:
            struck_parts.append({'row_idx': ri_xl, 'pno': pno,
                                  'pname': pname, 'desc': desc})

        is_hl = ri_xl in highlight_map
        if is_hl and pno and not is_pno_struck:
            highlighted_parts.append({'row_idx': ri_xl, 'pno': pno,
                                       'pname': pname, 'desc': desc,
                                       'color': highlight_map[ri_xl]})

        qtys, qtys_struck = {}, {}
        for ci, vc_code in variant_cols.items():
            val = row[ci]
            if pd.notna(val) and str(val) not in ('nan', '', '0'):
                try:
                    q = float(val)
                    if strike_qty.get((ri_xl, vc_code)):
                        qtys_struck[vc_code] = q
                    else:
                        qtys[vc_code] = q
                except Exception:
                    pass

        rows.append({
            'row_idx': ri_xl, 'vc': vc, 'level': level,
            'pno': pno, 'pname': pname, 'desc': desc,
            'mat': mat, 'region': region,
            'qtys': qtys, 'qtys_struck': qtys_struck,
            'is_pno_struck': is_pno_struck,
            'is_section': vc.startswith('■'),
            'is_highlighted': is_hl,
        })

    if with_vc_specs:
        return rows, variant_cols, struck_parts, highlighted_parts, vc_specs
    return rows, variant_cols, struck_parts, highlighted_parts
