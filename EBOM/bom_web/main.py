"""
DYA BOM 검증 웹 서버 — FastAPI
"""
import os, sys, shutil, tempfile, uuid, re, json, threading
from typing import Optional
from fastapi import FastAPI, UploadFile, File, Request, Form
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

from auth import (init_db, create_user, get_user, verify_pw, create_token,
                  current_user, require_login, require_admin,
                  get_all_users, approve_user, reject_user, delete_user, set_role, set_name,
                  get_all_vehicle_codes, add_vehicle_code, update_vehicle_code, delete_vehicle_code,
                  get_vehicle_code_by_code, update_vehicle_code_by_code, delete_vehicle_code_by_code,
                  get_vehicle_by_id, get_vehicle_by_code_mfg,
                  get_production_qty_rows, upsert_production_qty, delete_production_qty,
                  get_production_summary,
                  save_stored_bom, list_stored_boms, get_stored_bom, delete_stored_bom,
                  update_stored_bom_meta, find_duplicate_by_hash,
                  list_bom_template_revisions, get_active_bom_template,
                  get_bom_template_revision, add_bom_template_revision,
                  activate_bom_template_revision, delete_bom_template_revision,
                  update_bom_template_note,
                  add_bom_generate_history, get_bom_generate_history_list,
                  get_bom_generate_history, delete_bom_generate_history,
                  add_ccc_upload, save_ccc_items, get_ccc_uploads, get_ccc_upload,
                  get_ccc_items, get_ccc_items_by_vehicle, delete_ccc_upload,
                  upsert_sales_price, get_sales_prices,
                  add_ebom_upload, save_ebom_items, replace_ebom_items, get_ebom_uploads, get_ebom_upload,
                  get_ebom_items, get_ebom_items_by_vehicle, get_ebom_board_revisions,
                  delete_ebom_upload,
                  get_all_country_codes, upsert_country_code, delete_country_code, get_country_code,
                  get_all_dev_stages, get_dev_stage_codes, upsert_dev_stage, delete_dev_stage,
                  get_all_part_names, upsert_part_name, delete_part_name,
                  get_all_fabric_codes, upsert_fabric_code, delete_fabric_code,
                  add_mbom_history, add_mbom_file, get_mbom_history, get_mbom_history_post,
                  get_mbom_file, get_mbom_files_by_post, delete_mbom_history,
                  get_latest_mbom_post_with_alc, get_mbom_posts_with_files,
                  EO_FIELDS, EO_COLUMN_ALIASES, EO_REASONS, EO_APPROVAL_ROLES,
                  EO_APPROVAL_STATUS, submit_eo_approval, act_eo_approval,
                  reopen_eo_approval, get_eo_approvals,
                  add_eo_mail, get_eo_mails, get_user_emails,
                  upsert_eo_notices, search_eo_notices,
                  get_eo_notice, update_eo_notice, delete_eo_notice, add_eo_file,
                  get_eo_files, get_eo_file, delete_eo_file, get_eo_stats,
                  set_eo_items, get_eo_items,
                  add_doc_post, get_doc_posts, get_doc_post, update_doc_post, delete_doc_post,
                  add_doc_file, get_doc_file, delete_doc_file,
                  EO_APPROVAL_TYPES, EO_TYPE_LABEL, EO_TYPE_REFERENCE,
                  parse_org_rows, upsert_org_members, get_org_members, get_org_tree,
                  get_org_stats, delete_org_member, clear_org_members,
                  CATIA_PART_GROUPS, CATIA_GROUP_LABEL, CATIA_STAGES,
                  CONVERT_MAP, get_convert_agent_key, rotate_convert_agent_key,
                  get_convert_queue, get_convert_stats, mark_convert_agent_seen,
                  parse_catia_filename, add_catia_file, find_catia_duplicate,
                  refresh_catia_derived, base_part_no, upsert_parts_from_catia,
                  get_catia_counts, backfill_parts_from_catia,
                  is_design_user, get_user_dept, get_design_keywords, get_bom_part_numbers,
                  DIST_STATUS, get_partners, create_partner_account, get_partner_accounts,
                  create_dist, update_dist, get_dist, search_dist, delete_dist,
                  get_dist_candidates, set_dist_files, get_dist_files,
                  set_dist_targets, get_dist_targets, send_dist,
                  log_dist_download, get_dist_downloads, can_partner_get, get_dist_stats,
                  CUST_EO_TYPES, CUST_EO_FIELDS, CUST_EO_FILE_KINDS,
                  get_customers, upsert_customer, delete_customer, seed_customers,
                  create_cust_eo, update_cust_eo, get_cust_eo, search_cust_eo, delete_cust_eo,
                  set_cust_eo_links, get_cust_eo_links, get_eo_cust_links, set_eo_cust_links,
                  set_cust_eo_rels, get_cust_eo_rels, get_cust_eo_drawings,
                  add_cust_eo_file, get_cust_eo_files, get_cust_eo_file, delete_cust_eo_file,
                  get_cust_eo_stats,
                  CATIA_STATES, CATIA_STATE_LABEL, get_catia_item, get_catia_items_map,
                  catia_checkout, catia_checkin, catia_set_state, catia_can_modify,
                  get_catia_item_log, get_catia_lock_stats,
                  get_catia_facets, search_catia_parts, get_catia_file,
                  update_catia_file, delete_catia_file, get_catia_stats,
                  PART_SPEC_FIELDS, upsert_parts_bulk, search_parts, get_part, update_part,
                  get_part_revs, add_part_file, get_part_files, get_part_file,
                  delete_part_file, get_parts_stats,
                  add_ebom_sheet, save_ebom_sheet_cells, get_ebom_sheets, get_ebom_sheet,
                  get_ebom_sheet_cells, set_ebom_sheet_layout,
                  acquire_ebom_sheet_lock, release_ebom_sheet_lock,
                  get_ebom_sheet_lock_state, apply_ebom_sheet_edits, get_ebom_sheet_revs,
                  get_ebom_sheet_applied_changes, delete_ebom_sheet,
                  get_ebom_sheet_cells_at, revert_ebom_sheet_to, drop_last_ebom_sheet_rev,
                  add_qpart_merge_post, add_qpart_merge_file, get_qpart_merge_history,
                  get_qpart_merge_post, get_qpart_merge_files_by_post, get_qpart_merge_file,
                  delete_qpart_merge_post, add_qpart_merge_run, get_qpart_merge_runs,
                  get_qpart_merge_run,
                  list_country_ppt_revisions, add_country_ppt_revision,
                  delete_country_ppt_revision, get_country_ppt_revision,
                  get_all_process_diagrams, get_process_diagram, add_process_diagram,
                  replace_process_diagram_file, delete_process_diagram,
                  get_flowchart_override, set_flowchart_override, clear_flowchart_override,
                  get_board_guide_image, set_board_guide_image, clear_board_guide_image,
                  MATERIAL_TYPES, get_ccc_matrix, upsert_ccc_matrix, delete_ccc_matrix,
                  get_ccc_codes_for_dropdown,
                  upsert_sales_price_v2, get_sales_prices_v2,
                  add_pel_history, update_pel_history, get_pel_history, get_pel_history_item,
                  delete_pel_history, PEL_STAGE_ORDER, PEL_COLUMN_DIVS,
                  add_pel_spec, get_pel_spec_list, get_pel_spec, delete_pel_spec,
                  get_pel_spec_latest_by_factory, get_pel_spec_row_levels,
                  add_sales_file, get_sales_file_list, get_sales_file, delete_sales_file,
                  update_sales_file_edits)

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(BASE_DIR, 'reports')
DATA_DIR    = os.path.join(BASE_DIR, 'data')
os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

PEL_CODE_PATH = os.path.join(DATA_DIR, 'pel_code_master.xlsx')
STORED_BOM_DIR = os.path.join(DATA_DIR, 'stored_boms')
os.makedirs(STORED_BOM_DIR, exist_ok=True)

init_db()

STATIC_DIR = os.path.join(BASE_DIR, 'static')
os.makedirs(STATIC_DIR, exist_ok=True)

app       = FastAPI(title='DYA BOM 검증 시스템')
app.mount('/static', StaticFiles(directory=STATIC_DIR), name='static')
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, 'templates'))


# ── 인증 페이지 ───────────────────────────────────────────────────────────────
@app.get('/login', response_class=HTMLResponse)
async def login_page(request: Request, next: str = '/'):
    user = current_user(request)
    if user:
        return RedirectResponse('/', status_code=302)
    return templates.TemplateResponse(request=request, name='login.html',
                                      context={'next': next, 'error': ''})


@app.post('/login', response_class=HTMLResponse)
async def login_post(request: Request,
                     username: str = Form(...),
                     password: str = Form(...),
                     next: str    = Form('/login')):
    user = get_user(username)
    error = ''
    if not user or not verify_pw(password, user['hashed_pw']):
        error = '아이디 또는 비밀번호가 올바르지 않습니다.'
    elif user['role'] == 'pending':
        error = '관리자 승인 대기 중입니다. 승인 후 이용하실 수 있습니다.'
    elif user['role'] == 'rejected':
        error = '승인이 거부된 계정입니다. 관리자에게 문의하세요.'

    if error:
        return templates.TemplateResponse(request=request, name='login.html',
                                          context={'next': next, 'error': error})

    token = create_token(username)
    redirect_url = next if next.startswith('/') else '/'
    resp = RedirectResponse(redirect_url, status_code=302)
    resp.set_cookie('bom_token', token, httponly=True, max_age=60*60*8, samesite='lax')
    return resp


@app.get('/logout')
async def logout():
    resp = RedirectResponse('/login', status_code=302)
    resp.delete_cookie('bom_token')
    return resp


@app.get('/register', response_class=HTMLResponse)
async def register_page(request: Request):
    return templates.TemplateResponse(request=request, name='register.html',
                                      context={'error': '', 'success': False})


@app.post('/register', response_class=HTMLResponse)
async def register_post(request: Request,
                        username: str = Form(...),
                        email:    str = Form(...),
                        password: str = Form(...),
                        password2:str = Form(...),
                        name:     str = Form(...),
                        dept:     str = Form('')):
    error = ''
    if password != password2:
        error = '비밀번호가 일치하지 않습니다.'
    elif len(password) < 6:
        error = '비밀번호는 6자 이상이어야 합니다.'
    elif len(username) < 3:
        error = '아이디는 3자 이상이어야 합니다.'
    elif not name.strip():
        error = '이름을 입력해주세요.'

    if not error:
        result = create_user(username, email, password, dept, name)
        if not result['ok']:
            error = result['msg']

    if error:
        return templates.TemplateResponse(request=request, name='register.html',
                                          context={'error': error, 'success': False})
    return templates.TemplateResponse(request=request, name='register.html',
                                      context={'error': '', 'success': True})


# ── 관리자 페이지 ─────────────────────────────────────────────────────────────
@app.get('/admin', response_class=HTMLResponse)
async def admin_page(request: Request):
    redir = require_admin(request)
    if redir: return redir
    users = get_all_users()
    me = current_user(request)
    vcodes = get_all_vehicle_codes()
    return templates.TemplateResponse(request=request, name='admin.html',
                                      context={'users': users, 'me': me, 'vcodes': vcodes})


@app.post('/admin/approve/{user_id}')
async def admin_approve(request: Request, user_id: int):
    if require_admin(request): return RedirectResponse('/login', status_code=302)
    approve_user(user_id)
    return RedirectResponse('/admin', status_code=302)


@app.post('/admin/reject/{user_id}')
async def admin_reject(request: Request, user_id: int):
    if require_admin(request): return RedirectResponse('/login', status_code=302)
    reject_user(user_id)
    return RedirectResponse('/admin', status_code=302)


@app.post('/admin/delete/{user_id}')
async def admin_delete(request: Request, user_id: int):
    if require_admin(request): return RedirectResponse('/login', status_code=302)
    delete_user(user_id)
    return RedirectResponse('/admin', status_code=302)


@app.post('/admin/role/{user_id}')
async def admin_role(request: Request, user_id: int, role: str = Form(...)):
    if require_admin(request): return RedirectResponse('/login', status_code=302)
    set_role(user_id, role)
    return RedirectResponse('/admin', status_code=302)


@app.post('/admin/name/{user_id}')
async def admin_name(request: Request, user_id: int, name: str = Form(...)):
    if require_admin(request): return RedirectResponse('/login', status_code=302)
    set_name(user_id, name)
    return RedirectResponse('/admin', status_code=302)


@app.post('/admin/vehicle-code/add')
async def admin_vehicle_add(request: Request,
                             code: str = Form(...), name: str = Form(...), memo: str = Form('')):
    if require_admin(request): return RedirectResponse('/login', status_code=302)
    add_vehicle_code(code, name, memo)
    return RedirectResponse('/admin#tab-vehicle', status_code=302)


@app.post('/admin/vehicle-code/delete/{code_id}')
async def admin_vehicle_delete(request: Request, code_id: int):
    if require_admin(request): return RedirectResponse('/login', status_code=302)
    delete_vehicle_code(code_id)
    return RedirectResponse('/admin#tab-vehicle', status_code=302)


@app.post('/admin/vehicle-code/edit/{code_id}')
async def admin_vehicle_edit(request: Request, code_id: int,
                              code: str = Form(...), name: str = Form(...), memo: str = Form('')):
    if require_admin(request): return RedirectResponse('/login', status_code=302)
    update_vehicle_code(code_id, code, name, memo)
    return RedirectResponse('/admin#tab-vehicle', status_code=302)


# ── 메인 페이지 ───────────────────────────────────────────────────────────────
@app.get('/', response_class=HTMLResponse)
async def index(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='index.html',
                                      context={'me': me, 'stages': get_dev_stage_codes()})


# ── BOM 검증 API ──────────────────────────────────────────────────────────────
@app.post('/validate')
def validate(request: Request, file: UploadFile = File(...)):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)

    if not file.filename.endswith(('.xlsx', '.xlsm')):
        return JSONResponse({'error': 'xlsx 또는 xlsm 파일만 지원합니다.'}, status_code=400)

    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        from bom_parser import parse_bom
        from validators import validate_bom
        from report import make_report

        rows, variant_cols, struck_parts, highlighted_parts, vc_specs = parse_bom(
            tmp_path, with_vc_specs=True)
        errors, lv1_by_vc = validate_bom(rows, variant_cols, vc_specs=vc_specs)

        report_id   = uuid.uuid4().hex[:10]
        report_path = os.path.join(REPORTS_DIR, f'BOM_검증_{report_id}.xlsx')
        make_report(file.filename, errors, lv1_by_vc, variant_cols,
                    struck_parts, highlighted_parts, report_path)

        return JSONResponse({
            'filename':          file.filename,
            'variant_count':     len(lv1_by_vc),
            'struck_count':      len(struck_parts),
            'highlighted_count': len(highlighted_parts),
            'err_count':         sum(1 for e in errors if e['severity'] == 'ERROR'),
            'warn_count':        sum(1 for e in errors if e['severity'] == 'WARNING'),
            'report_id':         report_id,
            'errors':            errors,
            'lv1_variants': [
                {
                    'vc':          vc,
                    'pno':         r['pno'],
                    'desc':        r['desc'],
                    'has_error':   any(e['variant'] == vc and e['severity'] == 'ERROR'   for e in errors),
                    'has_warning': any(e['variant'] == vc and e['severity'] == 'WARNING' for e in errors),
                }
                for vc, r in sorted(lv1_by_vc.items())
            ],
        })

    except Exception as ex:
        import traceback
        return JSONResponse({'error': f'파싱 오류: {str(ex)}',
                             'trace': traceback.format_exc()}, status_code=500)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.get('/viewer', response_class=HTMLResponse)
async def viewer_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='index.html',
                                      context={'me': me, 'stages': get_dev_stage_codes()})


VIEWER_FILES: dict = {}  # file_id -> (path, original_filename)

@app.post('/view-excel')
def view_excel(request: Request, file: UploadFile = File(...)):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)

    if not file.filename.endswith(('.xlsx', '.xlsm')):
        return JSONResponse({'error': 'xlsx 또는 xlsm 파일만 지원합니다.'}, status_code=400)

    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        from excel_viewer import parse_excel
        sheets = parse_excel(tmp_path)
        file_id = uuid.uuid4().hex[:12]
        VIEWER_FILES[file_id] = (tmp_path, file.filename)
        return JSONResponse({'filename': file.filename, 'sheets': sheets, 'file_id': file_id})
    except Exception as ex:
        import traceback
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        return JSONResponse({'error': str(ex), 'trace': traceback.format_exc()}, status_code=500)


@app.get('/download-excel/{file_id}')
async def download_excel(request: Request, file_id: str):
    redir = require_login(request)
    if redir: return redir
    if not re.fullmatch(r'[a-f0-9]{12}', file_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    entry = VIEWER_FILES.get(file_id)
    if not entry or not os.path.exists(entry[0]):
        return JSONResponse({'error': '파일을 찾을 수 없습니다. 다시 업로드해 주세요.'}, status_code=404)
    path, original_name = entry
    media = ('application/vnd.ms-excel.sheet.macroEnabled.12'
             if original_name.endswith('.xlsm')
             else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return FileResponse(path, filename=original_name, media_type=media)


# ── 차종 마스터 ───────────────────────────────────────────────────────────────
@app.get('/vehicles', response_class=HTMLResponse)
async def vehicles_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    rows = get_all_vehicle_codes(distinct=False)  # 마스터 편집: 생산코드별 전체 행
    return templates.TemplateResponse(request=request, name='vehicles.html',
                                      context={'me': me, 'rows': rows})


@app.get('/vehicles/api/list')
async def vehicles_api_list(request: Request):
    """드롭다운/자동완성용 차종 JSON"""
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    rows = get_all_vehicle_codes()  # distinct: 차종당 1개
    return JSONResponse({'vehicles': [{'code': r['code'], 'name': r['name'], 'memo': r.get('memo', ''),
                                       'mfg_code': r.get('mfg_code', ''), 'powertrain': r.get('powertrain', '')}
                                      for r in rows]})


@app.post('/vehicles/row')
async def vehicles_add_row(request: Request):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    try:
        item = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    code = str(item.get('code', '')).strip().upper()
    name = str(item.get('name', '')).strip()
    memo = str(item.get('memo', '')).strip()
    mfg_code = str(item.get('mfg_code', '')).strip().upper()
    powertrain = str(item.get('powertrain', '')).strip()
    if not code or not name:
        return JSONResponse({'error': '코드와 차종명은 필수입니다.'}, status_code=400)
    if get_vehicle_by_code_mfg(code, mfg_code):
        pair = f'{code} / {mfg_code}' if mfg_code else code
        return JSONResponse({'error': f'이미 등록된 조합: {pair}'}, status_code=400)
    r = add_vehicle_code(code, name, memo, mfg_code=mfg_code, powertrain=powertrain)
    if not r.get('ok'):
        return JSONResponse({'error': r.get('msg', '저장 실패')}, status_code=400)
    v = get_vehicle_by_code_mfg(code, mfg_code)
    return JSONResponse({'ok': True, 'code': code, 'id': v['id'] if v else None})


@app.post('/vehicles/row/{item_id:int}')
async def vehicles_update_row(request: Request, item_id: int):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    try:
        item = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    cur = get_vehicle_by_id(item_id)
    if not cur:
        return JSONResponse({'error': '차종 행을 찾을 수 없습니다.'}, status_code=404)
    new_code = str(item.get('code', cur['code'])).strip().upper()
    name = str(item.get('name', '')).strip()
    memo = str(item.get('memo', '')).strip()
    mfg_code = str(item.get('mfg_code', '')).strip().upper()
    powertrain = str(item.get('powertrain', '')).strip()
    if not new_code or not name:
        return JSONResponse({'error': '코드와 차종명은 필수입니다.'}, status_code=400)
    dup = get_vehicle_by_code_mfg(new_code, mfg_code)
    if dup and dup['id'] != item_id:
        pair = f'{new_code} / {mfg_code}' if mfg_code else new_code
        return JSONResponse({'error': f'이미 등록된 조합으로 변경 불가: {pair}'}, status_code=400)
    result = update_vehicle_code(item_id, new_code, name, memo, mfg_code=mfg_code, powertrain=powertrain)
    if not result.get('ok'):
        return JSONResponse({'error': result.get('msg', '저장 실패')}, status_code=400)
    return JSONResponse({'ok': True, 'code': new_code, 'id': item_id})


@app.post('/vehicles/row/{item_id:int}/delete')
async def vehicles_delete_row(request: Request, item_id: int):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    try:
        body = await request.json()
        password = str(body.get('password', ''))
    except Exception:
        password = ''
    if not password:
        return JSONResponse({'error': '비밀번호를 입력해주세요.'}, status_code=400)
    me = current_user(request)
    if not me:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    user_data = get_user(me['username'])
    if not user_data or not verify_pw(password, user_data['hashed_pw']):
        return JSONResponse({'error': '비밀번호가 일치하지 않습니다.'}, status_code=403)
    if not get_vehicle_by_id(item_id):
        return JSONResponse({'error': '차종 행을 찾을 수 없습니다.'}, status_code=404)
    delete_vehicle_code(item_id)
    return JSONResponse({'ok': True})


# ── 저장된 BOM 관리 (Step 2/3) ────────────────────────────────────────────────
@app.post('/bom-storage/save')
async def bom_storage_save(request: Request,
                            file: UploadFile = File(...),
                            vehicle_code: str = Form(...),
                            row_num: str = Form(...),
                            position: str = Form(...),
                            kind: str = Form(...),
                            stage: str = Form(''),
                            memo: str = Form(''),
                            force: str = Form('false')):
    """업로드된 파일을 메타데이터와 함께 영구 저장.
       동일 해시 파일이 같은 (차종/열/위치/kind)에 이미 있으면 409 응답 (force=true 면 무시하고 저장)."""
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    me = current_user(request)

    if kind not in ('verify', 'viewer'):
        return JSONResponse({'error': '잘못된 kind 값'}, status_code=400)
    fname = (file.filename or '').lower()
    if not fname.endswith(('.xlsx', '.xlsm', '.xls')):
        return JSONResponse({'error': 'xlsx/xlsm/xls 파일만 지원합니다.'}, status_code=400)

    # 파일 내용을 한 번 메모리로 읽어 해시 계산
    import hashlib
    file_bytes = await file.read()
    file_hash = hashlib.sha256(file_bytes).hexdigest()

    vc = vehicle_code.strip().upper()
    rn = str(row_num).strip()
    ps = position.strip()
    force_flag = str(force).lower() in ('true', '1', 'yes')

    # 중복 검사
    if not force_flag:
        dup = find_duplicate_by_hash(vc, rn, ps, kind, file_hash)
        if dup:
            return JSONResponse({
                'duplicate': True,
                'existing': {
                    'file_id': dup['file_id'],
                    'version_num': dup['version_num'],
                    'filename': dup['filename'],
                    'uploaded_at': dup['uploaded_at'],
                    'uploader': dup['uploader'],
                    'memo': dup.get('memo', ''),
                },
                'message': '동일한 파일이 이미 같은 위치에 저장되어 있습니다.'
            }, status_code=409)

    file_id = uuid.uuid4().hex[:16]
    ext = os.path.splitext(file.filename or '')[1] or '.xlsx'
    saved_path = os.path.join(STORED_BOM_DIR, f'{file_id}{ext}')
    with open(saved_path, 'wb') as f:
        f.write(file_bytes)

    result = save_stored_bom(
        vehicle_code=vc, row_num=rn, position=ps, kind=kind,
        filename=file.filename or 'unknown.xlsx',
        file_id=file_id, file_path=saved_path,
        uploader=me['username'], memo=memo.strip(),
        file_hash=file_hash, stage=stage.strip(),
    )
    return JSONResponse({'ok': True, 'file_id': file_id, 'version': result.get('version')})


@app.post('/bom-storage/{file_id}/meta')
async def bom_storage_update_meta(request: Request, file_id: str):
    """저장본의 차종/열/위치/메모 사후 수정 (admin)."""
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    entry = get_stored_bom(file_id)
    if not entry:
        return JSONResponse({'error': '저장본을 찾을 수 없습니다.'}, status_code=404)
    try:
        item = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    fields = {}
    if 'vehicle_code' in item: fields['vehicle_code'] = str(item['vehicle_code']).strip().upper()
    if 'row_num' in item:      fields['row_num']      = str(item['row_num']).strip()
    if 'position' in item:     fields['position']     = str(item['position']).strip()
    if 'stage' in item:        fields['stage']        = str(item['stage']).strip()
    if 'memo' in item:         fields['memo']         = str(item['memo']).strip()
    if not fields:
        return JSONResponse({'error': '수정할 필드가 없습니다.'}, status_code=400)
    ok = update_stored_bom_meta(file_id, **fields)
    if not ok:
        return JSONResponse({'error': '수정 실패'}, status_code=400)
    return JSONResponse({'ok': True})


@app.get('/bom-storage/versions')
async def bom_storage_versions(request: Request,
                                vehicle_code: str = '',
                                row_num: str = '',
                                position: str = '',
                                kind: str = ''):
    """저장된 BOM 목록 (조회 카드용)"""
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    rows = list_stored_boms(
        vehicle_code=vehicle_code.strip().upper() if vehicle_code else None,
        row_num=row_num.strip() if row_num else None,
        position=position.strip() if position else None,
        kind=kind.strip() if kind else None,
    )
    return JSONResponse({'versions': rows})


@app.get('/bom-storage/load/{file_id}')
async def bom_storage_load(request: Request, file_id: str):
    """저장된 파일을 재분석. kind에 따라 validate 또는 view-excel 결과 반환."""
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    entry = get_stored_bom(file_id)
    if not entry:
        return JSONResponse({'error': '저장된 BOM을 찾을 수 없습니다.'}, status_code=404)
    path = entry['file_path']
    if not os.path.exists(path):
        return JSONResponse({'error': '파일이 손상되거나 삭제되었습니다.'}, status_code=404)

    if entry['kind'] == 'verify':
        try:
            from bom_parser import parse_bom
            from validators import validate_bom
            from report import make_report
            rows, variant_cols, struck_parts, highlighted_parts, vc_specs = parse_bom(
                path, with_vc_specs=True)
            errors, lv1_by_vc = validate_bom(rows, variant_cols, vc_specs=vc_specs)
            report_id   = uuid.uuid4().hex[:10]
            report_path = os.path.join(REPORTS_DIR, f'BOM_검증_{report_id}.xlsx')
            make_report(entry['filename'], errors, lv1_by_vc, variant_cols,
                        struck_parts, highlighted_parts, report_path)
            return JSONResponse({
                'kind': 'verify',
                'meta': entry,
                'filename':          entry['filename'],
                'variant_count':     len(lv1_by_vc),
                'struck_count':      len(struck_parts),
                'highlighted_count': len(highlighted_parts),
                'err_count':         sum(1 for e in errors if e['severity'] == 'ERROR'),
                'warn_count':        sum(1 for e in errors if e['severity'] == 'WARNING'),
                'report_id':         report_id,
                'errors':            errors,
                'lv1_variants': [
                    {'vc': vc, 'pno': r['pno'], 'desc': r['desc'],
                     'has_error':   any(e['variant'] == vc and e['severity'] == 'ERROR'   for e in errors),
                     'has_warning': any(e['variant'] == vc and e['severity'] == 'WARNING' for e in errors),}
                    for vc, r in sorted(lv1_by_vc.items())
                ],
            })
        except Exception as ex:
            import traceback
            return JSONResponse({'error': f'재분석 오류: {ex}', 'trace': traceback.format_exc()}, status_code=500)
    else:
        # viewer
        try:
            from excel_viewer import parse_excel
            sheets = parse_excel(path)
            view_id = uuid.uuid4().hex[:12]
            VIEWER_FILES[view_id] = (path, entry['filename'])
            return JSONResponse({
                'kind': 'viewer',
                'meta': entry,
                'filename': entry['filename'],
                'sheets': sheets,
                'file_id': view_id
            })
        except Exception as ex:
            import traceback
            return JSONResponse({'error': str(ex), 'trace': traceback.format_exc()}, status_code=500)


@app.post('/bom-storage/{file_id}/delete')
async def bom_storage_delete(request: Request, file_id: str):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    entry = get_stored_bom(file_id)
    if not entry:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    try:
        if os.path.exists(entry['file_path']):
            os.unlink(entry['file_path'])
    except Exception:
        pass
    delete_stored_bom(file_id)
    return JSONResponse({'ok': True})


# ── M-BOM 코드 변경 (생관 — 양식 수신 대기 중) ────────────────────────────────
@app.get('/m-bom', response_class=HTMLResponse)
async def m_bom_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='m_bom.html',
                                      context={'me': me})


# ── BOM 자동 생성 (부품사양서 → BOM) ──────────────────────────────────────────
GENERATED_BOMS: dict = {}  # file_id -> (out_path, filename, source_meta)
# source_meta: [{'plant': '광주', 'spec_path': ..., 'bre_path': ... or None}, ...]
BOM_PLANT_SLOTS = ['광주', '화성']


@app.get('/bom-generate', response_class=HTMLResponse)
async def bom_generate_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    guide_image = get_board_guide_image('bom_generate')
    return templates.TemplateResponse(request=request, name='auto_bom.html',
                                      context={'me': me, 'guide_image': guide_image})


def _save_upload(upfile: UploadFile, dest_path: str, exts: tuple) -> bool:
    """업로드 파일을 dest_path 에 저장. 확장자가 안 맞거나 파일이 없으면 False."""
    if upfile is None or not getattr(upfile, 'filename', ''):
        return False
    ext = os.path.splitext(upfile.filename)[1].lower()
    if ext not in exts:
        return False
    with open(dest_path, 'wb') as f:
        shutil.copyfileobj(upfile.file, f)
    return True


def _build_sources_from_slots(file_id: str, slots: list) -> tuple:
    """slots: [(plant, pel_upload, bre_upload), ...]
    저장 + BRE 파싱까지 마친 (sources, source_meta, errors) 반환.
    sources 는 generate_bom()에 그대로 넘길 수 있는 형태."""
    sources, source_meta, errors = [], [], []
    for plant, pel_file, bre_file in slots:
        if pel_file is None or not getattr(pel_file, 'filename', ''):
            continue
        ext = os.path.splitext(pel_file.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            errors.append(f'{plant} 부품사양서는 xlsx/xls 파일만 지원합니다.')
            continue
        spec_path = os.path.join(REPORTS_DIR, f'spec_{file_id}_{plant}{ext}')
        _save_upload(pel_file, spec_path, ('.xlsx', '.xls'))

        bre_path, bre_info = None, None
        if bre_file is not None and getattr(bre_file, 'filename', ''):
            bext = os.path.splitext(bre_file.filename)[1].lower()
            if bext in ('.xlsm', '.xlsx', '.xls'):
                bre_path = os.path.join(REPORTS_DIR, f'bre_{file_id}_{plant}{bext}')
                _save_upload(bre_file, bre_path, ('.xlsm', '.xlsx', '.xls'))
                try:
                    from bom_generator import parse_bre
                    bre_info = parse_bre(bre_path)
                except Exception:
                    bre_info = None

        sources.append({'spec_path': spec_path, 'bre_info': bre_info, 'plant_label': plant})
        source_meta.append({'plant': plant, 'spec_path': spec_path, 'bre_path': bre_path})
    return sources, source_meta, errors


@app.post('/bom-generate/upload')
def bom_generate_upload(request: Request,
                        pel_gj: UploadFile = File(None), pel_hs: UploadFile = File(None),
                        bre_gj: UploadFile = File(None), bre_hs: UploadFile = File(None)):
    """차종(운전석 등) 기준 — 광주/화성 부품사양서(+선택 BRE)를 각각 올리면
    하나의 표준 BOM으로 합쳐서 생성한다. 공장 중 있는 것만 올려도 된다."""
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    me = current_user(request)

    file_id = uuid.uuid4().hex[:12]
    slots = [('광주', pel_gj, bre_gj), ('화성', pel_hs, bre_hs)]
    sources, source_meta, errors = _build_sources_from_slots(file_id, slots)
    if errors:
        return JSONResponse({'error': ' / '.join(errors)}, status_code=400)
    if not sources:
        return JSONResponse({'error': '광주 또는 화성 부품사양서를 최소 1개 업로드하세요.'}, status_code=400)

    out_name = f'BOM_자동생성_{file_id}.xlsx'
    out_path = os.path.join(REPORTS_DIR, out_name)
    active_tpl = get_active_bom_template()
    tpl_path = active_tpl['file_path'] if active_tpl else None

    try:
        from bom_generator import generate_bom
        result = generate_bom(sources, PEL_CODE_PATH, out_path, template_path=tpl_path)
        GENERATED_BOMS[file_id] = (out_path, 'BOM_자동생성.xlsx', source_meta)
        result['file_id'] = file_id
        if active_tpl:
            result['template_rev'] = active_tpl.get('rev_num')
            result['template_filename'] = active_tpl.get('filename')

        # 영구 이력 기록 — 서버 재시작 후에도 재다운로드 가능하도록 output_path를 DB에 남긴다
        hist_id = add_bom_generate_history(
            vehicle_info=result.get('vehicle_info', ''),
            pel_gj_filename=pel_gj.filename if pel_gj else '',
            pel_hs_filename=pel_hs.filename if pel_hs else '',
            bre_gj_filename=bre_gj.filename if bre_gj else '',
            bre_hs_filename=bre_hs.filename if bre_hs else '',
            template_rev=active_tpl.get('rev_num') if active_tpl else None,
            template_filename=active_tpl.get('filename') if active_tpl else '',
            vc_count=result.get('vc_count', 0),
            matched=result.get('matched', 0),
            unmatched=result.get('unmatched', 0),
            plants_used=json.dumps(result.get('plants_used', []), ensure_ascii=False),
            output_path=out_path,
            output_filename=out_name,
            uploaded_by=me['username'],
        )
        result['history_id'] = hist_id
        return JSONResponse(result)
    except Exception as ex:
        import traceback
        for sm in source_meta:
            for p in (sm.get('spec_path'), sm.get('bre_path')):
                if p and os.path.exists(p):
                    try: os.unlink(p)
                    except Exception: pass
        return JSONResponse({'error': f'BOM 생성 오류: {ex}',
                             'trace': traceback.format_exc()}, status_code=500)


@app.post('/bom-generate/regenerate/{file_id}')
def bom_generate_regenerate(request: Request, file_id: str):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    if not re.fullmatch(r'[a-f0-9]{12}', file_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    entry = GENERATED_BOMS.get(file_id)
    if not entry:
        return JSONResponse({'error': '원본 파일이 만료되었습니다. 다시 업로드해주세요.'}, status_code=404)
    out_path, orig_name, source_meta = entry
    sources = []
    for sm in source_meta:
        if not os.path.exists(sm['spec_path']):
            return JSONResponse({'error': f"{sm['plant']} 원본 파일이 없습니다. 다시 업로드해주세요."}, status_code=404)
        bre_info = None
        if sm.get('bre_path') and os.path.exists(sm['bre_path']):
            try:
                from bom_generator import parse_bre
                bre_info = parse_bre(sm['bre_path'])
            except Exception:
                bre_info = None
        sources.append({'spec_path': sm['spec_path'], 'bre_info': bre_info, 'plant_label': sm['plant']})

    active_tpl = get_active_bom_template()
    tpl_path = active_tpl['file_path'] if active_tpl else None
    try:
        from bom_generator import generate_bom
        result = generate_bom(sources, PEL_CODE_PATH, out_path, template_path=tpl_path)
        result['file_id'] = file_id
        if active_tpl:
            result['template_rev'] = active_tpl.get('rev_num')
            result['template_filename'] = active_tpl.get('filename')
        return JSONResponse(result)
    except Exception as ex:
        import traceback
        return JSONResponse({'error': f'재생성 오류: {ex}',
                             'trace': traceback.format_exc()}, status_code=500)


@app.get('/bom-generate/download/{file_id}')
async def bom_generate_download(request: Request, file_id: str):
    redir = require_login(request)
    if redir: return redir
    if not re.fullmatch(r'[a-f0-9]{12}', file_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    entry = GENERATED_BOMS.get(file_id)
    if not entry or not os.path.exists(entry[0]):
        return JSONResponse({'error': '파일을 찾을 수 없습니다. 다시 생성해주세요.'}, status_code=404)
    path, orig = entry[0], entry[1]
    base = os.path.splitext(orig)[0]
    dl_name = f'{base}_BOM.xlsx'
    return FileResponse(path, filename=dl_name,
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.get('/bom-generate/history')
async def bom_generate_history_list(request: Request):
    """영구 이력 목록 — 서버 재시작 후에도 유지(DB 기반)."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    items = get_bom_generate_history_list()
    for it in items:
        try:
            it['plants_used'] = json.loads(it.get('plants_used') or '[]')
        except Exception:
            it['plants_used'] = []
        it['file_exists'] = bool(it.get('output_path') and os.path.exists(it['output_path']))
    return JSONResponse({'items': items})


@app.get('/bom-generate/history/download/{item_id:int}')
async def bom_generate_history_download(request: Request, item_id: int):
    """이력에 남은 결과물 재다운로드 — 서버 재시작으로 GENERATED_BOMS(메모리)가 비어도 동작."""
    redir = require_login(request)
    if redir: return redir
    item = get_bom_generate_history(item_id)
    if not item or not item.get('output_path') or not os.path.exists(item['output_path']):
        return JSONResponse({'error': '파일을 찾을 수 없습니다. 재업로드가 필요합니다.'}, status_code=404)
    base = os.path.splitext(item.get('output_filename') or 'BOM.xlsx')[0]
    return FileResponse(item['output_path'], filename=f'{base}_BOM.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.post('/bom-generate/history/{item_id:int}/delete')
async def bom_generate_history_delete(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    me = current_user(request)
    item = get_bom_generate_history(item_id)
    if not item:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and item['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    delete_bom_generate_history(item_id)
    if item.get('output_path') and os.path.exists(item['output_path']):
        try: os.unlink(item['output_path'])
        except Exception: pass
    return JSONResponse({'ok': True})


# ── 표준화 BOM 템플릿 (리비전 관리) ────────────────────────────────────────────
BOM_TEMPLATE_DIR = os.path.join(DATA_DIR, 'bom_templates')
os.makedirs(BOM_TEMPLATE_DIR, exist_ok=True)
# 호환용: 첫 업로드 때 만들어진 단일 템플릿이 있으면 자동 마이그레이션
_LEGACY_TPL = os.path.join(DATA_DIR, 'bom_template.xlsx')
_LEGACY_META = os.path.join(DATA_DIR, 'bom_template_meta.json')


def _migrate_legacy_template():
    if not os.path.exists(_LEGACY_TPL):
        return
    if list_bom_template_revisions():
        return  # 이미 리비전 있음 → 마이그레이션 불필요
    import json
    meta = {}
    if os.path.exists(_LEGACY_META):
        try:
            with open(_LEGACY_META, 'r', encoding='utf-8') as f:
                meta = json.load(f)
        except Exception:
            pass
    fname = meta.get('filename') or 'bom_template.xlsx'
    uploader = meta.get('uploaded_by') or 'admin'
    new_path = os.path.join(BOM_TEMPLATE_DIR, 'rev_001.xlsx')
    shutil.copy2(_LEGACY_TPL, new_path)
    add_bom_template_revision(fname, new_path, uploader,
                              note='최초 업로드 (자동 마이그레이션)')


_migrate_legacy_template()


@app.get('/bom-generate/template/info')
async def bom_template_info(request: Request):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    active = get_active_bom_template()
    return JSONResponse({'exists': active is not None, 'info': active})


@app.get('/bom-generate/template/list')
async def bom_template_list(request: Request):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    revs = list_bom_template_revisions()
    # 파일 존재 여부도 같이 알려줌
    for r in revs:
        r['file_exists'] = os.path.exists(r.get('file_path', ''))
    return JSONResponse({'revisions': revs})


@app.post('/bom-generate/template/upload')
def bom_template_upload(request: Request,
                              file: UploadFile = File(...),
                              note: str = Form('')):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    me = current_user(request)
    if not me or me.get('role') != 'admin':
        return JSONResponse({'error': '관리자만 업로드할 수 있습니다.'}, status_code=403)
    fname = (file.filename or '').lower()
    if not fname.endswith('.xlsx'):
        return JSONResponse({'error': 'xlsx 파일만 지원합니다.'}, status_code=400)

    # 다음 리비전 번호 계산
    existing = list_bom_template_revisions()
    next_rev = (max((r['rev_num'] for r in existing), default=0)) + 1
    save_path = os.path.join(BOM_TEMPLATE_DIR, f'rev_{next_rev:03d}.xlsx')
    with open(save_path, 'wb') as f:
        shutil.copyfileobj(file.file, f)

    info = add_bom_template_revision(
        filename=file.filename or f'bom_template_rev{next_rev}.xlsx',
        file_path=save_path,
        uploaded_by=me.get('username', ''),
        note=(note or '').strip(),
    )
    return JSONResponse({'ok': True, 'info': info})


@app.post('/bom-generate/template/{rev_id}/activate')
async def bom_template_activate(request: Request, rev_id: int):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    me = current_user(request)
    if not me or me.get('role') != 'admin':
        return JSONResponse({'error': '관리자만 변경할 수 있습니다.'}, status_code=403)
    ok = activate_bom_template_revision(rev_id)
    if not ok:
        return JSONResponse({'error': '해당 리비전을 찾을 수 없습니다.'}, status_code=404)
    return JSONResponse({'ok': True, 'active': get_active_bom_template()})


@app.post('/bom-generate/template/{rev_id}/delete')
async def bom_template_delete(request: Request, rev_id: int):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    me = current_user(request)
    if not me or me.get('role') != 'admin':
        return JSONResponse({'error': '관리자만 삭제할 수 있습니다.'}, status_code=403)
    info = delete_bom_template_revision(rev_id)
    if not info:
        return JSONResponse({'error': '해당 리비전을 찾을 수 없습니다.'}, status_code=404)
    # 물리 파일 삭제
    fp = info.get('file_path')
    if fp and os.path.exists(fp):
        try: os.unlink(fp)
        except Exception: pass
    return JSONResponse({'ok': True, 'deleted_rev': info.get('rev_num'),
                         'active': get_active_bom_template()})


@app.post('/bom-generate/template/{rev_id}/note')
async def bom_template_set_note(request: Request, rev_id: int, note: str = Form('')):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    me = current_user(request)
    if not me or me.get('role') != 'admin':
        return JSONResponse({'error': '관리자만 변경할 수 있습니다.'}, status_code=403)
    ok = update_bom_template_note(rev_id, (note or '').strip())
    if not ok:
        return JSONResponse({'error': '해당 리비전을 찾을 수 없습니다.'}, status_code=404)
    return JSONResponse({'ok': True})


@app.get('/bom-generate/template/{rev_id}/download')
async def bom_template_download_rev(request: Request, rev_id: int):
    redir = require_login(request)
    if redir: return redir
    info = get_bom_template_revision(rev_id)
    if not info or not os.path.exists(info['file_path']):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(info['file_path'], filename=info['filename'],
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.get('/bom-generate/template/download')
async def bom_template_download(request: Request):
    redir = require_login(request)
    if redir: return redir
    active = get_active_bom_template()
    if not active or not os.path.exists(active['file_path']):
        return JSONResponse({'error': '활성 템플릿이 없습니다.'}, status_code=404)
    return FileResponse(active['file_path'], filename=active['filename'],
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── PEL CODE 마스터 ───────────────────────────────────────────────────────────
PEL_STD_COLS = ['구분', 'CODE', '사양', '설명', '비고', '옵션그룹', '표시순서']  # 표준 7컬럼
PEL_COL_ALIASES = {
    '사양': ['사양', '명칭', 'NAME', 'SPEC'],
    '비고': ['비고', '분류', 'CATEGORY', 'CLASS', 'NOTE', 'REMARK'],
    '구분': ['구분', 'TYPE', 'KIND'],
    'CODE': ['CODE', 'PEL', 'PELCODE'],
    '설명': ['설명', 'DESCRIPTION', 'DESC'],
    '옵션그룹': ['옵션그룹', '옵션 그룹', 'GROUP', 'OPTGROUP', '그룹'],
    '표시순서': ['표시순서', '표시 순서', 'ORDER', 'SORT', '순서'],
}


def _normalize_pel_df(df):
    """기존 컬럼명을 표준 5컬럼으로 매핑/정렬. 누락 컬럼은 빈 컬럼 추가."""
    import pandas as pd
    rename = {}
    used = set()
    for std, aliases in PEL_COL_ALIASES.items():
        for c in df.columns:
            if c in used: continue
            cs = str(c).strip()
            cu = cs.upper()
            for al in aliases:
                if cs == al or cu == al.upper():
                    rename[c] = std
                    used.add(c)
                    break
            if c in used: break
    df = df.rename(columns=rename)
    # 표준 컬럼 외 잔여 컬럼은 그대로 유지(끝쪽), 표준 컬럼이 없으면 추가
    for std in PEL_STD_COLS:
        if std not in df.columns:
            df[std] = ''
    # 표준 컬럼 먼저, 나머지 뒤
    extras = [c for c in df.columns if c not in PEL_STD_COLS]
    df = df[PEL_STD_COLS + extras]
    return df


def _gubun_sort_key(v):
    """구분 정렬 키 (튜플 — hashable): '1열' → (1,''), 그 외는 큰 값."""
    s = str(v or '').strip()
    if not s: return (9999, '')
    m = re.match(r'^(\d+)\s*열', s)
    if m: return (int(m.group(1)), '')
    if s.isdigit(): return (int(s), '')
    return (9000, s)


def _code_sort_key(v):
    """CODE 자연정렬 키 (튜플 — hashable): 숫자는 (0, int), 문자는 (1, str)."""
    s = str(v or '').strip()
    parts = []
    for t in re.findall(r'\d+|\D+', s):
        if t.isdigit():
            parts.append((0, int(t)))
        else:
            parts.append((1, t.lower()))
    return tuple(parts)


def _load_pel_df():
    """PEL 마스터 DataFrame을 표준 컬럼으로 로드 + 기본 정렬 (구분→CODE). 없으면 빈 DF."""
    import pandas as pd
    if not os.path.exists(PEL_CODE_PATH):
        return pd.DataFrame(columns=PEL_STD_COLS)
    # dtype=str 로 읽는다. 숫자로 추론되면(예: 표시순서 int64) 행 저장 시
    # 문자열 대입이 pandas TypeError를 일으켜 500이 난다.
    df = pd.read_excel(PEL_CODE_PATH, sheet_name=0, dtype=str).fillna('')
    df = _normalize_pel_df(df)
    # 기본 정렬: 구분(1열→2열→3열…) → CODE(영숫자 자연정렬)
    if '구분' in df.columns and 'CODE' in df.columns and len(df):
        order = sorted(
            range(len(df)),
            key=lambda i: (_gubun_sort_key(df['구분'].iloc[i]),
                           _code_sort_key(df['CODE'].iloc[i])),
        )
        df = df.iloc[order].reset_index(drop=True)
    return df


def _save_pel_df(df):
    df.to_excel(PEL_CODE_PATH, index=False)


def _read_pel_code():
    """기존 시그니처 호환: (cols, rows, mtime) 반환."""
    if not os.path.exists(PEL_CODE_PATH):
        return [], [], None
    try:
        df = _load_pel_df()
        cols = [str(c) for c in df.columns]
        def _cell(c):
            s = str(c)
            if s.endswith('.0') and s[:-2].lstrip('-').isdigit():  # 10.0 → 10 (표시순서 등)
                return s[:-2]
            return '' if s == 'nan' else s
        rows = [[_cell(c) for c in row] for row in df.values.tolist()]
        from datetime import datetime
        mtime = datetime.fromtimestamp(os.path.getmtime(PEL_CODE_PATH)).strftime('%Y-%m-%d %H:%M')
        return cols, rows, mtime
    except Exception:
        return [], [], None


@app.get('/pel-code', response_class=HTMLResponse)
async def pel_code_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    cols, rows, mtime = _read_pel_code()
    return templates.TemplateResponse(request=request, name='pel_code.html',
                                      context={'me': me, 'cols': cols, 'rows': rows,
                                               'mtime': mtime or '-'})


@app.post('/pel-code/upload')
def pel_code_upload(request: Request, file: UploadFile = File(...)):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    fname = (file.filename or '').lower()
    if not fname.endswith(('.xls', '.xlsx')):
        return JSONResponse({'error': 'xls 또는 xlsx 파일만 지원합니다.'}, status_code=400)
    suffix = os.path.splitext(fname)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name
    try:
        import pandas as pd
        df = pd.read_excel(tmp_path, sheet_name=0)
        df.to_excel(PEL_CODE_PATH, index=False)
        return RedirectResponse('/pel-code', status_code=302)
    except Exception as ex:
        return JSONResponse({'error': f'파일 처리 오류: {ex}'}, status_code=400)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.get('/pel-code/download')
async def pel_code_download(request: Request):
    redir = require_login(request)
    if redir: return redir
    if not os.path.exists(PEL_CODE_PATH):
        return JSONResponse({'error': '파일이 없습니다.'}, status_code=404)
    return FileResponse(PEL_CODE_PATH, filename='PEL_CODE_마스터.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.get('/pel-code/api/list')
async def pel_code_api_list(request: Request, q: str = ''):
    """사이드 패널/검색용 PEL 마스터 JSON 조회"""
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    cols, rows, mtime = _read_pel_code()
    if q:
        ql = q.lower().strip()
        rows = [r for r in rows if any(ql in str(c).lower() for c in r)]
    total = len(rows)
    truncated = total > 300
    return JSONResponse({
        'cols': cols,
        'rows': rows[:300],
        'total': total,
        'truncated': truncated,
        'mtime': mtime or '',
    })


@app.post('/pel-code/row')
async def pel_code_add_row(request: Request):
    """단일 행 추가 — 관리자(admin) 만 가능."""
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    try:
        item = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    code = str(item.get('CODE', item.get('code', ''))).strip()
    if not code:
        return JSONResponse({'error': 'CODE는 필수 입력입니다.'}, status_code=400)
    df = _load_pel_df()
    existing = set(str(x).strip() for x in df['CODE'].astype(str))
    if code in existing:
        return JSONResponse({'error': f'이미 존재하는 CODE: {code}'}, status_code=400)
    new_row = {
        '구분': str(item.get('구분', '')).strip(),
        'CODE': code,
        '사양': str(item.get('사양', item.get('명칭', ''))).strip(),
        '설명': str(item.get('설명', '')).strip(),
        '비고': str(item.get('비고', item.get('분류', ''))).strip(),
        '옵션그룹': str(item.get('옵션그룹', '')).strip(),
        '표시순서': str(item.get('표시순서', '')).strip(),
    }
    import pandas as pd
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    _save_pel_df(df)
    return JSONResponse({'ok': True, 'code': code, 'total': len(df)})


@app.post('/pel-code/row/{code}')
async def pel_code_update_row(request: Request, code: str):
    """단일 행 업데이트 — 코드 변경도 가능. 관리자만 가능."""
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    try:
        item = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    old_code = code.strip()
    new_code = str(item.get('CODE', item.get('code', old_code))).strip()
    if not new_code:
        return JSONResponse({'error': 'CODE는 필수 입력입니다.'}, status_code=400)
    df = _load_pel_df()
    mask = df['CODE'].astype(str).str.strip() == old_code
    if not mask.any():
        return JSONResponse({'error': f'CODE {old_code}을(를) 찾을 수 없습니다.'}, status_code=404)
    # CODE 변경 시 중복 체크
    if new_code != old_code:
        if (df['CODE'].astype(str).str.strip() == new_code).any():
            return JSONResponse({'error': f'이미 존재하는 CODE로 변경 불가: {new_code}'}, status_code=400)
    idx = df.index[mask][0]
    df.at[idx, '구분'] = str(item.get('구분', df.at[idx, '구분'])).strip()
    df.at[idx, 'CODE'] = new_code
    df.at[idx, '사양'] = str(item.get('사양', item.get('명칭', df.at[idx, '사양']))).strip()
    df.at[idx, '설명'] = str(item.get('설명', df.at[idx, '설명'])).strip()
    df.at[idx, '비고'] = str(item.get('비고', item.get('분류', df.at[idx, '비고']))).strip()
    if '옵션그룹' in df.columns: df.at[idx, '옵션그룹'] = str(item.get('옵션그룹', df.at[idx, '옵션그룹'])).strip()
    if '표시순서' in df.columns: df.at[idx, '표시순서'] = str(item.get('표시순서', df.at[idx, '표시순서'])).strip()
    _save_pel_df(df)
    return JSONResponse({'ok': True, 'code': new_code})


@app.post('/pel-code/row/{code}/delete')
async def pel_code_delete_row(request: Request, code: str):
    """단일 행 삭제 — 관리자 + 비밀번호 재확인."""
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)

    # 본인 비밀번호 재확인
    try:
        body = await request.json()
        password = str(body.get('password', ''))
    except Exception:
        password = ''
    if not password:
        return JSONResponse({'error': '비밀번호를 입력해주세요.'}, status_code=400)

    me = current_user(request)
    if not me:
        return JSONResponse({'error': '로그인이 필요합니다.'}, status_code=401)
    user_data = get_user(me['username'])
    if not user_data or not verify_pw(password, user_data['hashed_pw']):
        return JSONResponse({'error': '비밀번호가 일치하지 않습니다.'}, status_code=403)

    df = _load_pel_df()
    mask = df['CODE'].astype(str).str.strip() == code.strip()
    if not mask.any():
        return JSONResponse({'error': f'CODE {code}을(를) 찾을 수 없습니다.'}, status_code=404)
    df = df[~mask].reset_index(drop=True)
    _save_pel_df(df)
    return JSONResponse({'ok': True, 'remaining': len(df)})


@app.post('/pel-code/bulk-add')
async def pel_code_bulk_add(request: Request):
    """누락 PEL 코드 일괄 추가 — 관리자만 가능."""
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청 형식'}, status_code=400)
    items = body.get('items') or []
    if not isinstance(items, list) or not items:
        return JSONResponse({'error': '추가할 항목이 없습니다.'}, status_code=400)

    df = _load_pel_df()
    existing = set(str(x).strip() for x in df['CODE'].astype(str))

    def pick(item, keys):
        for k in keys:
            v = item.get(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return ''

    new_rows, added, skipped = [], 0, 0
    for item in items:
        code = pick(item, ['code', 'CODE'])
        spec = pick(item, ['사양', 'name', '명칭', 'spec'])
        if not code or not spec:
            skipped += 1
            continue
        if code in existing:
            skipped += 1
            continue
        new_rows.append({
            '구분': pick(item, ['구분', 'gubun']),
            'CODE': code,
            '사양': spec,
            '설명': pick(item, ['설명', 'desc', 'description']),
            '비고': pick(item, ['비고', '분류', 'category', 'note', 'remark']),
        })
        existing.add(code)
        added += 1

    if new_rows:
        import pandas as pd
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
        _save_pel_df(df)

    return JSONResponse({'added': added, 'skipped': skipped, 'total': len(items)})


@app.get('/download/{report_id}')
async def download(request: Request, report_id: str):
    redir = require_login(request)
    if redir: return redir
    if not re.fullmatch(r'[a-f0-9]{10}', report_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    path = os.path.join(REPORTS_DIR, f'BOM_검증_{report_id}.xlsx')
    if not os.path.exists(path):
        return JSONResponse({'error': '리포트를 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(path, filename='BOM_검증_리포트.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── CCC 게시판 ────────────────────────────────────────────────────────────────
CCC_FILES: dict = {}   # file_id -> (path, original_filename)

@app.get('/ccc', response_class=HTMLResponse)
async def ccc_page(request: Request, vehicle: str = '', stage: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    vcodes = get_all_vehicle_codes()
    uploads = get_ccc_uploads(vehicle or None, stage or None)
    return templates.TemplateResponse(request=request, name='ccc.html',
                                      context={'me': me, 'vcodes': vcodes,
                                               'uploads': uploads,
                                               'sel_vehicle': vehicle,
                                               'sel_stage': stage,
                                               'stages': get_dev_stage_codes()})


@app.post('/ccc/upload')
def ccc_upload(request: Request,
                     vehicle_code: str = Form(...),
                     stage: str = Form(...),
                     revision: str = Form('VER.1'),
                     description: str = Form(''),
                     file: UploadFile = File(...)):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xlsm', '.pptx'):
        return JSONResponse({'error': 'xlsx, xlsm, pptx 파일만 지원합니다.'}, status_code=400)

    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        from ccc_parser import parse_ccc_file
        items = parse_ccc_file(tmp_path, ext.lstrip('.'))

        file_id = uuid.uuid4().hex[:12]
        CCC_FILES[file_id] = (tmp_path, file.filename)

        upload_id = add_ccc_upload(vehicle_code, stage, revision, description,
                                   file.filename, file_id, ext.lstrip('.'), me['username'])
        if items:
            save_ccc_items(upload_id, items)

        return JSONResponse({'ok': True, 'upload_id': upload_id,
                             'item_count': len(items), 'items': items, 'file_id': file_id})
    except Exception as ex:
        import traceback
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        return JSONResponse({'error': str(ex), 'trace': traceback.format_exc()}, status_code=500)


@app.get('/ccc/items/{upload_id}')
async def ccc_items_api(request: Request, upload_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    upload = get_ccc_upload(upload_id)
    if not upload:
        return JSONResponse({'error': '없음'}, status_code=404)
    items = get_ccc_items(upload_id)
    return JSONResponse({'upload': upload, 'items': items})


@app.post('/ccc/delete/{upload_id}')
async def ccc_delete(request: Request, upload_id: int):
    redir = require_login(request)
    if redir: return redir
    delete_ccc_upload(upload_id)
    return RedirectResponse('/ccc', status_code=302)


@app.get('/ccc/download/{file_id}')
async def ccc_download(request: Request, file_id: str):
    redir = require_login(request)
    if redir: return redir
    if not re.fullmatch(r'[a-f0-9]{12}', file_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    entry = CCC_FILES.get(file_id)
    if not entry or not os.path.exists(entry[0]):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    path, name = entry
    return FileResponse(path, filename=name)


# ── 영업 단가 입력 ─────────────────────────────────────────────────────────────
@app.get('/sales/price', response_class=HTMLResponse)
async def sales_price_page(request: Request, vehicle: str = '', stage: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    vcodes = get_all_vehicle_codes()
    return templates.TemplateResponse(request=request, name='sales_price.html',
                                      context={'me': me, 'vcodes': vcodes,
                                               'sel_vehicle': vehicle,
                                               'sel_stage': stage,
                                               'stages': get_dev_stage_codes()})


@app.post('/sales/price/save')
async def sales_price_save(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    errors = []
    saved = 0
    for row in body.get('rows', []):
        try:
            price_val = row.get('unit_price')
            if price_val is None or str(price_val).strip() == '':
                continue
            price_f = float(str(price_val).replace(',', ''))
            upsert_sales_price(
                row['vehicle_code'], row['stage'],
                row['part_no_10'], row['ccc_code'],
                price_f, row.get('currency', 'KRW'),
                row.get('effective_date', ''), me['username']
            )
            saved += 1
        except Exception as ex:
            errors.append(f"{row.get('part_no_13','?')}: {ex}")
    return JSONResponse({'ok': True, 'saved': saved, 'errors': errors})


@app.get('/sales/price/data')
async def sales_price_data(request: Request, vehicle: str = '', stage: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    prices = get_sales_prices(vehicle or None, stage or None)
    ccc_items = get_ccc_items_by_vehicle(vehicle, stage or None) if vehicle else []
    return JSONResponse({'prices': prices, 'ccc_items': ccc_items})


# ── E-BOM 게시판 ──────────────────────────────────────────────────────────────
EBOM_BOARD_DIR = os.path.join(DATA_DIR, 'ebom_board')
os.makedirs(EBOM_BOARD_DIR, exist_ok=True)

@app.get('/ebom-board', response_class=HTMLResponse)
async def ebom_board_page(request: Request, vehicle: str = '', stage: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    vcodes = get_all_vehicle_codes()
    return templates.TemplateResponse(request=request, name='ebom_board.html', context={
        'me': me, 'vcodes': vcodes,
        'sel_vehicle': vehicle, 'sel_stage': stage,
        'stages': get_dev_stage_codes(),
    })


_EBOM_PNO_PAT = re.compile(r'^\d{5}-?[A-Z0-9]{5}$')


def _ebom_norm(x):
    return re.sub(r'\s', '', str(x)) if x is not None else ''


def _detect_pno_col(df):
    """변경후 품번 컬럼 자동 탐지 — 헤더 '변경후' 우선, 없으면 10자리 최다 컬럼(최좌측=변경후)."""
    import pandas as pd
    for ri in range(min(25, df.shape[0])):
        for ci in range(min(30, df.shape[1])):
            if str(df.iat[ri, ci]).strip() == '변경후':
                return ci
    counts = {}
    for ci in range(min(30, df.shape[1])):
        n = sum(1 for ri in range(df.shape[0]) if _EBOM_PNO_PAT.match(_ebom_norm(df.iat[ri, ci])))
        if n >= 2:
            counts[ci] = n
    if not counts:
        return None
    mx = max(counts.values())
    return min(ci for ci, n in counts.items() if n >= mx * 0.6)


def _detect_spec_col(df, pno_col: int):
    """N열(사양) 컬럼 탐지. 사양은 «T&P+A/LEATHER+PWR»처럼 «+»로 이어 붙인 형태라
       «+»를 포함한 텍스트가 압도적으로 많은 열이 사양 열이다(실측: 대상 파일에서
       13열 215건 vs 2위 2건). 품명·재질 열은 «+»가 거의 없어 오탐 위험이 낮다.
       열 위치는 파일마다 달라 고정할 수 없으므로 시트당 한 번만 계산한다."""
    import pandas as pd
    best, best_cnt = None, 0
    for c in range(pno_col + 1, min(df.shape[1], pno_col + 12)):
        cnt = 0
        for ri in range(df.shape[0]):
            v = df.iat[ri, c]
            if pd.notna(v):
                s = str(v)
                if '+' in s and len(s) <= 120:
                    cnt += 1
        if cnt > best_cnt:
            best, best_cnt = c, cnt
    # 몇 건 안 되면 우연일 수 있으니 채택하지 않는다
    return best if best_cnt >= 5 else None


_PNO_RE = re.compile(r'^X?[0-9][0-9A-Z]{4}-[0-9A-Z]{4,6}$')


def _parts_from_sheet_cells(cells: list) -> list:
    """편집된 «셀»에서 품번·품명·레벨을 뽑는다. 원본 파일을 다시 읽으면 편집분이 빠지므로
       DB 의 셀을 본다.

       양식이 두 가지라 열 위치를 «고정하지 않고 탐지»한다(실측):
        · HKMC 양식 — 레벨이 별도 열(3열), 품번 10열, 품명 12열
        · DYA 표준  — 레벨을 들여쓰기로 표현(B~I 중 값이 있는 열 위치가 곧 레벨)
    """
    by_row = {}
    for r, c, v in cells:
        s = str(v or '').strip()
        if s:
            by_row.setdefault(r, {})[c] = s

    # ① 품번 열 — 품번 모양 값이 가장 많은 열
    pcount = {}
    for cols in by_row.values():
        for ci, val in cols.items():
            if _PNO_RE.match(val.replace(' ', '').upper()):
                pcount[ci] = pcount.get(ci, 0) + 1
    if not pcount:
        return []
    pcol = max(pcount, key=pcount.get)

    # ② 레벨은 «한 열»에 있지 않다 — B~I(2~9열)에 들여쓰기로 흩어져 있고, 값이 있는
    #    첫 열의 위치가 곧 레벨이다(실측: 열2=51행·열3=111·열4=131 … 열9=1).
    #    한 열만 고르면 대부분 놓친다(처음에 그렇게 만들었다가 309/427 이 레벨 없음으로 나왔다).
    # ③ 품명 열 — 품번 열 바로 오른쪽에서 «한글/영문 문자열»이 가장 많은 열
    ncount = {}
    for cols in by_row.values():
        if not _PNO_RE.match((cols.get(pcol) or '').replace(' ', '').upper()):
            continue
        for ci in range(pcol + 1, pcol + 5):
            val = cols.get(ci) or ''
            if len(val) >= 4 and not val.replace('.', '').isdigit():
                ncount[ci] = ncount.get(ci, 0) + 1
    ncol = max(ncount, key=ncount.get) if ncount else None

    out, seen = [], set()
    for _r, cols in by_row.items():
        pno = (cols.get(pcol) or '').replace(' ', '').upper()
        if not _PNO_RE.match(pno) or pno in seen:
            continue
        seen.add(pno)
        lv = None
        for ci in range(2, min(10, pcol)):
            val = cols.get(ci)
            if not val:
                continue
            # 칸에 적힌 숫자가 곧 레벨인 양식도 있고, 위치만으로 표현하는 양식도 있다
            lv = int(val) if (val.isdigit() and 0 <= int(val) <= 9) else (ci - 1)
            break
        out.append({'pno': pno, 'part_name': (cols.get(ncol) or '') if ncol else '',
                    'level': lv})
    return out


def _parse_ebom_xlsx(path: str, position: str = '') -> list:
    """
    E-BOM xlsx 파싱 (양식 변동 대응):
      · 모든 시트 스캔 → 위치(운전석/조수석)로 BOM 시트 필터, 이력/사양현황 시트 제외
      · '변경후' 품번 컬럼 자동 탐지 (파일마다 J/L열 등으로 다름)
      · 레벨 = B~I(1~8) 중 첫 non-empty 컬럼 위치 (1레벨 = B열 표시)
      · 여러 변형 시트(LHD/RHD·기본차/환경차)의 품번은 합집합·중복제거
    """
    import pandas as pd
    try:
        xl = pd.ExcelFile(path)
    except Exception:
        return []

    p = position or ''
    if '운전' in p:
        kws = ['운전']
    elif ('조수' in p) or ('동승' in p):
        kws = ['동승', '조수']
    else:
        kws = None
    SKIP = ('EXP', '이력', '사양 현황', '사양현황', '변경 이력', '표지', 'COVER')

    seen = {}
    for sh in xl.sheet_names:
        if kws and not any(k in sh for k in kws):
            continue
        if any(s in sh for s in SKIP):
            continue
        try:
            df = pd.read_excel(path, header=None, sheet_name=sh)
        except Exception:
            continue
        if df.shape[0] < 5 or df.shape[1] < 6:
            continue
        pno_col = _detect_pno_col(df)
        if pno_col is None:
            continue
        spec_col = _detect_spec_col(df, pno_col)
        for ri in range(df.shape[0]):
            raw = df.iat[ri, pno_col]
            if pd.isna(raw):
                continue
            pno = str(raw).strip()
            if not _EBOM_PNO_PAT.match(_ebom_norm(pno)):
                continue
            level = None
            for lc in range(1, 9):
                if lc < df.shape[1]:
                    v = df.iat[ri, lc]
                    if pd.notna(v) and _ebom_norm(v) not in ('', 'nan'):
                        level = lc
                        break
            if level is None:
                continue
            desc = ''
            for dc in range(pno_col + 1, min(pno_col + 5, df.shape[1])):
                dv = df.iat[ri, dc]
                if pd.notna(dv):
                    s = str(dv).strip()
                    if s and not _EBOM_PNO_PAT.match(_ebom_norm(s)):
                        desc = s
                        break
            spec = ''
            if spec_col is not None and spec_col < df.shape[1]:
                sv = df.iat[ri, spec_col]
                if pd.notna(sv):
                    spec = str(sv).strip()
            key = re.sub(r'[\s\-]', '', pno).upper()
            if key not in seen:
                seen[key] = {'level': level, 'pno': pno, 'description': desc,
                             'qty': '', 'variant_code': '', 'spec': spec}
    items = list(seen.values())

    return items


@app.get('/ebom-board/list')
async def ebom_board_list(request: Request, vehicle: str = '', row_num: str = '', position: str = ''):
    """특정 (차종,열,위치)의 리비전 이력 — PEL 이력관리와 동일한 게시판형 조회"""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not (vehicle and row_num and position):
        return JSONResponse({'items': []})
    items = get_ebom_board_revisions(vehicle, row_num, position)
    return JSONResponse({'items': items})


@app.post('/ebom-board/upload')
def ebom_board_upload(
    request: Request,
    vehicle_code: str = Form(...),
    stage: str = Form(''),
    row_num: str = Form(...),
    position: str = Form(...),
    variant: str = Form(''),
    revision: str = Form('VER.1'),
    description: str = Form(''),
    file: UploadFile = File(...),
):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)

    if not row_num.strip() or not position.strip():
        return JSONResponse({'error': '열과 위치를 선택하세요.'}, status_code=400)

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return JSONResponse({'error': 'xlsx/xls 파일만 업로드 가능합니다.'}, status_code=400)

    import uuid
    file_id = uuid.uuid4().hex[:12]
    saved_path = os.path.join(EBOM_BOARD_DIR, f'{file_id}{ext}')
    content = file.file.read()
    with open(saved_path, 'wb') as f:
        f.write(content)

    items = _parse_ebom_xlsx(saved_path, position=position.strip())
    upload_id = add_ebom_upload(vehicle_code, stage, revision, description,
                                file.filename, file_id, me['username'],
                                row_num=row_num.strip(), position=position.strip(),
                                file_path=saved_path, variant=variant.strip())
    if items:
        save_ebom_items(upload_id, items)

    lv1 = sum(1 for it in items if it.get('level') == 1)
    return JSONResponse({
        'ok': True,
        'upload_id': upload_id,
        'items_count': len(items),
        'lv1_count': lv1,
        'uploaded_by': me['username'],
        'message': f'{len(items)}개 품목이 등록되었습니다. (1레벨 {lv1}개)' if items else '파일은 저장되었으나 품번을 찾지 못했습니다. BOM 구조를 확인해주세요.'
    })


@app.get('/ebom-board/items/{upload_id}')
async def ebom_board_items(request: Request, upload_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    items = get_ebom_items(upload_id)
    return JSONResponse({'items': items})


@app.get('/ebom-board/download/{upload_id}')
async def ebom_board_download(request: Request, upload_id: int):
    redir = require_login(request)
    if redir: return redir
    up = get_ebom_upload(upload_id)
    if not up or not up.get('file_path') or not os.path.exists(up['file_path']):
        return JSONResponse({'error': '첨부 파일이 없습니다.'}, status_code=404)
    return FileResponse(up['file_path'], filename=up['filename'])


@app.post('/ebom-board/reparse/{upload_id}')
def ebom_board_reparse(request: Request, upload_id: int):
    """저장된 원본 파일을 새 파서로 다시 파싱 (구버전 파서로 0개였던 파일 복구용)."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    up = get_ebom_upload(upload_id)
    if not up:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and up['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 재파싱할 수 있습니다.'}, status_code=403)
    if not up.get('file_path') or not os.path.exists(up['file_path']):
        return JSONResponse({'error': '원본 파일이 없어 재파싱할 수 없습니다.'}, status_code=404)
    items = _parse_ebom_xlsx(up['file_path'], position=up.get('position', ''))
    replace_ebom_items(upload_id, items)
    lv1 = sum(1 for it in items if it.get('level') == 1)
    return JSONResponse({'ok': True, 'items_count': len(items), 'lv1_count': lv1,
                         'message': f'{len(items)}개 재파싱 (1레벨 {lv1}개)'})


@app.get('/api/ebom/parts')
async def api_ebom_parts(request: Request, vehicle: str = '', stage: str = ''):
    """영업단가 입력/M-BOM비교에서 호출하는 E-BOM 품목 API — 열/위치별 최신 업로드를 합쳐서 반환"""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not vehicle:
        return JSONResponse({'error': '차종을 선택하세요.'}, status_code=400)
    items = get_ebom_items_by_vehicle(vehicle)   # 차종 단독 (단계 무관)
    return JSONResponse({'vehicle': vehicle, 'stage': stage, 'items': items, 'count': len(items)})


@app.post('/ebom-board/delete/{upload_id}')
async def ebom_board_delete(request: Request, upload_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    up = get_ebom_upload(upload_id)
    if not up:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and up['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    delete_ebom_upload(upload_id)
    return JSONResponse({'ok': True})


# ── 국가코드 게시판 ────────────────────────────────────────────────────────────
COUNTRY_PPT_DIR = os.path.join(DATA_DIR, 'country_ppt')
os.makedirs(COUNTRY_PPT_DIR, exist_ok=True)
COUNTRY_PPT_FILES: dict = {}  # rev_id -> path (세션 캐시)

PROCESS_DIAGRAM_DIR = os.path.join(DATA_DIR, 'process_diagrams')
os.makedirs(PROCESS_DIAGRAM_DIR, exist_ok=True)


@app.get('/country-codes', response_class=HTMLResponse)
async def country_codes_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    codes = get_all_country_codes()
    ppt_revs = list_country_ppt_revisions()
    for r in ppt_revs:
        r['file_exists'] = os.path.exists(r.get('file_path', ''))
    return templates.TemplateResponse(request=request, name='country_codes.html',
                                      context={'me': me, 'codes': codes, 'ppt_revs': ppt_revs})


@app.get('/api/country-codes')
async def api_country_codes(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    codes = get_all_country_codes()
    return JSONResponse({'codes': codes})


@app.post('/country-codes/save')
async def country_codes_save(request: Request):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    rows = body.get('rows', [])
    for row in rows:
        code = str(row.get('code', '')).strip().upper()
        if not code:
            continue
        upsert_country_code(
            code,
            str(row.get('region', '')).strip(),
            str(row.get('countries', '')).strip(),
            int(row.get('display_order', 0)),
            code1=str(row.get('code1', '')).strip(),
            code2=str(row.get('code2', '')).strip(),
            hkmc_code=str(row.get('hkmc_code', '')).strip(),
        )
    return JSONResponse({'ok': True, 'saved': len(rows)})


@app.post('/country-codes/delete/{code}')
async def country_codes_delete(request: Request, code: str):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    delete_country_code(code)
    return JSONResponse({'ok': True})


# ── 개발단계 마스터 ────────────────────────────────────────────────────────────
@app.get('/dev-stages', response_class=HTMLResponse)
async def dev_stages_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='dev_stages.html',
                                      context={'me': me, 'stages': get_all_dev_stages()})


@app.get('/api/dev-stages')
async def api_dev_stages(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'stages': get_all_dev_stages()})


@app.post('/dev-stages/save')
async def dev_stages_save(request: Request):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    body = await request.json()
    rows = body.get('rows', [])
    for row in rows:
        code = str(row.get('code', '')).strip()
        if not code:
            continue
        upsert_dev_stage(code, str(row.get('name', '')).strip(), int(row.get('display_order', 0)))
    return JSONResponse({'ok': True, 'saved': len(rows)})


@app.post('/dev-stages/delete/{code}')
async def dev_stages_delete(request: Request, code: str):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    delete_dev_stage(code)
    return JSONResponse({'ok': True})


# ── 파트 네임 정의 ─────────────────────────────────────────────────────────────
@app.get('/part-names', response_class=HTMLResponse)
async def part_names_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='part_names.html',
                                      context={'me': me, 'rows': get_all_part_names(),
                                               'vcodes': get_all_vehicle_codes()})


@app.get('/api/part-names')
async def api_part_names(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'rows': get_all_part_names(vehicle)})


@app.post('/part-names/save')
async def part_names_save(request: Request):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    body = await request.json()
    saved = 0
    for row in body.get('rows', []):
        key = str(row.get('part_key', '')).strip()
        if not key:
            continue
        upsert_part_name(key, str(row.get('part_name', '')).strip(),
                         vehicle_code=str(row.get('vehicle_code', '')).strip(),
                         display_order=int(row.get('display_order', 0)))
        saved += 1
    return JSONResponse({'ok': True, 'saved': saved})


@app.post('/part-names/delete/{row_id}')
async def part_names_delete(request: Request, row_id: int):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    delete_part_name(row_id)
    return JSONResponse({'ok': True})


# ── M-BOM 시스템 ──────────────────────────────────────────────────────────────
MBOM_HISTORY_DIR = os.path.join(DATA_DIR, 'mbom_history')
os.makedirs(MBOM_HISTORY_DIR, exist_ok=True)

MBOM_FILE_SLOTS = ['Q파트 종합', 'FRT LH', 'FRT RH', 'RR BACK LH', 'RR CUSH', 'RR BACK RH']


@app.get('/mbom-history', response_class=HTMLResponse)
async def mbom_history_page(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='mbom_history.html', context={
        'alc2_masters': _alc2_all_info(), 'alc2_kinds': ALC2_MASTER_KINDS,
        'me': me, 'vcodes': get_all_vehicle_codes(), 'sel_vehicle': vehicle,
        'stages': get_dev_stage_codes(), 'slots': MBOM_FILE_SLOTS,
    })


@app.get('/mbom-history/list')
async def mbom_history_list(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not vehicle:
        return JSONResponse({'items': []})
    return JSONResponse({'items': get_mbom_history(vehicle)})


@app.post('/mbom-history/upload')
async def mbom_history_upload(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    form = await request.form()
    vehicle_code = str(form.get('vehicle_code', '')).strip().upper()
    title = str(form.get('title', '')).strip()
    if not vehicle_code or not title:
        return JSONResponse({'error': '차종과 제목은 필수입니다.'}, status_code=400)
    post_id = add_mbom_history(
        vehicle_code, str(form.get('stage', '')).strip(),
        str(form.get('revision', 'VER.1')).strip() or 'VER.1',
        title, str(form.get('description', '')).strip(), me['username'])
    saved_files = 0
    for i, slot in enumerate(MBOM_FILE_SLOTS):
        f = form.get(f'file{i}')
        if f is None or not getattr(f, 'filename', ''):
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        fid = uuid.uuid4().hex[:16]
        path = os.path.join(MBOM_HISTORY_DIR, f'{fid}{ext}')
        with open(path, 'wb') as out:
            shutil.copyfileobj(f.file, out)
        add_mbom_file(post_id, slot, f.filename, fid, path)
        saved_files += 1
    return JSONResponse({'ok': True, 'id': post_id, 'files': saved_files,
                         'uploaded_by': me['username']})


@app.get('/mbom-history/download/{file_row_id}')
async def mbom_history_download(request: Request, file_row_id: int):
    redir = require_login(request)
    if redir: return redir
    f = get_mbom_file(file_row_id)
    if not f or not os.path.exists(f['file_path']):
        return JSONResponse({'error': '파일이 없습니다.'}, status_code=404)
    return FileResponse(f['file_path'], filename=f['filename'])


@app.post('/mbom-history/delete/{post_id}')
async def mbom_history_delete(request: Request, post_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    # 삭제 전 권한 확인 (본인 또는 관리자)
    posts_all = get_mbom_history_post(post_id)
    if not posts_all:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and posts_all['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    result = delete_mbom_history(post_id)
    for p in result.get('paths', []):
        try:
            if os.path.exists(p): os.unlink(p)
        except Exception:
            pass
    return JSONResponse({'ok': True})


# ── Q파트 ALC 통합 게시판 (mbom_history와 별개 신설 게시판) ─────────────────────
QPART_MERGE_DIR = os.path.join(DATA_DIR, 'qpart_merge')
os.makedirs(QPART_MERGE_DIR, exist_ok=True)
QPART_MERGE_RESULT_DIR = os.path.join(REPORTS_DIR, 'qpart_merge')
os.makedirs(QPART_MERGE_RESULT_DIR, exist_ok=True)


@app.get('/qpart-merge', response_class=HTMLResponse)
async def qpart_merge_page(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    import alc2_convert
    return templates.TemplateResponse(request=request, name='qpart_merge.html', context={
        'me': me, 'vcodes': get_all_vehicle_codes(), 'sel_vehicle': vehicle,
        'stages': get_dev_stage_codes(), 'slots': ['Q파트 종합'] + alc2_convert.ALC_SLOTS,
    })


@app.get('/qpart-merge/list')
async def qpart_merge_list(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not vehicle:
        return JSONResponse({'items': []})
    return JSONResponse({'items': get_qpart_merge_history(vehicle)})


@app.post('/qpart-merge/upload')
async def qpart_merge_upload(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    import alc2_convert
    slots = ['Q파트 종합'] + alc2_convert.ALC_SLOTS
    form = await request.form()
    vehicle = str(form.get('vehicle', '')).strip().upper()
    title = str(form.get('title', '')).strip()
    if not vehicle:
        return JSONResponse({'error': '차종은 필수입니다.'}, status_code=400)
    post_id = add_qpart_merge_post(vehicle, str(form.get('dev_stage', '')).strip(), title, me['username'])
    saved_files = 0
    for i, slot in enumerate(slots):
        f = form.get(f'file{i}')
        if f is None or not getattr(f, 'filename', ''):
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        fid = uuid.uuid4().hex[:16]
        path = os.path.join(QPART_MERGE_DIR, f'{fid}{ext}')
        with open(path, 'wb') as out:
            shutil.copyfileobj(f.file, out)
        add_qpart_merge_file(post_id, slot, f.filename, path)
        saved_files += 1
    return JSONResponse({'ok': True, 'id': post_id, 'files': saved_files, 'uploaded_by': me['username']})


@app.get('/qpart-merge/download/{file_row_id}')
async def qpart_merge_download(request: Request, file_row_id: int):
    redir = require_login(request)
    if redir: return redir
    f = get_qpart_merge_file(file_row_id)
    if not f or not os.path.exists(f['file_path']):
        return JSONResponse({'error': '파일이 없습니다.'}, status_code=404)
    return FileResponse(f['file_path'], filename=f['filename'])


@app.post('/qpart-merge/delete/{post_id}')
async def qpart_merge_delete(request: Request, post_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    post = get_qpart_merge_post(post_id)
    if not post:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and post['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    result = delete_qpart_merge_post(post_id)
    for p in result.get('paths', []):
        try:
            if p and os.path.exists(p): os.unlink(p)
        except Exception:
            pass
    return JSONResponse({'ok': True})


@app.post('/qpart-merge/run/{post_id}')
def qpart_merge_run(request: Request, post_id: int):
    if require_login(request):
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    import alc2_convert
    files = get_qpart_merge_files_by_post(post_id)
    by_slot = {f['slot']: f['file_path'] for f in files if f.get('file_path') and os.path.exists(f['file_path'])}
    qpart = by_slot.get('Q파트 종합')
    if not qpart:
        return JSONResponse({'error': "'Q파트 종합' 파일이 이 게시글에 없습니다."}, status_code=400)
    alc_paths = {s: by_slot.get(s) for s in alc2_convert.ALC_SLOTS}
    missing_slots = [s for s in alc2_convert.ALC_SLOTS if not alc_paths.get(s)]
    from bom_generator import load_pel_master
    mpel = load_pel_master(PEL_CODE_PATH).get('data', {})
    rid = uuid.uuid4().hex[:10]
    out_path = os.path.join(QPART_MERGE_RESULT_DIR, f'QPARTMERGE_{rid}.xlsx')
    try:
        result = alc2_convert.build_qpart_merge(qpart, alc_paths, mpel, out_path)
    except Exception as ex:
        return JSONResponse({'error': f'변환 오류: {ex}'}, status_code=500)
    post = get_qpart_merge_post(post_id)
    out_filename = f"{post['vehicle']}_Q파트ALC통합_{rid}.xlsx" if post else f'QPARTMERGE_{rid}.xlsx'
    run_id = add_qpart_merge_run(post_id, out_path, out_filename,
                                 result['spec_col_count'], result['row_count'], me['username'])
    try:
        grid = alc2_convert.read_grid(out_path)
    except Exception as ex:
        return JSONResponse({'error': f'그리드 변환 오류: {ex}'}, status_code=500)
    return JSONResponse({'ok': True, 'run_id': run_id, 'missing_slots': missing_slots,
                         'row_count': result['row_count'], 'spec_col_count': result['spec_col_count'],
                         'grid': grid})


@app.get('/qpart-merge/runs/{post_id}')
async def qpart_merge_runs_list(request: Request, post_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'items': get_qpart_merge_runs(post_id)})


@app.get('/qpart-merge/grid/{run_id}')
def qpart_merge_grid(request: Request, run_id: int):
    if require_login(request):
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    import alc2_convert
    run = get_qpart_merge_run(run_id)
    if not run or not os.path.exists(run['output_path']):
        return JSONResponse({'error': '결과 파일을 찾을 수 없습니다.'}, status_code=404)
    try:
        grid = alc2_convert.read_grid(run['output_path'])
    except Exception as ex:
        return JSONResponse({'error': f'그리드 변환 오류: {ex}'}, status_code=500)
    return JSONResponse({'ok': True, 'run': run, 'grid': grid})


@app.get('/qpart-merge/download-result/{run_id}')
async def qpart_merge_download_result(request: Request, run_id: int):
    redir = require_login(request)
    if redir: return redir
    run = get_qpart_merge_run(run_id)
    if not run or not os.path.exists(run['output_path']):
        return JSONResponse({'error': '결과 파일을 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(run['output_path'], filename=run['output_filename'] or '병합결과.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 통합 ALC2 마스터 (마스터 데이터 저장 · ALC-2 채번 기준) ─────────────────────
ALC2_MASTER_DIR = os.path.join(DATA_DIR, 'alc2_master')
os.makedirs(ALC2_MASTER_DIR, exist_ok=True)
# 3종 마스터: 표준(매칭·채번) / 열별사양 / 최종 산출물 서식(★REV)
ALC2_MASTER_KINDS = {
    'standard': '표준 마스터 (MES OX표기용 · 전체열 통합)',
    'spec': '열별 사양 마스터 (ALC코드 집 생성용)',
    'format': '최종 산출물 서식 (★통합 ALC2 코드 REV)',
}
ALC2_MASTER_PATH = os.path.join(ALC2_MASTER_DIR, 'alc2_standard.xlsx')  # 매칭용(하위호환)


def _alc2_path(kind):
    return os.path.join(ALC2_MASTER_DIR, f'alc2_{kind}.xlsx')


def _alc2_meta(kind):
    return os.path.join(ALC2_MASTER_DIR, f'meta_{kind}.json')


def _alc2_master_info(kind='standard'):
    p = _alc2_path(kind)
    if not os.path.exists(p):
        return {'exists': False, 'kind': kind, 'label': ALC2_MASTER_KINDS.get(kind, kind)}
    meta = {}
    try:
        meta = json.load(open(_alc2_meta(kind), encoding='utf-8'))
    except Exception:
        pass
    return {'exists': True, 'kind': kind, 'label': ALC2_MASTER_KINDS.get(kind, kind),
            'filename': meta.get('filename', os.path.basename(p)),
            'uploaded_by': meta.get('uploaded_by', ''), 'uploaded': meta.get('uploaded', '')}


def _alc2_all_info():
    return {k: _alc2_master_info(k) for k in ALC2_MASTER_KINDS}


@app.get('/alc2-master/info')
async def alc2_master_info(request: Request):
    if require_login(request):
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'masters': _alc2_all_info()})


@app.post('/alc2-master/upload/{kind}')
async def alc2_master_upload(request: Request, kind: str, file: UploadFile = File(...)):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    if kind not in ALC2_MASTER_KINDS:
        return JSONResponse({'error': '잘못된 마스터 종류'}, status_code=400)
    me = current_user(request)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return JSONResponse({'error': 'xlsx/xls만 가능합니다.'}, status_code=400)
    with open(_alc2_path(kind), 'wb') as f:
        shutil.copyfileobj(file.file, f)
    import datetime
    json.dump({'filename': file.filename, 'uploaded_by': me['username'],
               'uploaded': datetime.datetime.now().strftime('%Y-%m-%d %H:%M')},
              open(_alc2_meta(kind), 'w', encoding='utf-8'), ensure_ascii=False)
    return JSONResponse({'ok': True, **_alc2_master_info(kind)})


@app.get('/alc2-master/download/{kind}')
async def alc2_master_download(request: Request, kind: str):
    if require_login(request):
        return RedirectResponse('/login')
    if kind not in ALC2_MASTER_KINDS or not os.path.exists(_alc2_path(kind)):
        return JSONResponse({'error': '등록된 마스터가 없습니다.'}, status_code=404)
    info = _alc2_master_info(kind)
    return FileResponse(_alc2_path(kind), filename=info.get('filename', f'alc2_{kind}.xlsx'))


ALC2_LEDGER_CACHE = {}   # path -> (mtime, size, cols, hdr_row)


def _alc2_ledger_cols(path):
    """대장의 열 위치를 헤더명으로 탐색 (mtime 캐시 — 매번 2.4초 걸리므로)."""
    import alc2_ledger
    try:
        st = os.stat(path); ck = (st.st_mtime, st.st_size)
    except OSError:
        ck = None
    hit = ALC2_LEDGER_CACHE.get(path)
    if hit and hit[0] == ck:
        return hit[1], hit[2]
    cols, last_no = alc2_ledger.find_columns(path)
    if ck:
        ALC2_LEDGER_CACHE[path] = (ck, cols, last_no)
    return cols, last_no


def _alc2_write_ledger(src, dst, rows, option_marks=None, meta_values=None, meta_cols=None):
    """★REV 서식(헤더 색상·열 구조)만 물려받고, 데이터 영역은 이번 변환 결과로 교체한다.
       기존 대장 이력을 그대로 두면 이전 파일과 비교가 안 되므로 8행부터 새로 채운다.
       openpyxl 왕복(15초) 대신 zip+XML 직접 조작(0.6초).
       option_marks: {kmc20: set(col_letter)} — 서식에 이미 있는 ERGO/LUMBAR SUPPORT/THORAX...
       고정 옵션 열에 O를 채운다 (alc2_convert.build_option_marks 결과).
       meta_values/meta_cols: F(DRV TYPE)/H(사양지)/K~Y(좌석별 품번·원단코드·KMC코드)
       (alc2_convert.build_meta_values / alc2_ledger.find_meta_columns 결과)."""
    import alc2_ledger
    cols, first_row = _alc2_ledger_cols(src)
    if 'kmc' not in cols:
        shutil.copy2(src, dst)
        return 0
    option_marks = option_marks or {}
    meta_values = meta_values or {}
    meta_cols = meta_cols or {}
    seat_cols = meta_cols.get('seats', {})
    vals = []
    for i, r in enumerate(rows, 1):
        kmc20 = r.get('kmc20', '')
        v = {cols['kmc']: kmc20}
        if 'no' in cols:
            v[cols['no']] = i
        if 'vehicle' in cols:
            v[cols['vehicle']] = r.get('vehicle', '')
        if 'alc2' in cols:
            v[cols['alc2']] = r.get('alc2', '')
        for col in option_marks.get(kmc20, ()):
            v[col] = 'O'
        mv = meta_values.get(kmc20)
        if mv:
            if meta_cols.get('dt') and mv.get('dt'):
                v[meta_cols['dt']] = mv['dt']
            if meta_cols.get('region') and mv.get('region'):
                v[meta_cols['region']] = mv['region']
            for top, sv in mv.get('seats', {}).items():
                sc = seat_cols.get(top)
                if not sc:
                    continue
                if sv.get('partno'):
                    v[sc['partno']] = sv['partno']
                if sv.get('fabric'):
                    v[sc['fabric']] = sv['fabric']
                if sv.get('kmc'):
                    v[sc['kmc']] = sv['kmc']
        vals.append(v)
    return alc2_ledger.replace_rows(src, dst, vals, first_row)


# ── mbom-history 게시글에서 ALC-2 생성 실행 ───────────────────────────────────
@app.post('/mbom-history/alc2-run/{post_id}')
def mbom_history_alc2_run(request: Request, post_id: int):
    if require_login(request):
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    import alc2_convert
    # 매칭·채번 원본 = ★최종 산출물 서식(통합 ALC2 코드 대장) 우선, 없으면 표준 마스터
    master_path = _alc2_path('format') if os.path.exists(_alc2_path('format')) else ALC2_MASTER_PATH
    if not os.path.exists(master_path):
        return JSONResponse({'error': '먼저 이 화면 상단 [기준 마스터 3종]에서 «★최종 산출물 서식» 또는 «표준 마스터»를 등록하세요.'}, status_code=400)
    files = get_mbom_files_by_post(post_id)
    by_slot = {f['slot']: f['file_path'] for f in files if f.get('file_path') and os.path.exists(f['file_path'])}
    qpart = by_slot.get('Q파트 종합')
    if not qpart:
        return JSONResponse({'error': "'Q파트 종합' 파일이 이 게시글에 없습니다."}, status_code=400)
    alc_paths = {s: by_slot.get(s) for s in alc2_convert.ALC_SLOTS}
    missing_slots = [s for s in alc2_convert.ALC_SLOTS if not alc_paths.get(s)]
    from bom_generator import load_pel_master
    mpel = load_pel_master(PEL_CODE_PATH).get('data', {})
    try:
        full = alc2_convert.analyze(qpart, alc_paths, master_path, mpel)  # 6파일 1회씩만 로드
    except Exception as ex:
        return JSONResponse({'error': f'변환 오류: {ex}'}, status_code=500)
    res = {'rows': full['rows'], 'stats': full['stats']}
    _ox = full['ox']
    dt_warnings = []
    hkmc_map = {}
    try:
        hkmc_map = {c['hkmc_code']: c for c in get_all_country_codes() if c.get('hkmc_code')}
        dt_warnings = alc2_convert.check_frt_dt(qpart, alc_paths, hkmc_map)['warnings']
    except Exception as ex:
        dt_warnings = [f'전석 DT/국가코드 검증 중 오류: {ex}']
    rid = uuid.uuid4().hex[:10]
    # ① ★통합 ALC2 코드 대장 — 원본 서식 그대로 복사 + 신규 코드만 이어붙임
    fmt_path = _alc2_path('format')
    tpl_used, ledger_added = '', 0
    if os.path.exists(fmt_path):
        try:
            import alc2_ledger
            option_cols = alc2_ledger.find_option_columns(fmt_path)
            option_marks = alc2_convert.build_option_marks(qpart, alc_paths, mpel, option_cols)
            meta_cols = alc2_ledger.find_meta_columns(fmt_path)
            meta_values = alc2_convert.build_meta_values(qpart, alc_paths, hkmc_map)
            lout = os.path.join(REPORTS_DIR, f'ALC2LEDGER_{rid}.xlsx')
            ledger_added = _alc2_write_ledger(fmt_path, lout, res['rows'], option_marks,
                                              meta_values, meta_cols)
            tpl_used = _alc2_master_info('format').get('filename', '★통합 ALC2 코드')
            ALC2_LEDGERS[rid] = lout
        except Exception:
            tpl_used, ledger_added = '', 0
    # ② 판정결과 · O·X 통합코드집 리포트
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = 'ALC2 판정결과'
    hdr = ['NO', '차종', '국가(KEY01)', 'KMC ALC-2 CODE', 'DYA ALC-2', '판정', '상세']
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', start_color='1A237E')
    yellow = PatternFill('solid', start_color='FFF9C4'); red = PatternFill('solid', start_color='FFCDD2')
    for i, r in enumerate(res['rows'], 1):
        ws.append([i, r['vehicle'], r.get('key01', ''), r['kmc20'], r['alc2'], r['status'], r['detail']])
        if r['status'] == '신규승인필요':
            for c in ws[i + 1]: c.fill = yellow
        elif r['status'] == '원본누락':
            for c in ws[i + 1]: c.fill = red
    # O/X 통합코드집 시트 (PEL CODE 마스터 기반) — 6파일만으로 생성
    ox_cols = 0
    try:
        from openpyxl.styles import Alignment
        ox = _ox
        ox_cols = len(ox['columns'])
        a2 = {r['kmc20']: r['alc2'] for r in res['rows']}
        ws2 = wb.create_sheet('O·X 통합코드집')
        center = Alignment(horizontal='center', vertical='center')
        fx = ['NO', '차종', '국가', 'KMC ALC-2', 'DYA ALC-2']
        ws2.append(fx + sum([[g['group']] + [''] * (g['span'] - 1) for g in ox['groups']], []))
        ws2.append([''] * len(fx) + [c['spec'] for c in ox['columns']])
        gfill = PatternFill('solid', start_color='283593'); hfill = PatternFill('solid', start_color='1A237E')
        hfont = Font(bold=True, color='FFFFFF')
        for cc in ws2[1]: cc.fill = gfill; cc.font = hfont; cc.alignment = center
        for cc in ws2[2]: cc.fill = hfill; cc.font = hfont; cc.alignment = center
        cidx = len(fx) + 1
        for g in ox['groups']:
            if g['span'] > 1:
                ws2.merge_cells(start_row=1, start_column=cidx, end_row=1, end_column=cidx + g['span'] - 1)
            cidx += g['span']
        for k in range(1, len(fx) + 1):
            ws2.merge_cells(start_row=1, start_column=k, end_row=2, end_column=k)
            ws2.cell(1, k).value = fx[k - 1]; ws2.cell(1, k).fill = hfill
        for i, r in enumerate(ox['rows'], 1):
            mk = set(r['marks'])
            ws2.append([i, r['vehicle'], r.get('key01', ''), r['kmc20'], a2.get(r['kmc20'], '')]
                       + ['O' if c['spec'] in mk else '' for c in ox['columns']])
            for cc in range(len(fx) + 1, len(fx) + 1 + len(ox['columns'])):
                ws2.cell(i + 2, cc).alignment = center
    except Exception:
        pass
    out = os.path.join(REPORTS_DIR, f'ALC2RES_{rid}.xlsx')
    wb.save(out); ALC2_RESULTS[rid] = out
    from datetime import datetime
    day = datetime.now().strftime('%Y%m%d')
    ALC2_RESULT_NAMES[rid] = 'ALC2_판정결과·OX통합코드집_%s.xlsx' % day
    ALC2_LEDGER_NAMES[rid] = '★통합 ALC2 코드_%s_REV(변환 %d건).xlsx' % (day, ledger_added)
    return JSONResponse({'ok': True, 'result_id': rid, 'stats': res['stats'], 'template': tpl_used,
                         'ledger_added': ledger_added, 'has_ledger': bool(tpl_used),
                         'unknown_pel': full.get('unknown_pel', []), 'dt_warnings': dt_warnings,
                         'ox_cols': ox_cols, 'missing_slots': missing_slots, 'rows': res['rows'][:200]})


@app.get('/mbom-history/alc2-result/{result_id}')
async def mbom_history_alc2_result(request: Request, result_id: str):
    if require_login(request):
        return RedirectResponse('/login')
    if not re.fullmatch(r'[a-f0-9]{10}', result_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    path = ALC2_RESULTS.get(result_id)
    if not path or not os.path.exists(path):
        return JSONResponse({'error': '결과 만료 — 다시 실행하세요.'}, status_code=404)
    return FileResponse(path, filename=ALC2_RESULT_NAMES.get(result_id, 'DYA_ALC2_판정결과.xlsx'),
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.get('/mbom-history/alc2-ledger/{result_id}')
async def mbom_history_alc2_ledger(request: Request, result_id: str):
    """★통합 ALC2 코드 대장 (원본 서식 + 신규코드 반영)."""
    if require_login(request):
        return RedirectResponse('/login')
    if not re.fullmatch(r'[a-f0-9]{10}', result_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    path = ALC2_LEDGERS.get(result_id)
    if not path or not os.path.exists(path):
        return JSONResponse({'error': '결과 만료 — 다시 실행하세요.'}, status_code=404)
    return FileResponse(path, filename=ALC2_LEDGER_NAMES.get(result_id, '★통합 ALC2 코드_REV.xlsx'),
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 원단코드 마스터 (마스터 데이터) ────────────────────────────────────────────
@app.get('/fabric-master', response_class=HTMLResponse)
async def fabric_master_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='fabric_master.html', context={
        'me': me, 'fabrics': get_all_fabric_codes(),
    })


# ── 원단코드 마스터 API ────────────────────────────────────────────────────────
@app.get('/api/fabric-codes')
async def api_fabric_codes(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'fabrics': get_all_fabric_codes()})


@app.post('/fabric-codes/save')
async def fabric_codes_save(request: Request):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    body = await request.json()
    n = 0
    for row in body.get('rows', []):
        code = str(row.get('code', '')).strip()
        if not code:
            continue
        upsert_fabric_code(code, str(row.get('fabric_code', '')).strip(),
                           str(row.get('name', '')).strip(), str(row.get('stitch_color', '')).strip(),
                           str(row.get('base_color', '')).strip(), int(row.get('display_order', 0)),
                           hkmc_code=str(row.get('hkmc_code', '')).strip())
        n += 1
    return JSONResponse({'ok': True, 'saved': n})


@app.post('/fabric-codes/delete/{code}')
async def fabric_codes_delete(request: Request, code: str):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    delete_fabric_code(code)
    return JSONResponse({'ok': True})


# ── DYA ALC-2 코드 생성 ────────────────────────────────────────────────────────

def _parse_qpart_xlsx(path: str) -> list:
    """GY 종합(Q파트) 파싱 — 차종|구분|KEY01~KEY06 (KEY02~06 4자리 코드 행만)"""
    import pandas as pd
    df = pd.read_excel(path, header=None, sheet_name=0)
    rows = []
    for i in range(1, len(df)):
        r = df.iloc[i]
        def s(c):
            v = r[c] if c < len(r) else None
            return str(v).strip() if pd.notna(v) else ''
        keys = [s(c) for c in range(3, 8)]   # KEY02~KEY06
        if not all(keys) or any(len(k) != 4 for k in keys):
            continue
        rows.append({
            'vehicle': s(0), 'gubun': s(1), 'key01': s(2),
            'keys': keys, 'kmc20': ''.join(keys),
        })
    del df
    return rows


def _parse_alc2_master_xlsx(path: str) -> dict:
    """통합 ALC2 코드 마스터 파싱 — {KMC 20자리: {'alc2':5자리, 'vehicle':차종}}"""
    import pandas as pd
    df = pd.read_excel(path, sheet_name=0, header=None)
    m = {}
    for i in range(len(df)):
        try:
            a2 = str(df.iat[i, 2]).strip()
            k = str(df.iat[i, 3]).strip()
            veh = str(df.iat[i, 1]).strip()
        except Exception:
            continue
        if len(k) == 20 and k.isalnum() and len(a2) == 5:
            m[k] = {'alc2': a2, 'vehicle': veh}
    del df
    return m


@app.on_event('startup')
def _alc2_warmup():
    """대장(3천행×283열) 캐시를 기동 직후 백그라운드로 미리 적재.
       안 하면 재시작 후 첫 변환이 30초대가 된다."""
    import threading

    def run():
        try:
            p = _alc2_path('format')
            if not os.path.exists(p):
                p = ALC2_MASTER_PATH
            if os.path.exists(p):
                import alc2_convert
                alc2_convert.read_master(p)
                _alc2_ledger_cols(p)
        except Exception:
            pass

    threading.Thread(target=run, daemon=True).start()


ALC2_RESULTS: dict = {}        # result_id -> 판정·O·X 리포트 path
ALC2_RESULT_NAMES: dict = {}   # result_id -> 리포트 다운로드 파일명
ALC2_LEDGERS: dict = {}        # result_id -> ★통합 ALC2 코드 대장 path
ALC2_LEDGER_NAMES: dict = {}   # result_id -> 대장 다운로드 파일명


@app.post('/mbom-alc2-gen/run')
def mbom_alc2_run(request: Request,
                        qpart: UploadFile = File(...),
                        master: UploadFile = File(...)):
    """Q파트 종합 + 기존 통합 ALC2 마스터 → 행별 DYA ALC-2 매칭/신규 판정"""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)

    import tempfile
    tmp_q = tmp_m = None
    try:
        # 순차 처리 (서버 메모리 보호): 파일 하나씩 저장→파싱→해제
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as t:
            shutil.copyfileobj(qpart.file, t); tmp_q = t.name
        q_rows = _parse_qpart_xlsx(tmp_q)
        os.unlink(tmp_q); tmp_q = None

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as t:
            shutil.copyfileobj(master.file, t); tmp_m = t.name
        master_map = _parse_alc2_master_xlsx(tmp_m)
        os.unlink(tmp_m); tmp_m = None

        if not q_rows:
            return JSONResponse({'error': 'Q파트 파일에서 KEY02~KEY06(4자리×5) 행을 찾지 못했습니다.'}, status_code=400)
        if not master_map:
            return JSONResponse({'error': '통합 ALC2 마스터에서 코드(5자리+20자리) 행을 찾지 못했습니다.'}, status_code=400)

        # 국가코드 해석 (HKMC 코드 → 지역/1순위)
        cmap = {}
        for c in get_all_country_codes():
            hk = (c.get('hkmc_code') or '').strip().upper()
            if hk:
                cmap[hk] = {'region': c.get('region', ''), 'code1': (c.get('code1') or '').strip().upper(),
                            'kcode': c.get('code', '')}

        # 신규 채번 준비: 마스터에 이미 쓰인 ALC2 수집 → (지역+차종)별 다음 순번
        used = set(v['alc2'] for v in master_map.values())

        results, matched, newcnt = [], 0, 0
        seen_new = {}
        for row in q_rows:
            k20 = row['kmc20']
            hit = master_map.get(k20)
            key01 = row['key01'].upper()
            cinfo = cmap.get(key01, {})
            veh_letter = row['keys'][0][-1] if row['keys'][0] else '?'
            if hit:
                matched += 1
                results.append({**row, 'alc2': hit['alc2'], 'status': 'OK',
                                'region': cinfo.get('region', ''), 'kcode': cinfo.get('kcode', '')})
            else:
                # 신규 — 제안 채번: [지역1순위][원단?][순번2][차종문자] (원단코드는 사양 확인 후 확정)
                if k20 in seen_new:
                    prop = seen_new[k20]
                else:
                    r1 = cinfo.get('code1') or '?'
                    nn = 1
                    while True:
                        cand = f"{r1}?{nn:02d}{veh_letter}"
                        if cand not in used:
                            used.add(cand); break
                        nn += 1
                    prop = cand
                    seen_new[k20] = prop
                    newcnt += 1
                results.append({**row, 'alc2': prop, 'status': 'NEW',
                                'region': cinfo.get('region', ''), 'kcode': cinfo.get('kcode', '')})

        # 결과 엑셀 생성
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook(); ws = wb.active; ws.title = 'ALC2 생성 결과'
        hdr = ['NO', '차종', '구분', 'KEY01', '국가', 'K코드',
               'KEY02', 'KEY03', 'KEY04', 'KEY05', 'KEY06',
               'KMC ALC-2 (20자리)', 'DYA ALC-2', '판정']
        ws.append(hdr)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='1A237E')
        new_fill = PatternFill('solid', start_color='FFF9C4')
        for i, r in enumerate(results, 1):
            ws.append([i, r['vehicle'], r['gubun'], r['key01'], r['region'], r['kcode'],
                       *r['keys'], r['kmc20'], r['alc2'],
                       '기존 매칭' if r['status'] == 'OK' else '신규 채번 필요(원단? 확인)'])
            if r['status'] == 'NEW':
                for c in ws[i + 1]:
                    c.fill = new_fill
        result_id = uuid.uuid4().hex[:10]
        out_path = os.path.join(REPORTS_DIR, f'ALC2_{result_id}.xlsx')
        wb.save(out_path)
        ALC2_RESULTS[result_id] = out_path

        return JSONResponse({
            'ok': True, 'result_id': result_id,
            'total': len(results), 'matched': matched, 'new_codes': newcnt,
            'rows': results[:300],
            'truncated': len(results) > 300,
        })
    except Exception as ex:
        import traceback
        return JSONResponse({'error': f'처리 오류: {ex}', 'trace': traceback.format_exc()}, status_code=500)
    finally:
        for p in (tmp_q, tmp_m):
            if p and os.path.exists(p):
                try: os.unlink(p)
                except Exception: pass


@app.get('/mbom-alc2-gen/download/{result_id}')
async def mbom_alc2_download(request: Request, result_id: str):
    redir = require_login(request)
    if redir: return redir
    if not re.fullmatch(r'[a-f0-9]{10}', result_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    path = ALC2_RESULTS.get(result_id)
    if not path or not os.path.exists(path):
        return JSONResponse({'error': '결과가 만료되었습니다. 다시 실행해주세요.'}, status_code=404)
    return FileResponse(path, filename='DYA_ALC2_생성결과.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 설계변경 통보서(EO) 게시판 ────────────────────────────────────────────────
EO_FILE_DIR = os.path.join(DATA_DIR, 'eo_files')
os.makedirs(EO_FILE_DIR, exist_ok=True)


@app.get('/eo', response_class=HTMLResponse)
async def eo_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='eo.html', context={
        'me': me, 'stats': get_eo_stats(), 'fields': EO_FIELDS,
    })


@app.get('/eo/list')
async def eo_list(request: Request, q: str = '', vehicle: str = '', status: str = '',
                  date_from: str = '', date_to: str = '', limit: int = 500, offset: int = 0):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    res = search_eo_notices(q.strip(), vehicle.strip(), status.strip(),
                            date_from.strip(), date_to.strip(), min(limit, 2000), offset)
    res['stats'] = get_eo_stats()
    return JSONResponse(res)


@app.get('/eo/detail/{eo_id}')
async def eo_detail(request: Request, eo_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    e = get_eo_notice(eo_id)
    if not e:
        return JSONResponse({'error': '통보서를 찾을 수 없습니다.'}, status_code=404)
    me = current_user(request)
    approvals = get_eo_approvals(eo_id)
    # 내 차례인지(= 가장 앞선 pending 단계의 결재자인지) 서버가 판단해 버튼 노출을 맞춘다.
    pending = next((a for a in approvals if a['status'] == 'pending'), None)
    my_turn = bool(pending) and (
        me['username'] in str(pending['approver']) or
        (me.get('name') or '\x00') in str(pending['approver']))
    return JSONResponse({'ok': True, 'eo': e, 'files': get_eo_files(eo_id),
                         'items': get_eo_items(eo_id),
                         'approvals': approvals, 'mails': get_eo_mails(eo_id),
                         'my_turn': my_turn, 'is_admin': me['role'] == 'admin',
                         'status_labels': EO_APPROVAL_STATUS,
                         'reasons': EO_REASONS, 'roles': EO_APPROVAL_ROLES})


@app.post('/eo/save')
async def eo_save(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    fields = {k: v for k, v in (body.get('fields') or {}).items() if k in EO_FIELDS}
    eo_id = body.get('id')
    if eo_id:
        # 결재 진행·완료 문서는 내용을 바꿀 수 없다(화면 비활성만으로는 못 막는다).
        cur = get_eo_notice(int(eo_id)) or {}
        st = cur.get('approval_status') or 'draft'
        if st in ('submitted', 'in_progress', 'approved'):
            return JSONResponse({'error': f'«{EO_APPROVAL_STATUS.get(st, st)}» 상태에서는 수정할 수 없습니다. '
                                          '반려 후 재상신하거나 관리자에게 문의하세요.'}, status_code=400)
        res = update_eo_notice(int(eo_id), fields)
        return JSONResponse({'ok': True, 'id': int(eo_id), **res})
    eo_no = str(fields.get('eo_no', '')).strip()
    if not eo_no:
        return JSONResponse({'error': 'EO 번호는 필수입니다.'}, status_code=400)
    # upsert 는 «일괄 등록이 상세를 지우지 않도록» 목록 항목만 다룬다. 화면에서 새로
    # 만들 때는 상세까지 저장해야 하므로, 생성 후 전체 필드로 한 번 더 갱신한다.
    res = upsert_eo_notices([fields], me['username'])
    row = search_eo_notices(q=eo_no, limit=1)
    new_id = row['items'][0]['id'] if row['items'] else None
    if new_id:
        update_eo_notice(new_id, fields)
    return JSONResponse({'ok': True, 'id': new_id, **res})


@app.post('/eo/import')
def eo_import(request: Request, file: UploadFile = File(...)):
    """설계변경 통보서 목록 엑셀 일괄 등록. 헤더 이름으로 열을 자동 인식한다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    import pandas as pd
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls', '.xlsm'):
        return JSONResponse({'error': 'xlsx/xlsm/xls 파일만 지원합니다.'}, status_code=400)
    tmp = os.path.join(EO_FILE_DIR, f'_imp_{uuid.uuid4().hex[:10]}{ext}')
    with open(tmp, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    try:
        raw = pd.read_excel(tmp, header=None, sheet_name=0)
        # 헤더 행 탐색 — «EO 번호» 계열 문구가 있는 행
        hdr_row, colmap = None, {}
        for ri in range(min(12, len(raw))):
            vals = ['' if pd.isna(v) else str(v).strip() for v in raw.iloc[ri].tolist()]
            found = {}
            for ci, v in enumerate(vals):
                key = re.sub(r'\s+', '', v).upper()
                for fld, aliases in EO_COLUMN_ALIASES.items():
                    if any(re.sub(r'\s+', '', a).upper() == key for a in aliases):
                        found.setdefault(fld, ci)
            if 'eo_no' in found and len(found) >= 3:
                hdr_row, colmap = ri, found
                break
        if hdr_row is None:
            return JSONResponse({'error': '«EO 번호» 헤더를 찾지 못했습니다. 통보서 목록 양식인지 확인하세요.'},
                                status_code=400)
        rows = []
        for ri in range(hdr_row + 1, len(raw)):
            rec = {}
            for fld, ci in colmap.items():
                v = raw.iat[ri, ci] if ci < raw.shape[1] else None
                if pd.isna(v):
                    rec[fld] = ''
                elif fld == 'eo_date':
                    s = str(v).strip()
                    rec[fld] = re.sub(r'[^0-9]', '', s)[:8] if re.search(r'\d', s) else ''
                else:
                    rec[fld] = str(v).strip()
            if rec.get('eo_no'):
                rows.append(rec)
        if not rows:
            return JSONResponse({'error': 'EO 번호가 있는 행을 찾지 못했습니다.'}, status_code=400)
        res = upsert_eo_notices(rows, me['username'])
    except Exception as ex:
        return JSONResponse({'error': f'엑셀 처리 오류: {ex}'}, status_code=400)
    finally:
        try: os.unlink(tmp)
        except OSError: pass
    res.update({'ok': True, 'parsed': len(rows), 'columns': sorted(colmap),
                'stats': get_eo_stats()})
    return JSONResponse(res)


@app.get('/eo/users')
async def eo_users(request: Request):
    """결재선 선택 창에 쓸 목록. 조직도가 등록돼 있으면 «사업장>부서>사람» 트리를 주고,
       아직 없으면 로그인 계정 목록으로 대체한다(조직도 없이도 결재가 돌아가야 하므로)."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    tree = get_org_tree()
    accounts = []
    for u in get_all_users():
        if u.get('role') not in ('user', 'admin'):
            continue
        accounts.append({'emp_id': u['username'], 'name': u.get('name') or u['username'],
                         'dept': u.get('dept') or '', 'position': '',
                         'email': u.get('email') or '', 'site': '',
                         'has_account': True})
    if not tree:
        # 조직도 미등록 — 계정만으로 한 덩어리 트리를 만들어 준다
        by = {}
        for a in accounts:
            by.setdefault(a['dept'] or '(부서 미지정)', []).append(a)
        tree = [{'site': '계정 목록',
                 'depts': [{'dept': d, 'members': m} for d, m in sorted(by.items())]}]
    return JSONResponse({'tree': tree, 'types': EO_APPROVAL_TYPES,
                         'ref_code': EO_TYPE_REFERENCE,
                         'org_ready': bool(get_org_stats()['total']),
                         'stats': get_org_stats()})


# ── 조직도 (사내 PLM 1회 등록) ────────────────────────────────────────────────
@app.get('/org', response_class=HTMLResponse)
async def org_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='org.html', context={
        'me': me, 'stats': get_org_stats(), 'types': EO_APPROVAL_TYPES,
    })


@app.get('/org/list')
async def org_list(request: Request, q: str = '', dept: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    items = get_org_members(q.strip(), dept.strip())
    accounts = {u['username'] for u in get_all_users() if u.get('role') in ('user', 'admin')}
    for it in items:
        it['has_account'] = it['emp_id'] in accounts
    return JSONResponse({'items': items, 'stats': get_org_stats(), 'tree': get_org_tree()})


@app.post('/org/upload')
def org_upload(request: Request, file: UploadFile = File(...)):
    """조직도 엑셀 일괄 등록. 열 이름으로 자동 인식하므로 열 순서는 상관없다."""
    redir = require_admin(request)
    if redir: return JSONResponse({'error': '관리자만 등록할 수 있습니다.'}, status_code=403)
    me = current_user(request)
    name = (file.filename or '').lower()
    if not name.endswith(('.xlsx', '.xlsm', '.xls', '.csv')):
        return JSONResponse({'error': '엑셀(.xlsx) 또는 CSV 파일을 올려 주세요.'}, status_code=400)
    tmp = os.path.join(tempfile.gettempdir(), f'org_{uuid.uuid4().hex[:8]}_{name[-40:]}')
    with open(tmp, 'wb') as out:
        shutil.copyfileobj(file.file, out)
    try:
        table = []
        if name.endswith('.csv'):
            import csv
            for enc in ('utf-8-sig', 'cp949'):
                try:
                    with open(tmp, newline='', encoding=enc) as fh:
                        table = [r for r in csv.reader(fh)]
                    break
                except UnicodeDecodeError:
                    continue
        else:
            import openpyxl
            wb = openpyxl.load_workbook(tmp, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                table.append(['' if c is None else str(c) for c in row])
            wb.close()
        parsed = parse_org_rows(table)
        if parsed.get('error'):
            return JSONResponse({'error': parsed['error'] +
                                 ' 첫 행에 «아이디(사번)»·«성명» 같은 열 제목이 있어야 합니다.'},
                                status_code=400)
        res = upsert_org_members(parsed['rows'], me['username'])
        res['columns'] = parsed['columns']
        res['stats'] = get_org_stats()
        return JSONResponse(res)
    except Exception as ex:
        return JSONResponse({'error': f'읽기 실패: {ex}'}, status_code=400)
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


@app.post('/org/paste')
async def org_paste(request: Request):
    """PLM 화면에서 표를 그대로 복사해 붙여넣는 경로. 엑셀 없이도 등록할 수 있게 한다."""
    redir = require_admin(request)
    if redir: return JSONResponse({'error': '관리자만 등록할 수 있습니다.'}, status_code=403)
    me = current_user(request)
    body = await request.json()
    text = str(body.get('text') or '')
    table = [re.split(r'\t|,|\s{2,}', ln.rstrip()) for ln in text.splitlines() if ln.strip()]
    parsed = parse_org_rows(table)
    if parsed.get('error'):
        return JSONResponse({'error': parsed['error'] +
                             ' 첫 줄에 «아이디  성명  부서» 같은 열 제목을 같이 붙여넣어 주세요.'},
                            status_code=400)
    res = upsert_org_members(parsed['rows'], me['username'])
    res['columns'] = parsed['columns']
    res['stats'] = get_org_stats()
    return JSONResponse(res)


@app.post('/org/delete/{emp_id}')
async def org_delete(request: Request, emp_id: str):
    redir = require_admin(request)
    if redir: return JSONResponse({'error': '관리자만 삭제할 수 있습니다.'}, status_code=403)
    return JSONResponse(delete_org_member(emp_id))


@app.post('/org/clear')
async def org_clear(request: Request):
    redir = require_admin(request)
    if redir: return JSONResponse({'error': '관리자만 초기화할 수 있습니다.'}, status_code=403)
    return JSONResponse({'ok': True, 'deleted': clear_org_members(), 'stats': get_org_stats()})


# ── 전자결재 ──────────────────────────────────────────────────────────────────
@app.post('/eo/approval/submit/{eo_id}')
async def eo_approval_submit(request: Request, eo_id: int):
    """상신 — 결재선을 만들고 1단계를 승인대기로 둔다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    res = submit_eo_approval(eo_id, body.get('line') or [], me['username'])
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg', '상신 실패')}, status_code=400)
    return JSONResponse(res)


@app.post('/eo/approval/act/{eo_id}')
async def eo_approval_act(request: Request, eo_id: int):
    """승인 / 반려. 본인 차례만 처리 가능하며 관리자는 대결할 수 있다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    res = act_eo_approval(eo_id, me['username'], str(body.get('action', '')),
                          str(body.get('comment', ''))[:500], is_admin=(me['role'] == 'admin'))
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg', '처리 실패')}, status_code=400)
    return JSONResponse(res)


@app.post('/eo/approval/reopen/{eo_id}')
async def eo_approval_reopen(request: Request, eo_id: int):
    """반려된 문서를 작성중으로 되돌려 재상신할 수 있게 한다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse(reopen_eo_approval(eo_id))


# ── 메일 전송 ─────────────────────────────────────────────────────────────────
# SMTP가 설정되지 않았으면 «실제 발송 없이 이력만» 남긴다. 설정 없이 자동 발송되어
# 엉뚱한 사람에게 메일이 나가는 사고를 막기 위한 기본값이다.
def _smtp_config() -> dict:
    return {
        'host': os.environ.get('BOM_SMTP_HOST', '').strip(),
        'port': int(os.environ.get('BOM_SMTP_PORT', '587') or 587),
        'user': os.environ.get('BOM_SMTP_USER', '').strip(),
        'password': os.environ.get('BOM_SMTP_PASS', '').strip(),
        'sender': os.environ.get('BOM_SMTP_FROM', '').strip(),
        'use_tls': os.environ.get('BOM_SMTP_TLS', '1') != '0',
    }


def _send_mail(to_list: list, subject: str, body: str) -> tuple:
    """반환: (status, detail). SMTP 미설정이면 ('skipped', 사유)."""
    cfg = _smtp_config()
    if not cfg['host'] or not cfg['sender']:
        return 'skipped', 'SMTP 미설정 — 발송하지 않고 이력만 기록했습니다.'
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.header import Header
        msg = MIMEText(body, 'plain', 'utf-8')
        msg['Subject'] = Header(subject, 'utf-8')
        msg['From'] = cfg['sender']
        msg['To'] = ', '.join(to_list)
        with smtplib.SMTP(cfg['host'], cfg['port'], timeout=20) as s:
            if cfg['use_tls']:
                s.starttls()
            if cfg['user']:
                s.login(cfg['user'], cfg['password'])
            s.sendmail(cfg['sender'], to_list, msg.as_string())
        return 'sent', f'{len(to_list)}명에게 발송'
    except Exception as ex:
        return 'failed', str(ex)[:300]


@app.get('/eo/mail/{eo_id}')
async def eo_mail_list(request: Request, eo_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    cfg = _smtp_config()
    return JSONResponse({'items': get_eo_mails(eo_id),
                         'smtp_ready': bool(cfg['host'] and cfg['sender'])})


@app.post('/eo/mail/{eo_id}')
async def eo_mail_send(request: Request, eo_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    e = get_eo_notice(eo_id)
    if not e:
        return JSONResponse({'error': '통보서를 찾을 수 없습니다.'}, status_code=404)
    body = await request.json()
    # 수신자: 직접 입력한 주소 + 결재선·참조자 이름을 계정 이메일로 변환
    addrs = [a.strip() for a in re.split(r'[,;\s]+', str(body.get('to', ''))) if '@' in a]
    names = [n.strip() for n in re.split(r'[,;]', str(body.get('names', ''))) if n.strip()]
    resolved = get_user_emails(names)
    addrs += [v for v in resolved.values() if v]
    addrs = sorted(set(a for a in addrs if a))
    if not addrs:
        return JSONResponse({'error': '수신자를 찾지 못했습니다. 이메일 주소를 직접 입력하거나 '
                                      '계정에 등록된 이름을 지정하세요.'}, status_code=400)
    subject = str(body.get('subject') or f"[설계변경 통보] {e['eo_no']}")[:200]
    text = str(body.get('body') or '')
    if not text:
        text = (f"설계변경 통보서 {e['eo_no']}\n"
                f"일자: {e.get('eo_date','')}\n차종: {e.get('vehicle_code','')}\n"
                f"진행상태: {EO_APPROVAL_STATUS.get(e.get('approval_status') or 'draft', '')}\n\n"
                f"{e.get('content','')}\n")
    status, detail = _send_mail(addrs, subject, text)
    add_eo_mail(eo_id, ', '.join(addrs), subject, text, status, detail, me['username'])
    return JSONResponse({'ok': True, 'status': status, 'detail': detail,
                         'to': addrs, 'unresolved': [n for n in names if n not in resolved],
                         'items': get_eo_mails(eo_id)})


@app.post('/eo/items/{eo_id}')
async def eo_items_save(request: Request, eo_id: int):
    """품목현황 저장. 품번만 적어도 품목 마스터에서 품명을 채워 준다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not get_eo_notice(eo_id):
        return JSONResponse({'error': '통보서를 찾을 수 없습니다.'}, status_code=404)
    body = await request.json()
    set_eo_items(eo_id, body.get('items') or [])
    return JSONResponse({'ok': True, 'items': get_eo_items(eo_id)})


# 확장자로 2D/3D를 자동 판정한다 — 통보서 도면현황이 이 둘을 나눠 보여주기 때문
EO_2D_EXTS = ('.pdf', '.dwg', '.dxf', '.catdrawing', '.tif', '.tiff', '.png', '.jpg', '.jpeg')
EO_3D_EXTS = ('.catpart', '.catproduct', '.stp', '.step', '.igs', '.iges', '.jt', '.stl')


def _eo_doc_kind(filename: str) -> str:
    low = filename.lower()
    if low.endswith(EO_3D_EXTS):
        return '3d'
    if low.endswith(EO_2D_EXTS):
        return '2d'
    return 'doc'


@app.post('/eo/file/{eo_id}')
def eo_file_upload(request: Request, eo_id: int,
                   purpose: str = Form(''), doc_kind: str = Form(''),
                   file: UploadFile = File(...)):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if not get_eo_notice(eo_id):
        return JSONResponse({'error': '통보서를 찾을 수 없습니다.'}, status_code=404)
    ext = os.path.splitext(file.filename)[1].lower()
    saved = os.path.join(EO_FILE_DIR, f'{uuid.uuid4().hex[:12]}{ext}')
    with open(saved, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    kind = doc_kind.strip() or _eo_doc_kind(file.filename)
    size = os.path.getsize(saved)
    add_eo_file(eo_id, file.filename, saved, me['username'], doc_kind=kind,
                purpose=purpose.strip(), size_no=str(size),
                file_type=(ext.lstrip('.') or 'file'))
    return JSONResponse({'ok': True, 'files': get_eo_files(eo_id)})


@app.get('/eo/file/view/{file_id}')
async def eo_file_view(request: Request, file_id: int):
    redir = require_login(request)
    if redir: return redir
    f = get_eo_file(file_id)
    if not f or not os.path.exists(f['file_path']):
        return JSONResponse({'error': '파일이 없습니다.'}, status_code=404)
    return FileResponse(f['file_path'], filename=f['filename'])


@app.post('/eo/file/delete/{file_id}')
async def eo_file_delete(request: Request, file_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    f = get_eo_file(file_id)
    if not f:
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and f['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    info = delete_eo_file(file_id)
    if info:
        try:
            if os.path.exists(info['file_path']): os.unlink(info['file_path'])
        except OSError:
            pass
    return JSONResponse({'ok': True})


@app.post('/eo/delete/{eo_id}')
async def eo_delete(request: Request, eo_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    e = get_eo_notice(eo_id)
    if not e:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and (e.get('created_by') or '') != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    for p in delete_eo_notice(eo_id):
        try:
            if p and os.path.exists(p): os.unlink(p)
        except OSError:
            pass
    return JSONResponse({'ok': True})


# ── 문서 게시판 (이용방법 / RFP) ──────────────────────────────────────────────
# usage : 사이트 이용방법 — 관리자만 열람·작성
# rfp   : PLM/ERP 업체 제출용 RFP — 관리자가 작성, 로그인 사용자는 열람
DOC_FILE_DIR = os.path.join(DATA_DIR, 'doc_files')
os.makedirs(DOC_FILE_DIR, exist_ok=True)
DOC_KINDS = {
    # 열람은 전 사용자, 작성·수정은 관리자만(_doc_guard 의 need_write 로 분리).
    'usage': {'title': '사이트 이용방법', 'icon': '📘', 'admin_only': False,
              'desc': '이 시스템의 게시판별 사용법과 운영 규칙입니다. 작성·수정은 관리자만 할 수 있습니다.'},
    'rfp':   {'title': 'PLM / ERP RFP', 'icon': '📑', 'admin_only': False,
              'desc': 'PLM·ERP 업체에 제출할 요구사항을 정리합니다. '
                      '이미 자체 구축한 기능은 «구축 완료»로 표시해 도입 범위를 명확히 합니다.',
              'download': {
                  'url': '/guide/rfp/download',
                  'title': '제안요청서 전문 (Word)',
                  'desc': '업체에 그대로 보낼 수 있는 제안요청서입니다. 기준정보 문제 정의(실측 근거), '
                          '수립 원칙 6가지, 자체 구축 현황, PLM·ERP 요구사항, 평가 기준, '
                          '부록(실측 데이터·용어 매핑)으로 구성됩니다. '
                          'SCM·MES·더존 ERP 연계는 범위에서 제외했습니다.'}},
}

RFP_DOC_PATH = os.path.join(DATA_DIR, 'rfp', 'DYA_PLM_ERP_RFP.docx')


@app.get('/guide/rfp/download')
def guide_rfp_download(request: Request):
    """RFP Word 문서. 없으면 그 자리에서 생성한다(배포 후 최초 1회)."""
    redir = require_login(request)
    if redir: return redir
    if not os.path.exists(RFP_DOC_PATH):
        try:
            sys.path.insert(0, os.path.join(BASE_DIR, 'tools'))
            import gen_rfp_docx
            gen_rfp_docx.build()
        except Exception as ex:
            return JSONResponse({'error': f'문서를 생성하지 못했습니다: {ex}'}, status_code=500)
    if not os.path.exists(RFP_DOC_PATH):
        return JSONResponse({'error': '문서를 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(
        RFP_DOC_PATH, filename='대유에이피_PLM_ERP_도입_제안요청서.docx',
        media_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')


def _doc_guard(request: Request, kind: str, need_write: bool = False):
    """열람·작성 권한 확인. 반환값이 있으면 그대로 응답으로 돌려준다."""
    if kind not in DOC_KINDS:
        return JSONResponse({'error': '잘못된 게시판입니다.'}, status_code=404)
    if require_login(request):
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    admin_only = DOC_KINDS[kind]['admin_only']
    if (admin_only or need_write) and me['role'] != 'admin':
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    return None


@app.get('/guide/{kind}', response_class=HTMLResponse)
async def docs_page(request: Request, kind: str):
    if kind not in DOC_KINDS:
        return RedirectResponse('/', status_code=302)
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    if DOC_KINDS[kind]['admin_only'] and me['role'] != 'admin':
        return RedirectResponse('/', status_code=302)
    return templates.TemplateResponse(request=request, name='docs.html', context={
        'me': me, 'kind': kind, 'meta': DOC_KINDS[kind],
        'can_edit': me['role'] == 'admin',
    })


@app.get('/guide/{kind}/list')
async def docs_list(request: Request, kind: str):
    g = _doc_guard(request, kind)
    if g: return g
    return JSONResponse({'items': get_doc_posts(kind)})


@app.post('/guide/{kind}/save')
async def docs_save(request: Request, kind: str):
    g = _doc_guard(request, kind, need_write=True)
    if g: return g
    me = current_user(request)
    body = await request.json()
    title = str(body.get('title', '')).strip()
    if not title:
        return JSONResponse({'error': '제목을 입력하세요.'}, status_code=400)
    cat = str(body.get('category', '')).strip()
    txt = str(body.get('body', ''))
    so = int(body.get('sort_order') or 0)
    pid = body.get('id')
    if pid:
        res = update_doc_post(int(pid), cat, title, txt, me['username'], so)
        if not res.get('ok'):
            return JSONResponse({'error': res.get('msg', '저장 실패')}, status_code=400)
        return JSONResponse({'ok': True, 'id': int(pid), **res})
    new_id = add_doc_post(kind, cat, title, txt, me['username'], so)
    return JSONResponse({'ok': True, 'id': new_id, 'revision': 1, 'changed': True})


@app.post('/guide/{kind}/delete/{post_id}')
async def docs_delete(request: Request, kind: str, post_id: int):
    g = _doc_guard(request, kind, need_write=True)
    if g: return g
    for p in delete_doc_post(post_id):
        try:
            if p and os.path.exists(p): os.unlink(p)
        except OSError:
            pass
    return JSONResponse({'ok': True})


@app.post('/guide/{kind}/file/{post_id}')
def docs_file_upload(request: Request, kind: str, post_id: int, file: UploadFile = File(...)):
    g = _doc_guard(request, kind, need_write=True)
    if g: return g
    me = current_user(request)
    if not get_doc_post(post_id):
        return JSONResponse({'error': '문서를 찾을 수 없습니다.'}, status_code=404)
    ext = os.path.splitext(file.filename)[1].lower()
    saved = os.path.join(DOC_FILE_DIR, f'{uuid.uuid4().hex[:12]}{ext}')
    with open(saved, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    add_doc_file(post_id, file.filename, saved, me['username'])
    return JSONResponse({'ok': True})


@app.get('/guide/file/view/{file_id}')
async def docs_file_view(request: Request, file_id: int):
    redir = require_login(request)
    if redir: return redir
    f = get_doc_file(file_id)
    if not f or not os.path.exists(f['file_path']):
        return JSONResponse({'error': '파일이 없습니다.'}, status_code=404)
    return FileResponse(f['file_path'], filename=f['filename'])


@app.post('/guide/{kind}/file/delete/{file_id}')
async def docs_file_delete(request: Request, kind: str, file_id: int):
    g = _doc_guard(request, kind, need_write=True)
    if g: return g
    info = delete_doc_file(file_id)
    if info:
        try:
            if os.path.exists(info['file_path']): os.unlink(info['file_path'])
        except OSError:
            pass
    return JSONResponse({'ok': True})


# ── 품목 게시판 (PLM 연동 대상 품목 마스터) ───────────────────────────────────
# BOM 엑셀을 올리면 전 레벨 품번·품명이 자동 등록되고, 품목별 스펙(재질·중량·
# MS SPEC·도면 등)을 사람이 채운다. 품번이 전사 연결 키다.
PART_FILE_DIR = os.path.join(DATA_DIR, 'part_files')
os.makedirs(PART_FILE_DIR, exist_ok=True)
PART_DRAWING_EXTS = ('.pdf', '.dwg', '.dxf', '.png', '.jpg', '.jpeg', '.tif', '.tiff')


# ══════════════════════════════════════════════════════════════════════════════
# 협력사 도면 배포
# ══════════════════════════════════════════════════════════════════════════════
# 협력사 계정(role='partner')은 «배포 게시판만» 볼 수 있어야 한다.
# 아래 목록 밖의 주소로 가면 배포 화면으로 되돌린다(화면을 숨기는 게 아니라 서버가 막는다).
PARTNER_ALLOWED = ('/dist', '/dist/list', '/dist/detail', '/dist/get',
                   '/logout', '/login', '/static')


@app.middleware('http')
async def partner_scope_guard(request: Request, call_next):
    """협력사 계정이 내부 게시판에 못 들어가게 막는다."""
    path = request.url.path
    if not path.startswith(('/static', '/login', '/logout', '/favicon')):
        try:
            me = current_user(request)
        except Exception:
            me = None
        if me and me.get('role') == 'partner':
            if not any(path == a or path.startswith(a + '/') for a in PARTNER_ALLOWED):
                if path.startswith('/catia') or path.startswith('/parts') \
                        or path.startswith('/eo') or path.startswith('/cust-eo'):
                    return JSONResponse({'error': '협력사 계정은 접근할 수 없습니다.'},
                                        status_code=403)
                return RedirectResponse('/dist', status_code=302)
    return await call_next(request)


@app.get('/dist', response_class=HTMLResponse)
async def dist_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    is_partner = (me['role'] == 'partner')
    return templates.TemplateResponse(request=request, name='dist.html', context={
        'me': me, 'is_partner': is_partner,
        'partners': [] if is_partner else get_partners(),
        'vcodes': get_all_vehicle_codes(),
        'stats': get_dist_stats(me.get('partner_code') or '' if is_partner else ''),
    })


@app.get('/dist/list')
async def dist_list(request: Request, q: str = '', vehicle: str = '', status: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    pc = (me.get('partner_code') or '') if me['role'] == 'partner' else ''
    res = search_dist(q.strip(), vehicle.strip(), status.strip(), pc)
    res['stats'] = get_dist_stats(pc)
    res['is_partner'] = (me['role'] == 'partner')
    res['status_labels'] = DIST_STATUS
    return JSONResponse(res)


@app.get('/dist/detail/{did}')
async def dist_detail(request: Request, did: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    d = get_dist(did)
    if not d:
        return JSONResponse({'error': '배포 건을 찾을 수 없습니다.'}, status_code=404)
    pc = (me.get('partner_code') or '') if me['role'] == 'partner' else ''
    if pc:
        if d['status'] != 'sent' or pc not in [t['partner_code'] for t in get_dist_targets(did)]:
            return JSONResponse({'error': '볼 수 없는 배포 건입니다.'}, status_code=403)
    files = get_dist_files(did)
    if pc:
        # 협력사에게는 «허용된 종류»만 보인다 — 원본 미제공이면 목록에서 아예 뺀다
        ORIG = ('.catpart', '.catproduct', '.catdrawing')
        files = [f for f in files
                 if (((f['ext'] or '').lower() in ORIG) and d['share_orig'])
                 or (((f['ext'] or '').lower() not in ORIG) and d['share_conv'])]
    out = {'ok': True, 'dist': d, 'files': files, 'targets': get_dist_targets(did),
           'is_partner': bool(pc), 'status_labels': DIST_STATUS}
    if not pc:
        out['downloads'] = get_dist_downloads(did)
    return JSONResponse(out)


@app.post('/dist/save')
async def dist_save(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if me['role'] == 'partner':
        return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    b = await request.json()
    fields = b.get('fields') or {}
    did = b.get('id')
    if did:
        update_dist(int(did), fields)
        return JSONResponse({'ok': True, 'id': int(did)})
    res = create_dist(fields, me['username'])
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=400)
    return JSONResponse(res)


@app.get('/dist/candidates')
async def dist_candidates(request: Request, eo_id: int = 0, part_nos: str = '',
                          vehicle: str = ''):
    """배포 후보 — «배포완료» 상태인 것만 실제로 고를 수 있다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if me['role'] == 'partner':
        return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    return JSONResponse(get_dist_candidates(eo_id, part_nos, vehicle))


@app.post('/dist/files/{did}')
async def dist_set_files(request: Request, did: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if me['role'] == 'partner':
        return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    b = await request.json()
    set_dist_files(did, b.get('file_ids') or [])
    return JSONResponse({'ok': True, 'files': get_dist_files(did)})


@app.post('/dist/targets/{did}')
async def dist_set_targets(request: Request, did: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if me['role'] == 'partner':
        return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    b = await request.json()
    set_dist_targets(did, b.get('codes') or [])
    return JSONResponse({'ok': True, 'targets': get_dist_targets(did)})


@app.post('/dist/send/{did}')
async def dist_send(request: Request, did: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if me['role'] == 'partner':
        return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    res = send_dist(did, me['username'])
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=400)
    return JSONResponse(res)


@app.post('/dist/delete/{did}')
async def dist_delete(request: Request, did: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if me['role'] == 'partner':
        return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    return JSONResponse(delete_dist(did))


@app.get('/dist/get/{did}/{file_id}')
async def dist_get_file(request: Request, did: int, file_id: int):
    """배포 파일 내려받기. 협력사는 권한을 확인하고 «받은 기록»을 남긴다."""
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    pc = (me.get('partner_code') or '') if me['role'] == 'partner' else ''
    if pc:
        ok, why = can_partner_get(did, file_id, pc)
        if not ok:
            return JSONResponse({'error': why}, status_code=403)
    f = get_catia_file(file_id)
    if not f or not os.path.exists(f.get('file_path') or ''):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    log_dist_download(did, file_id, me['username'], pc)
    return FileResponse(f['file_path'], filename=f['filename'],
                        media_type='application/octet-stream')


# ── 협력사 계정 관리 (관리자) ─────────────────────────────────────────────────
@app.get('/partners/list')
async def partners_list(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if me['role'] == 'partner':
        return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    return JSONResponse({'items': get_partners(), 'accounts': get_partner_accounts()})


@app.post('/partners/account')
async def partners_account(request: Request):
    """협력사 로그인 계정 생성. 비밀번호는 관리자가 직접 입력한다(자동 생성 안 함)."""
    redir = require_admin(request)
    if redir: return JSONResponse({'error': '관리자만 가능합니다.'}, status_code=403)
    b = await request.json()
    res = create_partner_account(b.get('username', ''), b.get('password', ''),
                                 b.get('partner_code', ''), b.get('partner_name', ''),
                                 b.get('email', ''))
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=400)
    return JSONResponse({'ok': True, 'accounts': get_partner_accounts()})


# ══════════════════════════════════════════════════════════════════════════════
# 고객 EO
# ══════════════════════════════════════════════════════════════════════════════
CUST_EO_DIR = os.path.join(DATA_DIR, 'cust_eo')
os.makedirs(CUST_EO_DIR, exist_ok=True)
seed_customers()


@app.get('/cust-eo', response_class=HTMLResponse)
async def cust_eo_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='cust_eo.html', context={
        'me': me, 'customers': get_customers(), 'vcodes': get_all_vehicle_codes(),
        'types': CUST_EO_TYPES, 'stats': get_cust_eo_stats(),
    })


@app.get('/cust-eo/list')
async def cust_eo_list(request: Request, q: str = '', cust: str = '', vehicle: str = '',
                       eo_type: str = '', dev_schedule: str = '', date_from: str = '',
                       date_to: str = '', part_no: str = '', limit: int = 500):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    res = search_cust_eo(q.strip(), cust.strip(), vehicle.strip(), eo_type.strip(),
                         dev_schedule.strip(), date_from.strip(), date_to.strip(),
                         part_no.strip(), min(limit, 2000))
    res['stats'] = get_cust_eo_stats()
    res['status_labels'] = EO_APPROVAL_STATUS
    return JSONResponse(res)


@app.get('/cust-eo/detail/{cid}')
async def cust_eo_detail(request: Request, cid: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    c = get_cust_eo(cid)
    if not c:
        return JSONResponse({'error': '고객EO를 찾을 수 없습니다.'}, status_code=404)
    me = current_user(request)
    approvals = get_eo_approvals(cid, doc_type='cust')
    pending = next((a for a in approvals if a['status'] == 'pending'), None)
    my_turn = bool(pending) and (
        me['username'] in str(pending['approver']) or
        (me.get('name') or '\x00') in str(pending['approver']))
    return JSONResponse({
        'ok': True, 'eo': c, 'files': get_cust_eo_files(cid),
        'links': get_cust_eo_links(cid), 'rels': get_cust_eo_rels(cid),
        'drawings': get_cust_eo_drawings(cid),
        'approvals': approvals, 'my_turn': my_turn,
        'is_admin': me['role'] == 'admin', 'status_labels': EO_APPROVAL_STATUS,
        'roles': EO_APPROVAL_ROLES, 'file_kinds': CUST_EO_FILE_KINDS,
    })


@app.post('/cust-eo/save')
async def cust_eo_save(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    fields = {k: v for k, v in (body.get('fields') or {}).items() if k in CUST_EO_FIELDS}
    cid = body.get('id')
    if cid:
        cur = get_cust_eo(int(cid)) or {}
        st = cur.get('approval_status') or 'draft'
        if st in ('submitted', 'in_progress', 'approved'):
            return JSONResponse({'error': f'«{EO_APPROVAL_STATUS.get(st, st)}» 상태에서는 '
                                          f'수정할 수 없습니다.'}, status_code=400)
        update_cust_eo(int(cid), fields)
        return JSONResponse({'ok': True, 'id': int(cid)})
    res = create_cust_eo(fields, me['username'])
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=400)
    return JSONResponse(res)


@app.post('/cust-eo/delete/{cid}')
async def cust_eo_delete(request: Request, cid: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse(delete_cust_eo(cid))


# ── 상호 연결 ─────────────────────────────────────────────────────────────────
@app.post('/cust-eo/links/{cid}')
async def cust_eo_set_links(request: Request, cid: int):
    """고객EO ↔ 내부 설계변경통보서 (여러 건 연결)."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    b = await request.json()
    set_cust_eo_links(cid, b.get('eo_ids') or [])
    return JSONResponse({'ok': True, 'links': get_cust_eo_links(cid),
                         'drawings': get_cust_eo_drawings(cid)})


@app.post('/cust-eo/rels/{cid}')
async def cust_eo_set_rels(request: Request, cid: int):
    """비슷한 내용의 다른 고객EO 연결."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    b = await request.json()
    set_cust_eo_rels(cid, b.get('ids') or [])
    return JSONResponse({'ok': True, 'rels': get_cust_eo_rels(cid)})


@app.get('/cust-eo/pick')
async def cust_eo_pick(request: Request, kind: str = 'eo', q: str = ''):
    """연결 창에서 쓸 목록. kind=eo → 내부 설계변경통보서, kind=cust → 고객EO."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if kind == 'cust':
        r = search_cust_eo(q=q.strip(), limit=200)
        return JSONResponse({'items': [
            {'id': i['id'], 'no': i['cust_eo_no'], 'sub': i['cust_code'],
             'date': i['eo_date'], 'content': (i['content'] or '')[:70]} for i in r['items']]})
    r = search_eo_notices(q=q.strip(), limit=200)
    return JSONResponse({'items': [
        {'id': i['id'], 'no': i['eo_no'], 'sub': i.get('vehicle_code') or '',
         'date': i.get('eo_date') or '', 'content': (i.get('content') or '')[:70]}
        for i in r['items']]})


@app.get('/eo/cust-links/{eo_id}')
async def eo_cust_links(request: Request, eo_id: int):
    """내부 설계변경통보서에서 본 고객EO 연결(반대 방향)."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'items': get_eo_cust_links(eo_id)})


@app.post('/eo/cust-links/{eo_id}')
async def eo_cust_links_set(request: Request, eo_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    b = await request.json()
    set_eo_cust_links(eo_id, b.get('ids') or [])
    return JSONResponse({'ok': True, 'items': get_eo_cust_links(eo_id)})


# ── 첨부파일 ─────────────────────────────────────────────────────────────────
@app.post('/cust-eo/file/{cid}')
def cust_eo_file_upload(request: Request, cid: int, kind: str = Form('etc'),
                        file: UploadFile = File(...)):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if not get_cust_eo(cid):
        return JSONResponse({'error': '고객EO를 찾을 수 없습니다.'}, status_code=404)
    safe = f"{uuid.uuid4().hex[:10]}_{re.sub(r'[^A-Za-z0-9._-]', '_', file.filename or 'f')[-80:]}"
    dest = os.path.join(CUST_EO_DIR, safe)
    size = 0
    with open(dest, 'wb') as out:
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            out.write(chunk); size += len(chunk)
    add_cust_eo_file(cid, kind, file.filename or safe, dest, size, me['username'])
    return JSONResponse({'ok': True, 'files': get_cust_eo_files(cid)})


@app.get('/cust-eo/file/view/{fid}')
async def cust_eo_file_view(request: Request, fid: int):
    redir = require_login(request)
    if redir: return redir
    f = get_cust_eo_file(fid)
    if not f or not os.path.exists(f.get('file_path') or ''):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(f['file_path'], filename=f['filename'])


@app.post('/cust-eo/file/delete/{fid}')
async def cust_eo_file_delete(request: Request, fid: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse(delete_cust_eo_file(fid))


# ── 전자결재 (내부 EO 와 같은 로직, doc_type 으로 구분) ────────────────────────
@app.post('/cust-eo/approval/submit/{cid}')
async def cust_eo_appr_submit(request: Request, cid: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    b = await request.json()
    res = submit_eo_approval(cid, b.get('line') or [], me['username'], doc_type='cust')
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=400)
    return JSONResponse(res)


@app.post('/cust-eo/approval/act/{cid}')
async def cust_eo_appr_act(request: Request, cid: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    b = await request.json()
    res = act_eo_approval(cid, me['username'], str(b.get('action', '')),
                          str(b.get('comment', ''))[:500],
                          is_admin=(me['role'] == 'admin'), doc_type='cust')
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=400)
    return JSONResponse(res)


@app.post('/cust-eo/approval/reopen/{cid}')
async def cust_eo_appr_reopen(request: Request, cid: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse(reopen_eo_approval(cid, doc_type='cust'))


# ── 고객 마스터 ───────────────────────────────────────────────────────────────
@app.get('/customers', response_class=HTMLResponse)
async def customers_page(request: Request):
    """고객 마스터 — 전사 공통 키라 마스터 데이터 메뉴에 둔다(ERP·SCM 과 코드가 맞아야 하므로)."""
    redir = require_login(request)
    if redir: return redir
    return templates.TemplateResponse(request=request, name='customers.html',
                                      context={'me': current_user(request)})


@app.get('/customers/list')
async def customers_list(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'items': get_customers(active_only=False)})


@app.post('/customers/save')
async def customers_save(request: Request):
    redir = require_admin(request)
    if redir: return JSONResponse({'error': '관리자만 수정할 수 있습니다.'}, status_code=403)
    b = await request.json()
    res = upsert_customer(b.get('code', ''), b.get('name', ''), b.get('sort_no', 0),
                          1 if b.get('active', True) else 0, b.get('note', ''))
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=400)
    return JSONResponse({'ok': True, 'items': get_customers(active_only=False)})


@app.post('/customers/delete/{code}')
async def customers_delete(request: Request, code: str):
    redir = require_admin(request)
    if redir: return JSONResponse({'error': '관리자만 삭제할 수 있습니다.'}, status_code=403)
    res = delete_customer(code)
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=400)
    return JSONResponse({'ok': True, 'items': get_customers(active_only=False)})


# ══════════════════════════════════════════════════════════════════════════════
# 카티아 2D/3D 파일 관리
# ══════════════════════════════════════════════════════════════════════════════
CATIA_DIR = os.path.join(DATA_DIR, 'catia')
os.makedirs(CATIA_DIR, exist_ok=True)
# 추출 규칙을 고쳤으면 기존 등록분도 여기서 한 번 따라온다(재업로드 불필요)
refresh_catia_derived()


@app.get('/catia', response_class=HTMLResponse)
async def catia_page(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='catia.html', context={
        'me': me, 'vcodes': get_all_vehicle_codes(), 'sel_vehicle': vehicle,
        'groups': CATIA_PART_GROUPS, 'stages': CATIA_STAGES,
        'stats': get_catia_stats(vehicle),
    })


@app.get('/catia/list')
async def catia_list(request: Request, vehicle: str = '', row_level: str = '', stage: str = '',
                     part_group: str = '', part_type: str = '', side: str = '',
                     kind: str = '', q: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    res = search_catia_parts(vehicle.strip(), row_level.strip(), stage.strip(),
                             part_group.strip(), part_type.strip(), side.strip(),
                             kind.strip(), q.strip())
    # 체크아웃·수명주기 상태를 품번마다 붙인다 + 내가 손댈 수 있는지도 서버가 판단해 준다
    me = current_user(request)
    items = get_catia_items_map([(it['vehicle_code'], it['part_no']) for it in res['items']])
    for it in res['items']:
        st = items.get((it['vehicle_code'], base_part_no(it['part_no']))) or {}
        it['lock'] = {'state': st.get('state') or 'work',
                      'locked_by': st.get('locked_by') or '',
                      'locked_at': st.get('locked_at') or '',
                      'released_rev': st.get('released_rev') or ''}
        ok, why = catia_can_modify(it['vehicle_code'], it['part_no'], me['username'],
                                   is_admin=(me['role'] == 'admin'))
        it['lock']['can_modify'] = ok
        it['lock']['why'] = why
        it['lock']['mine'] = (st.get('locked_by') or '') == me['username']
    res['facets'] = get_catia_facets(vehicle.strip())
    res['stats'] = get_catia_stats(vehicle.strip())
    res['stats'].update(get_catia_lock_stats(vehicle.strip()))
    res['group_label'] = CATIA_GROUP_LABEL
    res['states'] = CATIA_STATES
    res['me'] = me['username']
    # 배포·개정은 설계자만 — 화면에서 버튼을 감추기 위해 내려보낸다(서버에서도 막는다)
    is_d, why_d = is_design_user(me['username'], is_admin=(me['role'] == 'admin'))
    res['designer'] = {'ok': is_d, 'why': why_d, 'dept': get_user_dept(me['username'])[0]}
    return JSONResponse(res)


@app.post('/catia/upload')
def catia_upload(request: Request,
                 vehicle_code: str = Form(''), row_level: str = Form(''),
                 stage: str = Form(''), part_group: str = Form(''),
                 rel_paths: list[str] = Form([]),
                 files: list[UploadFile] = File(...)):
    """카티아 파일 다중 업로드. 품번·리비전·품명·EO·일자는 «파일명에서 자동 추출»한다.
       차종/열/부품군/단계만 화면에서 고른다 — 폴더를 파는 대신 열로 들고 있기 위해서.
       폴더째 올리면 rel_paths 에 원본 하위 경로가 같은 순서로 들어온다(어디서 왔는지 추적용)."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if not vehicle_code.strip():
        return JSONResponse({'error': '차종을 선택하세요.'}, status_code=400)

    added, dups, unparsed, skipped, blocked, total_size = [], [], [], [], [], 0
    for fi, uf in enumerate(files):
        if not uf or not uf.filename:
            continue
        base = os.path.basename(uf.filename)
        # 탐색기가 만든 부산물은 올리지 않는다(Thumbs.db 가 폴더마다 있다)
        if base.lower() in ('thumbs.db', 'desktop.ini', '.ds_store') or base.startswith('~$'):
            skipped.append(base)
            continue
        meta = parse_catia_filename(base)
        rel = rel_paths[fi] if fi < len(rel_paths) else ''
        src_dir = os.path.dirname(rel).strip('/')
        # 남이 체크아웃했거나 배포완료된 부품에는 새 리비전을 올릴 수 없다(PLM 기본 규칙).
        if meta['part_no']:
            ok, why = catia_can_modify(vehicle_code.strip(), meta['part_no'],
                                       me['username'], is_admin=(me['role'] == 'admin'))
            if not ok:
                blocked.append({'filename': base, 'part_no': meta['part_no'], 'reason': why})
                continue
        dup = find_catia_duplicate(vehicle_code.strip(), meta['part_no'], meta['rev'],
                                   meta['kind'], base, meta['ext'])
        if dup:
            # 폴더를 여러 겹으로 나눠 쓰면 같은 파일이 두 폴더에 들어 있는 일이 실제로 있다.
            # (NQ5 실측: 5.BACK COVER 와 8.BACK COVER 에 동일 파일이 MD5까지 같게 존재)
            dups.append({'filename': base, 'part_no': dup['part_no'], 'rev': dup['rev'],
                         'uploaded': dup['uploaded'], 'src_dir': src_dir})
            continue
        # 카티아 원본은 수십~수백 MB다. 통째로 메모리에 올리면 서버가 죽으므로 청크로 흘려 쓴다.
        safe = f"{uuid.uuid4().hex[:10]}_{re.sub(r'[^A-Za-z0-9._-]', '_', base)[-90:]}"
        dest = os.path.join(CATIA_DIR, safe)
        size = 0
        with open(dest, 'wb') as out:
            while True:
                chunk = uf.file.read(1024 * 1024)
                if not chunk:
                    break
                out.write(chunk); size += len(chunk)
        note = meta.get('note') or ''
        if src_dir:
            note = (note + ' / ' if note else '') + '원본 폴더: ' + src_dir
        meta.update(vehicle_code=vehicle_code.strip(), row_level=row_level.strip(),
                    stage=stage.strip(), part_group=part_group.strip() or 'ETC',
                    file_path=dest, size_no=size, note=note)
        add_catia_file(meta, me['username'])
        total_size += size
        added.append({'filename': base, 'part_no': meta['part_no'], 'rev': meta['rev'],
                      'part_name': meta['part_name'],
                      'kind': meta['kind'], 'ext': meta['ext'], 'parsed': meta['parsed'],
                      'src_dir': src_dir})
        if not meta['parsed']:
            unparsed.append(base)

    # 올린 품번을 품목 마스터에도 등록한다 — BOM 이 먼저든 도면이 먼저든 상관없게.
    # (이게 없어서 카티아 32품번 중 품목마스터 매칭이 0개였다)
    pm = upsert_parts_from_catia(
        [{'part_no': a['part_no'], 'part_name': a['part_name']} for a in added if a['part_no']],
        vehicle_code.strip(), me['username'])

    return JSONResponse({'ok': True, 'added': len(added), 'dup': len(dups),
                         'unparsed': len(unparsed), 'skipped': len(skipped),
                         'blocked': len(blocked), 'blocked_items': blocked,
                         'size': total_size, 'items': added, 'dups': dups,
                         'parts': pm,
                         'stats': get_catia_stats(vehicle_code.strip())})


# ── 체크아웃/체크인·수명주기 (PLM 기본 기능) ─────────────────────────────────
@app.post('/catia/checkout')
async def catia_checkout_route(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    b = await request.json()
    res = catia_checkout(str(b.get('vehicle') or ''), str(b.get('part_no') or ''),
                         me['username'], str(b.get('note') or '')[:200])
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=409)
    return JSONResponse(res)


@app.post('/catia/checkin')
async def catia_checkin_route(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    b = await request.json()
    res = catia_checkin(str(b.get('vehicle') or ''), str(b.get('part_no') or ''),
                        me['username'], str(b.get('comment') or '')[:300],
                        is_admin=(me['role'] == 'admin'), cancel=bool(b.get('cancel')))
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=409)
    return JSONResponse(res)


@app.post('/catia/state')
async def catia_state_route(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    b = await request.json()
    res = catia_set_state(str(b.get('vehicle') or ''), str(b.get('part_no') or ''),
                          str(b.get('state') or ''), me['username'],
                          str(b.get('comment') or '')[:300],
                          is_admin=(me['role'] == 'admin'))
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg')}, status_code=409)
    return JSONResponse(res)


@app.get('/catia/item')
async def catia_item_route(request: Request, vehicle: str = '', part_no: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'item': get_catia_item(vehicle, part_no),
                         'log': get_catia_item_log(vehicle, part_no),
                         'states': CATIA_STATES})



# ── 자동 변환 (사내 CATIA PC 의 대기 프로그램이 사용) ─────────────────────────
def _agent_ok(request: Request) -> bool:
    """변환 프로그램 전용 인증 — 사람 계정과 분리한다(무인 실행이라 비밀번호를 못 쓴다)."""
    key = request.headers.get('X-Agent-Key', '')
    return bool(key) and key == get_convert_agent_key()


@app.get('/catia/convert/queue')
async def catia_convert_queue(request: Request, limit: int = 20):
    """변환본이 아직 없는 원본 목록. 사내 CATIA PC 가 주기적으로 물어본다."""
    if not _agent_ok(request):
        redir = require_admin(request)
        if redir: return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    else:
        mark_convert_agent_seen()
    return JSONResponse({'items': get_convert_queue(min(limit, 100)),
                         'map': CONVERT_MAP, 'stats': get_convert_stats()})


@app.get('/catia/convert/source/{file_id}')
async def catia_convert_source(request: Request, file_id: int):
    """변환할 원본 내려받기."""
    if not _agent_ok(request):
        redir = require_login(request)
        if redir: return redir
    f = get_catia_file(file_id)
    if not f or not os.path.exists(f.get('file_path') or ''):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(f['file_path'], filename=f['filename'])


@app.post('/catia/convert/result/{file_id}')
def catia_convert_result(request: Request, file_id: int, file: UploadFile = File(...)):
    """변환본 올리기. 원본의 차종·품번·리비전·단계를 그대로 물려받아 같은 줄에 붙는다."""
    if not _agent_ok(request):
        return JSONResponse({'error': '권한이 없습니다.'}, status_code=403)
    src = get_catia_file(file_id)
    if not src:
        return JSONResponse({'error': '원본을 찾을 수 없습니다.'}, status_code=404)
    mark_convert_agent_seen()
    base = os.path.basename(file.filename or '')
    ext = os.path.splitext(base)[1].lower()
    want = CONVERT_MAP.get((src.get('ext') or '').lower())
    if want and ext != want:
        return JSONResponse({'error': f'{want} 파일이어야 합니다 (받은 것: {ext})'},
                            status_code=400)
    dup = find_catia_duplicate(src['vehicle_code'], src['part_no'], src['rev'],
                               src['kind'], base, ext)
    if dup:
        return JSONResponse({'ok': True, 'skipped': True, 'msg': '이미 변환본이 있습니다.'})
    safe = f"{uuid.uuid4().hex[:10]}_{re.sub(r'[^A-Za-z0-9._-]', '_', base)[-90:]}"
    dest = os.path.join(CATIA_DIR, safe)
    size = 0
    with open(dest, 'wb') as out:
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            out.write(chunk); size += len(chunk)
    meta = parse_catia_filename(base)
    # 파일명이 규칙 밖이어도 «원본의 정보»를 물려받아 반드시 같은 줄에 붙게 한다
    meta.update(part_no=src['part_no'], rev=src['rev'], rev_sort=src['rev_sort'],
                part_name=src['part_name'], part_type=src['part_type'], side=src['side'],
                eo_no=src['eo_no'], file_date=src['file_date'], parsed=1,
                vehicle_code=src['vehicle_code'], row_level=src['row_level'],
                stage=src['stage'], part_group=src['part_group'],
                kind=src['kind'], ext=ext, file_path=dest, size_no=size,
                note=f"자동 변환 (원본 {src['filename']})")
    add_catia_file(meta, 'convert-agent')
    return JSONResponse({'ok': True, 'size': size, 'stats': get_convert_stats()})


@app.get('/catia/convert/status')
async def catia_convert_status(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    out = {'stats': get_convert_stats()}
    if me['role'] == 'admin':
        out['agent_key'] = get_convert_agent_key()
    return JSONResponse(out)


@app.post('/catia/convert/rotate-key')
async def catia_convert_rotate(request: Request):
    redir = require_admin(request)
    if redir: return JSONResponse({'error': '관리자만 가능합니다.'}, status_code=403)
    return JSONResponse({'ok': True, 'agent_key': rotate_convert_agent_key()})


@app.get('/catia/viewer/{file_id}', response_class=HTMLResponse)
async def catia_viewer(request: Request, file_id: int):
    """도면 뷰어. PDF·이미지는 브라우저가 바로 열고, STEP·IGES 등은 브라우저에서
       OpenCASCADE(WASM)로 읽어 3D로 그린다. CATPart·CATDrawing 은 다쏘 전용 형식이라
       어떤 오픈소스도 못 읽으므로, STEP 으로 내보내 함께 올리라고 안내한다."""
    redir = require_login(request)
    if redir: return redir
    f = get_catia_file(file_id)
    if not f:
        return HTMLResponse('<h3 style="font-family:sans-serif;padding:40px">'
                            '파일을 찾을 수 없습니다.</h3>', status_code=404)
    return templates.TemplateResponse(request=request, name='viewer3d.html', context={'f': f})


@app.get('/catia/file/view/{file_id}')
async def catia_file_view(request: Request, file_id: int):
    redir = require_login(request)
    if redir: return redir
    f = get_catia_file(file_id)
    if not f or not f.get('file_path') or not os.path.exists(f['file_path']):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(f['file_path'], filename=f['filename'],
                        media_type='application/octet-stream')


@app.post('/catia/file/update/{file_id}')
async def catia_file_update(request: Request, file_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    body = await request.json()
    return JSONResponse(update_catia_file(file_id, body.get('fields') or {}))


@app.post('/catia/file/delete/{file_id}')
async def catia_file_delete(request: Request, file_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    f = get_catia_file(file_id)
    if not f:
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and f.get('uploaded_by') != me['username']:
        return JSONResponse({'error': '본인이 올린 파일 또는 관리자만 삭제할 수 있습니다.'},
                            status_code=403)
    # 남이 체크아웃했거나 배포완료된 부품의 파일은 지울 수 없다
    ok, why = catia_can_modify(f.get('vehicle_code') or '', f.get('part_no') or '',
                               me['username'], is_admin=(me['role'] == 'admin'))
    if not ok:
        return JSONResponse({'error': f'삭제할 수 없습니다 — {why}'}, status_code=409)
    return JSONResponse(delete_catia_file(file_id))


@app.get('/parts', response_class=HTMLResponse)
async def parts_page(request: Request, vehicle: str = '', q: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='parts.html', context={
        'me': me, 'vcodes': get_all_vehicle_codes(), 'sel_vehicle': vehicle,
        'sel_q': q, 'stats': get_parts_stats(),
    })


@app.get('/parts/list')
async def parts_list(request: Request, q: str = '', vehicle: str = '', level: str = '',
                     limit: int = 1000, offset: int = 0):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    res = search_parts(q.strip(), vehicle.strip().upper(), level.strip(),
                       min(limit, 3000), offset)
    # 카티아 게시판에 올라온 2D·3D 를 품번별로 붙여 준다 — 품목 목록에서 도면 유무가
    # 한눈에 보여야 하기 때문. 개발 품번(X접두)도 같은 부품으로 대조한다.
    counts = get_catia_counts([p['part_no'] for p in res['items']])
    # BOM 에 실제로 있는 품번인지 표시한다. 도면에서 등록된 개발 품번(X…)이 BOM 에 없으면
    # «X를 떼고 품번이 바뀐 것»일 수 있어 사용자가 찾아봐야 한다(사용자 제안 2026-08-02).
    bom = get_bom_part_numbers()
    for p in res['items']:
        p['catia'] = counts.get(p['part_no'], {'d2': 0, 'd3': 0, 'rev2': '', 'rev3': '',
                                               'catia_no': ''})
        pn = (p['part_no'] or '').upper()
        p['in_bom'] = (pn in bom) or (base_part_no(pn) in bom)
    res['stats'] = get_parts_stats()
    return JSONResponse(res)


@app.get('/parts/detail/{part_no}')
async def parts_detail(request: Request, part_no: str):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    p = get_part(part_no)
    if not p:
        return JSONResponse({'error': '품번을 찾을 수 없습니다.'}, status_code=404)
    # 카티아 게시판에 올라온 같은 부품의 2D·3D 를 함께 보여 준다(X 접두 개발품번 포함)
    catia = search_catia_parts(q=base_part_no(part_no))
    mine = [it for it in catia['items'] if base_part_no(it['part_no']) == base_part_no(part_no)]
    bom = get_bom_part_numbers()
    return JSONResponse({'ok': True, 'part': p, 'files': get_part_files(part_no),
                         'revs': get_part_revs(part_no), 'fields': PART_SPEC_FIELDS,
                         'catia': mine,
                         'in_bom': (part_no.upper() in bom) or (base_part_no(part_no) in bom)})


@app.post('/parts/save/{part_no}')
async def parts_save(request: Request, part_no: str):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    fields = {k: v for k, v in (body.get('fields') or {}).items() if k in PART_SPEC_FIELDS}
    res = update_part(part_no, fields, me['username'],
                      str(body.get('eo_no', ''))[:40], str(body.get('approval', '미상신'))[:20])
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg', '저장 실패')}, status_code=400)
    return JSONResponse(res)


@app.post('/parts/import')
def parts_import(request: Request,
                 vehicle_code: str = Form(...),
                 position: str = Form(''),
                 file: UploadFile = File(...)):
    """BOM 엑셀 → 전 레벨 품번·품명 자동 등록.
       기존 품목의 수기 입력 스펙은 덮어쓰지 않고 «빈 칸만» 채운다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls', '.xlsm'):
        return JSONResponse({'error': 'xlsx/xlsm/xls 파일만 지원합니다.'}, status_code=400)
    tmp = os.path.join(PART_FILE_DIR, f'_imp_{uuid.uuid4().hex[:10]}{ext}')
    with open(tmp, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    try:
        items = _parse_ebom_xlsx(tmp, position=position.strip())
        norm = [{'pno': it['pno'], 'part_name': it.get('description', ''), 'level': it.get('level')}
                for it in items if it.get('pno')]
        if not norm:
            return JSONResponse({'error': '품번을 찾지 못했습니다. BOM 구조를 확인하세요.'}, status_code=400)
        res = upsert_parts_bulk(norm, vehicle_code.strip().upper(), me['username'])
    except Exception as ex:
        return JSONResponse({'error': f'엑셀 처리 오류: {ex}'}, status_code=400)
    finally:
        try: os.unlink(tmp)
        except OSError: pass
    res['ok'] = True
    res['parsed'] = len(norm)
    res['stats'] = get_parts_stats()
    return JSONResponse(res)


@app.post('/parts/file/{part_no}')
def parts_file_upload(request: Request, part_no: str,
                      kind: str = Form('attach'), file: UploadFile = File(...)):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if not get_part(part_no):
        return JSONResponse({'error': '품번을 찾을 수 없습니다.'}, status_code=404)
    kind = 'drawing' if kind == 'drawing' else 'attach'
    ext = os.path.splitext(file.filename)[1].lower()
    if kind == 'drawing' and ext not in PART_DRAWING_EXTS:
        return JSONResponse({'error': f'도면은 {", ".join(PART_DRAWING_EXTS)} 만 가능합니다.'},
                            status_code=400)
    saved = os.path.join(PART_FILE_DIR, f'{uuid.uuid4().hex[:12]}{ext}')
    with open(saved, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    fid = add_part_file(part_no, kind, file.filename, saved, me['username'])
    return JSONResponse({'ok': True, 'id': fid, 'files': get_part_files(part_no)})


@app.get('/parts/file/view/{file_id}')
async def parts_file_view(request: Request, file_id: int):
    redir = require_login(request)
    if redir: return redir
    f = get_part_file(file_id)
    if not f or not os.path.exists(f['file_path']):
        return JSONResponse({'error': '파일이 없습니다.'}, status_code=404)
    return FileResponse(f['file_path'], filename=f['filename'])


@app.post('/parts/file/delete/{file_id}')
async def parts_file_delete(request: Request, file_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    f = get_part_file(file_id)
    if not f:
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and f['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    info = delete_part_file(file_id)
    if info:
        try:
            if os.path.exists(info['file_path']): os.unlink(info['file_path'])
        except OSError:
            pass
    return JSONResponse({'ok': True})


# ── E-BOM 시트 편집 게시판(2안) ───────────────────────────────────────────────
# 엑셀을 셀 단위로 DB에 적재해 웹에서 편집하고, 다운로드는 원본 워크북에 변경 셀만
# 덮어써 서식을 보존한다(alc2_convert.build_qpart_merge 에서 검증된 방식).
EBOM_SHEET_DIR = os.path.join(DATA_DIR, 'ebom_sheets')
os.makedirs(EBOM_SHEET_DIR, exist_ok=True)

# 서버 여유가 크지 않아(RAM 956MB, 서비스 상한 750MB) 적재 크기를 제한한다.
SHEET_MAX_ROWS, SHEET_MAX_COLS = 3000, 200

# 시트 처리 버전. 서식 추출이나 자동등록 규칙을 고치면 이 숫자를 올린다.
# 이미 등록된 시트는 열 때 이 값과 비교해 낮으면 원본 파일에서 자동으로 다시 뽑는다
# — 사용자가 재업로드할 필요가 없다. 원본을 file_path에 보관하기에 가능하다.
#   v1: 최초  /  v2: 열너비 min~max 범위 펼치기 + BOM 업로드 시 품목 자동등록
SHEET_PROC_VER = 2


def _refresh_sheet_processing(s: dict, username: str = '') -> dict:
    """구버전으로 처리된 시트를 원본 파일에서 다시 뽑아 갱신한다.
       **셀 값은 절대 건드리지 않는다** — 사용자가 편집한 내용이 들어 있기 때문.
       서식(layout)과 품목 자동등록만 다시 수행한다."""
    out = {'refreshed': False}
    if (s.get('proc_ver') or 0) >= SHEET_PROC_VER:
        return out
    path = s.get('file_path') or ''
    if not path or not os.path.exists(path):
        return out
    try:
        layout = _extract_layout(path, s.get('sheet_name') or '',
                                 s.get('n_rows') or 0, s.get('n_cols') or 0)
        set_ebom_sheet_layout(s['id'], json.dumps(layout, ensure_ascii=False), SHEET_PROC_VER)
        out['refreshed'] = True
        out['cols'] = len(layout.get('colw', {}))
    except Exception:
        return out
    # 품목 자동등록도 구버전에선 안 돌았으므로 이때 함께 채운다.
    try:
        items = _parse_ebom_xlsx(path)
        norm = [{'pno': it['pno'], 'part_name': it.get('description', ''), 'level': it.get('level')}
                for it in items if it.get('pno')]
        if norm:
            out['parts'] = upsert_parts_bulk(norm, s.get('vehicle_code') or '', username or 'system')
    except Exception:
        pass
    return out


def _load_sheet_cells(path: str):
    """엑셀 → [(row, col, value)]. read_only 스트리밍으로 읽어 메모리를 아낀다.
       반환: (sheet_name, n_rows, n_cols, cells)"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    cells, max_r, max_c = [], 0, 0
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri > SHEET_MAX_ROWS:
            break
        for ci, v in enumerate(row, start=1):
            if ci > SHEET_MAX_COLS:
                break
            if v is None:
                continue
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            s = str(v).strip()
            if not s:
                continue
            cells.append((ri, ci, s))
            if ri > max_r: max_r = ri
            if ci > max_c: max_c = ci
    name = ws.title
    wb.close()
    return name, max_r, max_c, cells


def _rgb(color):
    """openpyxl 색 → '#RRGGBB'. 테마색·인덱스색은 값을 특정할 수 없어 건너뛴다."""
    try:
        if color is None or color.type != 'rgb':
            return ''
        v = color.rgb
        if not isinstance(v, str) or len(v) != 8:
            return ''
        if v in ('00000000', 'FFFFFFFF'):
            return ''
        return '#' + v[2:]
    except Exception:
        return ''


def _extract_layout(path: str, sheet_name: str, max_r: int, max_c: int) -> dict:
    """엑셀 서식을 뽑아 화면을 원본과 똑같이 그리기 위한 정보로 만든다.
       열너비·행높이·병합·셀 스타일(배경/글꼴/정렬/테두리).
       셀마다 스타일을 통째로 담으면 커지므로 중복 제거해 {id:정의}+{셀:id}로 저장한다.
       스타일 읽기는 read_only 모드에서 불가능해 일반 모드로 한 번 더 연다(업로드 1회)."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # 열 너비는 «범위»로 저장된다 — <col min="2" max="5" width="2.33"/> 하나가 B~E를
    # 한꺼번에 지정한다. 열 문자로만 조회하면 범위 안의 중간 열(C·D·E 등)이 통째로
    # 빠져 기본값으로 그려지고 실제보다 훨씬 넓게 보인다(실측: 2.33→21px인데 96px).
    # 그래서 min~max를 펼쳐서 모든 열을 채운다.
    colw = {}
    for d in ws.column_dimensions.values():
        if not d.width:
            continue
        lo = max(1, int(d.min or 1))
        hi = min(max_c, int(d.max or lo))
        for c in range(lo, hi + 1):
            colw[c] = round(float(d.width), 2)
    # 지정이 없는 열은 시트 기본 너비를 쓴다(엑셀 기본 8.43자).
    default_w = ws.sheet_format.defaultColWidth or 8.43
    for c in range(1, max_c + 1):
        colw.setdefault(c, round(float(default_w), 2))

    # 행 높이도 같은 이유로 기본값을 함께 넘긴다.
    rowh = {}
    for r in range(1, max_r + 1):
        d = ws.row_dimensions.get(r)
        if d is not None and d.height:
            rowh[r] = round(float(d.height), 2)
    default_h = ws.sheet_format.defaultRowHeight or 15.0

    merges = []
    for m in ws.merged_cells.ranges:
        if m.min_row <= max_r and m.min_col <= max_c:
            merges.append([m.min_row, m.min_col, m.max_row, m.max_col])

    styles, cellstyle, sid_of = {}, {}, {}
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            bg = ''
            try:
                if cell.fill is not None and cell.fill.fill_type == 'solid':
                    bg = _rgb(cell.fill.start_color)
            except Exception:
                pass
            f = cell.font
            fc = _rgb(f.color) if f is not None else ''
            bold = bool(f.bold) if f is not None else False
            size = float(f.sz) if (f is not None and f.sz) else 0
            al = cell.alignment
            ha = (al.horizontal or '') if al is not None else ''
            va = (al.vertical or '') if al is not None else ''
            wrap = bool(al.wrap_text) if al is not None else False
            bd = ''
            try:
                b = cell.border
                bd = ''.join(('1' if (getattr(b, s) and getattr(b, s).style) else '0')
                             for s in ('top', 'right', 'bottom', 'left'))
                if bd == '0000':
                    bd = ''
            except Exception:
                pass
            if not (bg or fc or bold or ha or va or wrap or bd or (size and size != 11)):
                continue
            key = (bg, fc, bold, size, ha, va, wrap, bd)
            sid = sid_of.get(key)
            if sid is None:
                sid = str(len(sid_of) + 1)
                sid_of[key] = sid
                d = {}
                if bg: d['bg'] = bg
                if fc: d['fc'] = fc
                if bold: d['b'] = 1
                if size and size != 11: d['sz'] = size
                if ha: d['ha'] = ha
                if va: d['va'] = va
                if wrap: d['w'] = 1
                if bd: d['bd'] = bd
                styles[sid] = d
            cellstyle[f'{r}:{c}'] = sid
    wb.close()
    return {'colw': colw, 'rowh': rowh, 'merges': merges,
            'default_h': round(float(default_h), 2),
            'styles': styles, 'cellstyle': cellstyle}


@app.get('/ebom-sheet', response_class=HTMLResponse)
async def ebom_sheet_page(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='ebom_sheet.html', context={
        'me': me, 'vcodes': get_all_vehicle_codes(), 'sel_vehicle': vehicle,
        'stages': get_dev_stage_codes(),
    })


@app.get('/ebom-sheet/list')
async def ebom_sheet_list(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    items = get_ebom_sheets(vehicle)
    for it in items:
        st = get_ebom_sheet_lock_state(it['id'])
        it['lock'] = st
    return JSONResponse({'items': items})


@app.post('/ebom-sheet/upload')
def ebom_sheet_upload(request: Request,
                      vehicle_code: str = Form(...),
                      stage: str = Form(''),
                      title: str = Form(''),
                      file: UploadFile = File(...)):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xlsm'):
        return JSONResponse({'error': 'xlsx/xlsm 파일만 지원합니다. (셀 편집·서식 보존을 위해 xls는 제외)'},
                            status_code=400)
    fid = uuid.uuid4().hex[:12]
    saved = os.path.join(EBOM_SHEET_DIR, f'{fid}{ext}')
    with open(saved, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    try:
        sheet_name, n_rows, n_cols, cells = _load_sheet_cells(saved)
    except Exception as ex:
        try: os.unlink(saved)
        except Exception: pass
        return JSONResponse({'error': f'엑셀을 읽지 못했습니다: {ex}'}, status_code=400)
    if not cells:
        return JSONResponse({'error': '내용이 있는 셀을 찾지 못했습니다.'}, status_code=400)
    sid = add_ebom_sheet(vehicle_code.strip().upper(), stage.strip(), title.strip(),
                         file.filename, saved, sheet_name, n_rows, n_cols, me['username'])
    save_ebom_sheet_cells(sid, cells)
    # 엑셀 서식 추출 — 실패해도 값은 이미 저장됐으므로 서식 없이 동작하게 둔다.
    try:
        layout = _extract_layout(saved, sheet_name, n_rows, n_cols)
        set_ebom_sheet_layout(sid, json.dumps(layout, ensure_ascii=False))
        n_styles = len(layout.get('styles', {}))
    except Exception:
        n_styles = 0
    # BOM을 올리면 전 레벨 품번이 품목 DB에도 자동 등록된다(2D/3D 도면을 BOM과 함께
    # 관리하기 위함). 이미 있는 품목의 수기 스펙은 덮어쓰지 않고 빈 칸만 채운다.
    part_res = {}
    try:
        items = _parse_ebom_xlsx(saved)
        norm = [{'pno': it['pno'], 'part_name': it.get('description', ''), 'level': it.get('level')}
                for it in items if it.get('pno')]
        if norm:
            part_res = upsert_parts_bulk(norm, vehicle_code.strip().upper(), me['username'])
            part_res['parsed'] = len(norm)
    except Exception:
        part_res = {}
    return JSONResponse({'ok': True, 'id': sid, 'rows': n_rows, 'cols': n_cols,
                         'cells': len(cells), 'sheet_name': sheet_name, 'styles': n_styles,
                         'parts': part_res})


@app.get('/ebom-sheet/grid/{sheet_id}')
async def ebom_sheet_grid(request: Request, sheet_id: int, row_from: int = 1, row_to: int = 400,
                          rev: int = -1):
    """행 범위만 잘라서 내려준다 — 전체를 한 번에 보내면 브라우저·서버 모두 부담.

       rev 를 주면 «그 시점»을 보여 준다. 과거 시점은 읽기 전용이다 — 이미 지나간
       리비전을 고칠 수 있으면 이력이 뜻을 잃는다. 그래서 readonly 를 함께 내려
       화면이 편집을 막게 하고, 저장 요청도 서버에서 따로 거른다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    s = get_ebom_sheet(sheet_id)
    if not s:
        return JSONResponse({'error': '시트를 찾을 수 없습니다.'}, status_code=404)
    # 시스템이 개선된 뒤 처음 여는 구버전 시트면 원본에서 자동 갱신(재업로드 불필요)
    me = current_user(request)
    refreshed = _refresh_sheet_processing(s, (me or {}).get('username', ''))
    if refreshed.get('refreshed'):
        s = get_ebom_sheet(sheet_id)
    cur = s.get('current_rev') or 0
    past = (0 <= rev < cur)
    cells = (get_ebom_sheet_cells_at(sheet_id, rev, row_from, row_to) if past
             else get_ebom_sheet_cells(sheet_id, row_from, row_to))
    try:
        layout = json.loads(s.get('layout') or '{}')
    except ValueError:
        layout = {}
    s.pop('layout', None)          # 본문에 중복으로 싣지 않도록 제거
    return JSONResponse({'ok': True, 'sheet': s, 'lock': get_ebom_sheet_lock_state(sheet_id),
                         'row_from': row_from, 'row_to': row_to, 'layout': layout,
                         'refreshed': refreshed,
                         'rev': rev if past else cur, 'readonly': past,
                         'cells': [[r, c, v] for r, c, v in cells]})


@app.post('/ebom-sheet/lock/{sheet_id}')
async def ebom_sheet_lock(request: Request, sheet_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    res = acquire_ebom_sheet_lock(sheet_id, me['username'])
    return JSONResponse(res, status_code=200 if res.get('ok') else 409)


@app.post('/ebom-sheet/unlock/{sheet_id}')
async def ebom_sheet_unlock(request: Request, sheet_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    res = release_ebom_sheet_lock(sheet_id, me['username'], force=(me['role'] == 'admin'))
    return JSONResponse(res, status_code=200 if res.get('ok') else 403)


@app.post('/ebom-sheet/save/{sheet_id}')
async def ebom_sheet_save(request: Request, sheet_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    edits = body.get('edits') or []
    if not isinstance(edits, list):
        return JSONResponse({'error': '잘못된 요청입니다.'}, status_code=400)
    # 과거 리비전 화면에서 온 저장은 거른다. 화면에서도 막지만, 지나간 리비전을
    # 고칠 수 있으면 이력이 뜻을 잃으므로 서버에서 한 번 더 잠근다.
    base = body.get('rev')
    if base is not None:
        s0 = get_ebom_sheet(sheet_id)
        if s0 and int(base) != (s0.get('current_rev') or 0):
            return JSONResponse({'error': f'REV {base} 는 지나간 리비전이라 저장할 수 없습니다. '
                                          '최신 리비전에서 편집하세요.'}, status_code=409)
    res = apply_ebom_sheet_edits(sheet_id, me['username'], edits, str(body.get('note', ''))[:200])
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg', '저장 실패')}, status_code=409)
    # 시트에서 품번을 고치거나 새로 넣으면 품목 마스터에도 따라와야 한다.
    # (기존엔 업로드·재처리 때만 등록돼서 «편집»한 품번은 품목관리에 안 나타났다)
    try:
        s = get_ebom_sheet(sheet_id)
        if s and s.get('file_path'):
            cells = get_ebom_sheet_cells(sheet_id)
            norm = _parts_from_sheet_cells(cells)
            if norm:
                res['parts'] = upsert_parts_bulk(norm, s.get('vehicle_code') or '',
                                                 me['username'])
    except Exception as ex:
        res['parts_error'] = str(ex)[:150]
    return JSONResponse(res)


@app.get('/ebom-sheet/revs/{sheet_id}')
async def ebom_sheet_revs(request: Request, sheet_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    return JSONResponse({'items': get_ebom_sheet_revs(sheet_id)})


# 내려받기 파일을 만드는 동안 다른 요청이 겹쳐 들어오지 못하게 한다.
# openpyxl 은 서식까지 살리려면 워크북을 통째로 메모리에 올려야 해서 한 번에
# 수백 MB 를 쓴다. 느리다고 사용자가 버튼을 여러 번 누르면 그만큼 워크북이
# 동시에 쌓여 서버가 스왑으로 빠졌다(실측: 앱 507MB + 스왑 1.5GB, 응답 지연).
_SHEET_BUILD_LOCK = threading.Lock()


def _build_sheet_download(sheet_id: int, rev: Optional[int] = None) -> str:
    """원본 워크북을 열어 «바뀐 셀만» 덮어쓴다. 새로 만들지 않으므로 병합·색·열너비
       등 서식이 그대로 남는다. 변경 목록은 리비전에 쌓인 것을 순서대로 적용한다.

       같은 리비전이면 이미 만든 파일을 그대로 준다 — 내용이 같은데 매번 다시
       만들 이유가 없다. 리비전이 오르면 파일 이름이 바뀌므로 자동으로 새로 만든다."""
    import openpyxl
    s = get_ebom_sheet(sheet_id)
    if not s or not os.path.exists(s['file_path']):
        return ''
    want = (s['current_rev'] or 0) if rev is None else max(0, min(int(rev), s['current_rev'] or 0))
    out = os.path.join(REPORTS_DIR, f'EBOMSHEET_{sheet_id}_r{want}.xlsx')
    with _SHEET_BUILD_LOCK:
        # 잠금을 기다리는 사이 다른 요청이 이미 만들어 놨을 수 있으니 안에서 한 번 더 본다.
        if os.path.exists(out) and os.path.getmtime(out) >= os.path.getmtime(s['file_path']):
            return out
        wb = openpyxl.load_workbook(s['file_path'])
        try:
            ws = wb[s['sheet_name']] if s.get('sheet_name') in wb.sheetnames else wb.active
            # 원본은 손대지 않고 그대로 두므로 «원본 + (1..N)» 이 곧 REV N 시점이다.
            for (r, c), v in get_ebom_sheet_applied_changes(sheet_id, upto_rev=want).items():
                ws.cell(r, c).value = v
            wb.save(out)
        finally:
            # 워크북을 닫아야 수백 MB 가 곧바로 풀린다.
            wb.close()
    return out


@app.get('/ebom-sheet/download/{sheet_id}')
def ebom_sheet_download(request: Request, sheet_id: int, rev: int = -1):
    """엑셀 내려받기. rev 를 주면 «그 시점» 엑셀을 만든다(서식은 원본 그대로)."""
    redir = require_login(request)
    if redir: return redir
    s = get_ebom_sheet(sheet_id)
    if not s:
        return JSONResponse({'error': '시트를 찾을 수 없습니다.'}, status_code=404)
    cur = s.get('current_rev') or 0
    want = cur if rev < 0 else max(0, min(rev, cur))
    path = _build_sheet_download(sheet_id, want)
    if not path:
        return JSONResponse({'error': '원본 파일을 찾을 수 없습니다.'}, status_code=404)
    base = os.path.splitext(s['filename'])[0]
    return FileResponse(path, filename=f'{base}_REV{want}.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.post('/ebom-sheet/drop-rev/{sheet_id}')
async def ebom_sheet_drop_rev(request: Request, sheet_id: int):
    """마지막 리비전을 없던 일로 만든다(잘못 저장한 것 지우기)."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    res = drop_last_ebom_sheet_rev(sheet_id, me['username'], is_admin=(me['role'] == 'admin'))
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg', '지우지 못했습니다.')}, status_code=409)
    # 지운 번호가 다시 쓰일 수 있으므로 그 번호 이상의 «만들어 둔 엑셀»을 버린다.
    # 안 그러면 새 R3 를 받을 때 옛 R3 파일이 그대로 나간다.
    with _SHEET_BUILD_LOCK:
        for n in range(res['dropped'], (res['dropped'] or 0) + 200):
            p = os.path.join(REPORTS_DIR, f'EBOMSHEET_{sheet_id}_r{n}.xlsx')
            if os.path.exists(p):
                try: os.unlink(p)
                except OSError: pass
    return JSONResponse(res)


@app.post('/ebom-sheet/revert/{sheet_id}')
async def ebom_sheet_revert(request: Request, sheet_id: int):
    """REV 시점으로 되돌린다. 과거를 지우지 않고 새 리비전으로 쌓는다."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    try:
        rev = int(body.get('rev'))
    except (TypeError, ValueError):
        return JSONResponse({'error': '되돌릴 리비전을 지정하세요.'}, status_code=400)
    res = revert_ebom_sheet_to(sheet_id, me['username'], rev)
    if not res.get('ok'):
        return JSONResponse({'error': res.get('msg', '되돌리기 실패')}, status_code=409)
    return JSONResponse(res)


@app.post('/ebom-sheet/delete/{sheet_id}')
async def ebom_sheet_delete(request: Request, sheet_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    s = get_ebom_sheet(sheet_id)
    if not s:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and s['created_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    info = delete_ebom_sheet(sheet_id)
    if info and info.get('file_path'):
        try:
            if os.path.exists(info['file_path']): os.unlink(info['file_path'])
        except Exception:
            pass
    return JSONResponse({'ok': True})


def _ebom_compare_sources() -> list:
    """비교 대상 E-BOM 후보 — (차종,단계)별로 묶고 그 안의 열/위치별 최신본 수를 센다."""
    groups = {}
    for u in get_ebom_uploads():
        if not u.get('is_active', 1):
            continue
        key = (u.get('vehicle_code', ''), u.get('stage', ''))
        g = groups.setdefault(key, {'vehicle': key[0], 'stage': key[1],
                                    'rows': set(), 'revisions': set(), 'count': 0})
        g['count'] += 1
        if u.get('row_num'):
            g['rows'].add(u['row_num'])
        if u.get('revision'):
            g['revisions'].add(u['revision'])
    out = []
    for (veh, stage), g in sorted(groups.items()):
        out.append({
            'key': f'{veh}|{stage}', 'vehicle': veh, 'stage': stage, 'count': g['count'],
            'label': (f"{veh} / {stage or '-'} — 등록 {g['count']}건"
                      f"{' (열 ' + ','.join(sorted(g['rows'])) + ')' if g['rows'] else ''}"),
        })
    return out


def _mbom_compare_sources() -> list:
    """비교 대상 M-BOM 후보 — HKMC Q파트 & ALC 이력 게시글 중 ALC 파일이 있는 것."""
    out = []
    for r in get_mbom_posts_with_files():
        out.append({'id': r['id'], 'vehicle': r['vehicle_code'], 'revision': r['revision'],
                    'label': (f"{r['vehicle_code']} / {r['stage'] or '-'} / {r['revision']} — "
                              f"{(r['title'] or '')[:26]} (파일 {r['nfiles']}개)")})
    return out


@app.get('/ebom-mbom-compare', response_class=HTMLResponse)
async def ebom_mbom_compare_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='ebom_mbom_compare.html', context={
        'me': me,
        'ebom_sources': _ebom_compare_sources(),
        'mbom_sources': _mbom_compare_sources(),
    })


def _norm_pno(p: str) -> str:
    """품번 정규화 — 하이픈/공백 제거, 대문자 (88005 P1000 == 88005-P1000)"""
    return re.sub(r'[\s\-]', '', str(p or '')).upper()


def _parse_alc_partnos(path: str) -> dict:
    """열별 ALC 코드집에서 13자리 품번 추출 → {10자리: {'names':set, 'cccs':set, 'rows':n}}"""
    import pandas as pd
    df = pd.read_excel(path, header=None, sheet_name=0)
    pat = re.compile(r'^(\d{5})([A-Z0-9]{5})([A-Z0-9]{3})?$')
    out = {}
    for i in range(len(df)):
        raw = str(df.iat[i, 2]).strip() if df.shape[1] > 2 and pd.notna(df.iat[i, 2]) else ''
        m = pat.match(_norm_pno(raw))
        if not m:
            continue
        base = m.group(1) + m.group(2)
        ccc = m.group(3) or ''
        name = ''
        if df.shape[1] > 6 and pd.notna(df.iat[i, 6]):
            name = str(df.iat[i, 6]).strip()
        e = out.setdefault(base, {'names': set(), 'cccs': set(), 'rows': 0})
        e['rows'] += 1
        if ccc: e['cccs'].add(ccc)
        if name: e['names'].add(name)
    del df
    return out


COMPARE_RESULTS: dict = {}


def _ebom_mbom_compare_process(vehicle, stage, ebom_items, tmp_files):
    """무거운 파싱·엑셀 생성 부분 — 스레드풀에서 실행해 이벤트 루프를 막지 않도록 분리."""
    ebom = {}
    for it in ebom_items:
        base = _norm_pno(it.get('pno'))
        if len(base) == 10:
            ebom.setdefault(base, set()).add((it.get('description') or '').strip())

    mbom = {}
    file_names = []
    for filename, tmp in tmp_files:
        file_names.append(filename)
        part = _parse_alc_partnos(tmp)
        for base, e in part.items():
            m = mbom.setdefault(base, {'names': set(), 'cccs': set(), 'rows': 0, 'files': set()})
            m['rows'] += e['rows']; m['cccs'] |= e['cccs']; m['names'] |= e['names']
            m['files'].add(filename)

    if not mbom:
        return {'error': 'ALC 파일에서 품번을 찾지 못했습니다. 파일 형식을 확인하세요.'}

    def fmt(b): return b[:5] + '-' + b[5:]
    both   = sorted(set(ebom) & set(mbom))
    e_only = sorted(set(ebom) - set(mbom))
    m_only = sorted(set(mbom) - set(ebom))

    rows = []
    for b in both:
        rows.append({'pno': fmt(b), 'status': 'OK',
                     'ebom_name': ' / '.join(sorted(n for n in ebom[b] if n))[:60],
                     'mbom_name': ' / '.join(sorted(mbom[b]['names']))[:60],
                     'ccc_count': len(mbom[b]['cccs']),
                     'cccs': ','.join(sorted(mbom[b]['cccs'])[:15])})
    for b in e_only:
        rows.append({'pno': fmt(b), 'status': 'EBOM_ONLY',
                     'ebom_name': ' / '.join(sorted(n for n in ebom[b] if n))[:60],
                     'mbom_name': '', 'ccc_count': 0, 'cccs': ''})
    for b in m_only:
        rows.append({'pno': fmt(b), 'status': 'MBOM_ONLY', 'ebom_name': '',
                     'mbom_name': ' / '.join(sorted(mbom[b]['names']))[:60],
                     'ccc_count': len(mbom[b]['cccs']),
                     'cccs': ','.join(sorted(mbom[b]['cccs'])[:15])})

    # 결과 엑셀
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = 'E-BOM vs M-BOM'
    ws.append(['NO', '품번(10자리)', '판정', 'E-BOM 품명', 'M-BOM PART-NAME', 'CCC 수', 'CCC 목록'])
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', start_color='1A237E')
    fills = {'OK': None, 'EBOM_ONLY': PatternFill('solid', start_color='FFCDD2'),
             'MBOM_ONLY': PatternFill('solid', start_color='FFF9C4')}
    label = {'OK': '일치(OK)', 'EBOM_ONLY': 'E-BOM에만 존재 — 확인 필요', 'MBOM_ONLY': 'M-BOM에만 존재 — 확인 필요'}
    for i, r in enumerate(rows, 1):
        ws.append([i, r['pno'], label[r['status']], r['ebom_name'], r['mbom_name'], r['ccc_count'], r['cccs']])
        if fills[r['status']]:
            for c in ws[i + 1]: c.fill = fills[r['status']]
    result_id = uuid.uuid4().hex[:10]
    out = os.path.join(REPORTS_DIR, f'COMPARE_{result_id}.xlsx')
    wb.save(out)
    COMPARE_RESULTS[result_id] = out

    return {'ok': True, 'result_id': result_id,
            'vehicle': vehicle, 'stage': stage, 'files': file_names,
            'ebom_count': len(ebom), 'mbom_count': len(mbom),
            'ok_count': len(both), 'ebom_only': len(e_only), 'mbom_only': len(m_only),
            'rows': rows[:400], 'truncated': len(rows) > 400}


def _ebom_specs_for(vehicle: str, stage: str) -> tuple:
    """(차종,단계)의 열/위치별 최신 등록본에서 1레벨 품번 → 사양(N열)을 모은다.
       spec 컬럼이 비어 있으면(컬럼 추가 이전 등록본) file_path로 그 자리에서 재파싱해
       저장한다 — 별도 운영 절차 없이 채워지도록 하는 lazy backfill."""
    specs, meta, backfilled, no_spec = {}, {}, 0, []
    for u in get_ebom_uploads(vehicle_code=vehicle, stage=stage or None):
        if not u.get('is_active', 1):
            continue
        items = get_ebom_items(u['id'])
        lv1 = [it for it in items if (it.get('level') or 1) == 1 and it.get('pno')]
        if lv1 and not any((it.get('spec') or '').strip() for it in lv1):
            path = u.get('file_path') or ''
            if path and os.path.exists(path):
                try:
                    reparsed = _parse_ebom_xlsx(path, position=u.get('position', ''))
                    if reparsed and any(x.get('spec') for x in reparsed):
                        replace_ebom_items(u['id'], reparsed)
                        items = get_ebom_items(u['id'])
                        lv1 = [it for it in items if (it.get('level') or 1) == 1 and it.get('pno')]
                        backfilled += 1
                except Exception:
                    pass
        got = False
        for it in lv1:
            sp = (it.get('spec') or '').strip()
            if not sp:
                continue
            base = _norm_pno(it['pno'])[:10]
            if len(base) != 10:
                continue
            specs[base] = sp
            meta[base] = {'row_num': u.get('row_num') or '', 'position': u.get('position') or '',
                          'revision': u.get('revision') or ''}
            got = True
        if not got and lv1:
            no_spec.append(f"{u.get('row_num') or '-'}/{u.get('position') or '-'} ({u.get('filename','')[:24]})")
    return specs, meta, backfilled, no_spec


def _mbom_specs_for(post_id: int) -> tuple:
    """M-BOM 게시글의 ALC 파일들 → {10자리 품번: [PEL 코드]}. 슬롯명도 함께 돌려준다."""
    import alc2_convert
    out, slots = {}, []
    for f in get_mbom_files_by_post(post_id):
        path = f.get('file_path') or ''
        if not path or not os.path.exists(path):
            continue
        slot = f.get('slot') or ''
        if 'Q파트' in slot:           # 생산계획표라 품번·PEL 코드가 없다
            continue
        try:
            partno = alc2_convert.read_alc_partno(path)
            pels = alc2_convert.read_alc_pel(path)
        except Exception:
            continue
        slots.append(slot)
        for code, p13 in partno.items():
            base = _norm_pno(p13)[:10]
            if len(base) != 10:
                continue
            e = out.setdefault(base, set())
            e.update(pels.get(code, []))
    return out, slots


def _ebom_mbom_spec_compare(vehicle, stage, post_id):
    """축(옵션그룹) 단위 사양 대조. 무거운 파싱을 포함하므로 스레드풀에서 호출한다."""
    import spec_compare
    from bom_generator import load_pel_master

    vocab = validators_load_vocab()
    if not vocab:
        return {'error': 'PEL CODE 마스터를 읽을 수 없어 사양 비교를 할 수 없습니다.'}
    import validators as _v
    idx = _v._spec_index(vocab)
    master = load_pel_master(PEL_CODE_PATH).get('data', {})

    e_specs, e_meta, backfilled, no_spec = _ebom_specs_for(vehicle, stage)
    if not e_specs:
        return {'error': 'E-BOM 1레벨 사양(N열)을 찾지 못했습니다. 등록본에 N열 사양이 있는지 확인하세요.'}
    m_pels, slots = _mbom_specs_for(post_id)
    if not m_pels:
        return {'error': '선택한 M-BOM 게시글에서 ALC 품번을 찾지 못했습니다.'}

    both = sorted(set(e_specs) & set(m_pels))
    rows, all_axes = [], []
    for base in both:
        e_axes, unresolved = spec_compare.resolve_text(e_specs[base], vocab, idx)
        m_axes = spec_compare.resolve_pel_codes(m_pels[base], master, vocab)
        axis_results = spec_compare.compare_axes(e_axes, m_axes, unresolved)
        verdict, note = spec_compare.verdict_text(axis_results, unresolved)
        all_axes.append(axis_results)
        md = e_meta.get(base, {})
        rows.append({
            'pno': base[:5] + '-' + base[5:],
            'row_num': md.get('row_num', ''), 'position': md.get('position', ''),
            'ebom_rev': md.get('revision', ''),
            'ebom_spec': e_specs[base],
            'mbom_spec': ' + '.join(sorted({s for a in m_axes.values() for s in a})),
            'verdict': verdict, 'note': note,
            'axes': axis_results,
        })
    summary = spec_compare.summarize(all_axes)

    from collections import Counter
    vc = Counter(r['verdict'] for r in rows)
    return {
        'ok': True, 'vehicle': vehicle, 'stage': stage, 'slots': slots,
        'ebom_count': len(e_specs), 'mbom_count': len(m_pels),
        'matched': len(both),
        'ebom_only': len(set(e_specs) - set(m_pels)), 'mbom_only': len(set(m_pels) - set(e_specs)),
        'verdicts': dict(vc), 'summary': summary,
        'backfilled': backfilled, 'no_spec_uploads': no_spec,
        'rows': rows[:500], 'truncated': len(rows) > 500,
    }


def validators_load_vocab():
    import validators
    return validators.load_spec_vocab()


@app.post('/ebom-mbom-compare/run')
async def ebom_mbom_compare_run(request: Request):
    """E-BOM 1레벨 사양(N열) vs M-BOM ALC 코드집 사양 — 축(옵션그룹) 단위 대조"""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    from starlette.concurrency import run_in_threadpool
    form = await request.form()
    ebom_key = str(form.get('ebom_key', '')).strip()
    post_id = str(form.get('mbom_post_id', '')).strip()
    if not ebom_key or '|' not in ebom_key:
        return JSONResponse({'error': 'E-BOM을 선택하세요.'}, status_code=400)
    if not post_id.isdigit():
        return JSONResponse({'error': 'M-BOM을 선택하세요.'}, status_code=400)
    vehicle, stage = ebom_key.split('|', 1)

    result = await run_in_threadpool(_ebom_mbom_spec_compare, vehicle.strip().upper(),
                                     stage.strip(), int(post_id))
    if 'error' in result:
        return JSONResponse({'error': result['error']}, status_code=400)
    return JSONResponse(result)


@app.get('/ebom-mbom-compare/download/{result_id}')
async def ebom_mbom_compare_download(request: Request, result_id: str):
    redir = require_login(request)
    if redir: return redir
    if not re.fullmatch(r'[a-f0-9]{10}', result_id):
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    path = COMPARE_RESULTS.get(result_id)
    if not path or not os.path.exists(path):
        return JSONResponse({'error': '결과가 만료되었습니다.'}, status_code=404)
    return FileResponse(path, filename='EBOM_MBOM_비교결과.xlsx',
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.post('/country-codes/ppt/upload')
def country_ppt_upload(request: Request,
                              file: UploadFile = File(...),
                              note: str = Form('')):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    me = current_user(request)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.pptx', '.png', '.jpg', '.jpeg', '.gif', '.webp'):
        return JSONResponse({'error': 'pptx, png, jpg, jpeg, gif, webp 파일만 지원합니다.'}, status_code=400)

    existing = list_country_ppt_revisions()
    next_rev = (max((r['rev_num'] for r in existing), default=0)) + 1
    save_path = os.path.join(COUNTRY_PPT_DIR, f'rev_{next_rev:03d}{ext}')
    with open(save_path, 'wb') as f:
        shutil.copyfileobj(file.file, f)

    result = add_country_ppt_revision(
        filename=file.filename,
        file_path=save_path,
        file_ext=ext.lstrip('.'),
        uploaded_by=me['username'],
        note=(note or '').strip()
    )
    return JSONResponse({'ok': True, 'rev_num': result['rev_num']})


@app.get('/country-codes/ppt/{rev_id}/download')
async def country_ppt_download(request: Request, rev_id: int):
    redir = require_login(request)
    if redir: return redir
    info = get_country_ppt_revision(rev_id)
    if not info or not os.path.exists(info['file_path']):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(info['file_path'], filename=info['filename'])


@app.get('/country-codes/ppt/{rev_id}/view')
async def country_ppt_view(request: Request, rev_id: int):
    """이미지 파일 인라인 표시"""
    redir = require_login(request)
    if redir: return redir
    info = get_country_ppt_revision(rev_id)
    if not info or not os.path.exists(info['file_path']):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    ext = info.get('file_ext', '').lower()
    mime_map = {'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                'gif': 'image/gif', 'webp': 'image/webp'}
    if ext not in mime_map:
        return JSONResponse({'error': '이미지 파일이 아닙니다.'}, status_code=400)
    return FileResponse(info['file_path'], media_type=mime_map[ext])


@app.post('/country-codes/ppt/{rev_id}/delete')
async def country_ppt_delete(request: Request, rev_id: int):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    info = delete_country_ppt_revision(rev_id)
    if not info:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    fp = info.get('file_path')
    if fp and os.path.exists(fp):
        try: os.unlink(fp)
        except Exception: pass
    return JSONResponse({'ok': True})


# ── CCC 매트릭스 API ──────────────────────────────────────────────────────────
@app.get('/ccc/matrix')
async def ccc_matrix_get(request: Request, vehicle: str = '', stage: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not vehicle:
        return JSONResponse({'matrix': [], 'material_types': MATERIAL_TYPES})
    matrix = get_ccc_matrix(vehicle)   # 차종 단독 (단계 무관)
    return JSONResponse({'matrix': matrix, 'material_types': MATERIAL_TYPES})


@app.post('/ccc/matrix/save')
async def ccc_matrix_save(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    vehicle = str(body.get('vehicle_code', '')).strip()
    cells = body.get('cells', [])
    if not vehicle:
        return JSONResponse({'error': '차종을 선택하세요.'}, status_code=400)
    saved = 0
    for cell in cells:
        mat = str(cell.get('material_type', '')).strip()
        ctry = str(cell.get('country_code', '')).strip()
        ccc = str(cell.get('ccc_code', '')).strip()
        if mat and ctry:
            upsert_ccc_matrix(vehicle, mat, ctry, ccc, me['username'])
            saved += 1
    return JSONResponse({'ok': True, 'saved': saved})


@app.get('/api/ccc/matrix')
async def api_ccc_matrix(request: Request, vehicle: str = '', stage: str = ''):
    """영업단가 게시판에서 호출 — CCC 매트릭스 조회 (차종 단독)"""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    matrix = get_ccc_matrix(vehicle) if vehicle else []
    return JSONResponse({'matrix': matrix, 'material_types': MATERIAL_TYPES})


# ── 영업단가 v2 (매트릭스 형식) ────────────────────────────────────────────────
@app.get('/sales/price/v2', response_class=HTMLResponse)
async def sales_price_v2_page(request: Request, vehicle: str = '', stage: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    vcodes = get_all_vehicle_codes()
    return templates.TemplateResponse(request=request, name='sales_price.html',
                                      context={'me': me, 'vcodes': vcodes,
                                               'sel_vehicle': vehicle, 'sel_stage': stage})


@app.get('/sales/price/v2/data')
async def sales_price_v2_data(request: Request, vehicle: str = '', stage: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    # 연결은 차종 단독 — E-BOM은 (열,위치)별 최신, CCC는 차종별 최신 (단계 무관)
    ebom_items = get_ebom_items_by_vehicle(vehicle) if vehicle else []
    ccc_matrix = get_ccc_matrix(vehicle) if vehicle else []
    prices = get_sales_prices_v2(vehicle or None, None)
    country_codes = get_all_country_codes()
    return JSONResponse({
        'ebom_items': ebom_items,
        'ccc_matrix': ccc_matrix,
        'prices': prices,
        'country_codes': country_codes,
        'material_types': MATERIAL_TYPES,
    })


@app.post('/sales/price/v2/save')
async def sales_price_v2_save(request: Request):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    body = await request.json()
    errors = []
    saved = 0
    for row in body.get('rows', []):
        try:
            price_val = row.get('unit_price')
            if price_val is None or str(price_val).strip() == '':
                price_f = None
            else:
                price_f = float(str(price_val).replace(',', ''))
            upsert_sales_price_v2(
                vehicle_code=row['vehicle_code'],
                stage=row['stage'],
                part_no=row['part_no'],
                part_name=row.get('part_name', ''),
                material_type=row['material_type'],
                country_code=row.get('country_code', '-'),
                ccc_code=row.get('ccc_code', ''),
                unit_price=price_f,
                currency=row.get('currency', 'KRW'),
                effective_date=row.get('effective_date', ''),
                username=me['username'],
                compare_pno=row.get('compare_pno', '')
            )
            saved += 1
        except Exception as ex:
            errors.append(f"{row.get('part_no','?')}-{row.get('material_type','?')}-{row.get('country_code','?')}: {ex}")
    return JSONResponse({'ok': True, 'saved': saved, 'errors': errors})


# ── PEL 이력 관리 게시판 ───────────────────────────────────────────────────────
PEL_HISTORY_DIR = os.path.join(DATA_DIR, 'pel_history')
os.makedirs(PEL_HISTORY_DIR, exist_ok=True)


@app.get('/pel-history', response_class=HTMLResponse)
async def pel_history_page(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    vcodes = get_all_vehicle_codes()
    return templates.TemplateResponse(request=request, name='pel_history.html', context={
        'me': me, 'vcodes': vcodes, 'sel_vehicle': vehicle,
        'stages': get_dev_stage_codes(), 'column_divs': PEL_COLUMN_DIVS,
    })


@app.get('/pel-history/list')
async def pel_history_list(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not vehicle:
        return JSONResponse({'items': []})
    items = get_pel_history(vehicle)
    return JSONResponse({'items': items})


@app.post('/pel-history/upload')
def pel_history_upload(
    request: Request,
    vehicle_code: str = Form(...),
    stage: str = Form(''),
    column_div: str = Form(''),
    revision: str = Form('VER.1'),
    title: str = Form(...),
    description: str = Form(''),
    file: UploadFile = File(None),
):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)

    if not vehicle_code.strip():
        return JSONResponse({'error': '차종을 선택하세요.'}, status_code=400)
    if not title.strip():
        return JSONResponse({'error': '제목을 입력하세요.'}, status_code=400)

    filename = file_id = file_path = ''
    if file is not None and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        file_id = uuid.uuid4().hex[:16]
        file_path = os.path.join(PEL_HISTORY_DIR, f'{file_id}{ext}')
        with open(file_path, 'wb') as f:
            shutil.copyfileobj(file.file, f)
        filename = file.filename

    new_id = add_pel_history(
        vehicle_code.strip().upper(), stage.strip(), revision.strip(),
        title.strip(), description.strip(),
        filename, file_id, file_path, me['username'],
        column_div=column_div.strip()
    )
    return JSONResponse({'ok': True, 'id': new_id, 'uploaded_by': me['username']})


@app.post('/pel-history/update/{item_id}')
async def pel_history_update(
    request: Request, item_id: int,
    stage: str = Form(''),
    column_div: str = Form(''),
    revision: str = Form('VER.1'),
    title: str = Form(...),
    description: str = Form(''),
    file: UploadFile = File(None),
):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    item = get_pel_history_item(item_id)
    if not item:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and item['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 수정할 수 있습니다.'}, status_code=403)
    if not title.strip():
        return JSONResponse({'error': '제목을 입력하세요.'}, status_code=400)

    filename = file_id = file_path = None
    if file is not None and file.filename:
        # 새 파일 첨부 → 기존 파일 삭제 후 교체
        if item.get('file_path') and os.path.exists(item['file_path']):
            try: os.unlink(item['file_path'])
            except Exception: pass
        ext = os.path.splitext(file.filename)[1].lower()
        file_id = uuid.uuid4().hex[:16]
        file_path = os.path.join(PEL_HISTORY_DIR, f'{file_id}{ext}')
        with open(file_path, 'wb') as f:
            shutil.copyfileobj(file.file, f)
        filename = file.filename

    update_pel_history(
        item_id, stage.strip(), column_div.strip(), revision.strip(),
        title.strip(), description.strip(),
        filename=filename, file_id=file_id, file_path=file_path
    )
    return JSONResponse({'ok': True})


@app.get('/pel-history/download/{item_id}')
async def pel_history_download(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return redir
    item = get_pel_history_item(item_id)
    if not item or not item.get('file_path') or not os.path.exists(item['file_path']):
        return JSONResponse({'error': '첨부 파일이 없습니다.'}, status_code=404)
    return FileResponse(item['file_path'], filename=item['filename'])


@app.post('/pel-history/delete/{item_id}')
async def pel_history_delete(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    item = get_pel_history_item(item_id)
    if not item:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    # 관리자 또는 작성자 본인만 삭제 가능
    if me['role'] != 'admin' and item['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    info = delete_pel_history(item_id)
    if info and info.get('file_path') and os.path.exists(info['file_path']):
        try: os.unlink(info['file_path'])
        except Exception: pass
    return JSONResponse({'ok': True})


# ── PEL 사양변경 게시판 (부품사양서 → 사양수현황 그리드) ────────────────────────
PEL_SPEC_DIR = os.path.join(DATA_DIR, 'pel_spec')
os.makedirs(PEL_SPEC_DIR, exist_ok=True)


def _pt_mark(v) -> bool:
    return str(v).strip() in ('○', '●', 'O', 'o', '*', 'V', 'v', '√', '1')


def _extract_my_code(part_path: str) -> str:
    """부품사양서에서 '차종년식 : XX YY' 공장구분 코드 추출 (행/열 위치 무관하게 검색)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(part_path, data_only=True)
        ws = wb.active
        for r in range(1, min(14, ws.max_row + 1)):
            for c in range(1, min(24, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if v and isinstance(v, str):
                    m = re.search(r'차종\s*년?식\s*[:：]\s*([0-9A-Za-z]+\s+[0-9A-Za-z]+)', v)
                    if m:
                        wb.close()
                        return m.group(1).strip()
        wb.close()
    except Exception:
        pass
    return ''


ROW_LEVEL_OPTIONS = ['1열 LH', '1열 RH', '2열', '3열']


def _detect_row_level(filename: str) -> str:
    """파일명으로 열구분(1열LH/RH, 2열, 3열) 자동 추정 — 못 찾으면 빈 문자열(수동 선택)."""
    s = (filename or '').upper()
    if '3열' in s or '3RD' in s or 'THIRD' in s:
        return '3열'
    if '2열' in s or 'RR' in s or 'SECOND' in s or 'REAR' in s:
        return '2열'
    if 'FRT' in s or '1열' in s:
        if 'DRV' in s or 'DRIVER' in s or 'LH' in s:
            return '1열 LH'
        if 'PASS' in s or 'PASSENGER' in s or 'RH' in s:
            return '1열 RH'
    return ''


def _build_pt_map(part_path: str) -> dict:
    """VC → 기본차/환경차/공용 (BASE/EV1 열 스캔)."""
    import openpyxl
    pt_map = {}
    try:
        wb = openpyxl.load_workbook(part_path, data_only=True)
        ws = wb.active
        base_col = ev_col = None
        for r in range(1, min(14, ws.max_row + 1)):
            for c in range(1, min(70, ws.max_column + 1)):
                s = str(ws.cell(r, c).value or '').strip().upper()
                if s == 'BASE': base_col = c
                elif s in ('EV1', 'EV'): ev_col = c
            if base_col and ev_col: break
        if base_col or ev_col:
            for r in range(1, ws.max_row + 1):
                vraw = ws.cell(r, 1).value
                if vraw is None: continue
                vc = str(vraw).strip()
                if not vc or not vc[0].isdigit(): continue
                b = _pt_mark(ws.cell(r, base_col).value) if base_col else False
                e = _pt_mark(ws.cell(r, ev_col).value) if ev_col else False
                pt_map[vc] = '공용' if (b and e) else ('기본차' if b else ('환경차' if e else ''))
        wb.close()
    except Exception:
        pass
    return pt_map


def _transform_pel_spec_multi(sources: list) -> dict:
    """여러 부품사양서(공장별)를 하나의 사양수현황 그리드로 병합.
       sources: [{'path':..., 'factory':...}] — factory가 각 행의 공장 태그로 들어감.
       열은 모든 소스의 사양 합집합(옵션그룹/표시순서 정렬)."""
    from bom_generator import parse_part_spec, load_pel_master
    master = load_pel_master(PEL_CODE_PATH).get('data', {})
    col_defs, all_rows = {}, []
    for src in sources:
        path = src.get('path'); factory = src.get('factory', '') or ''
        try:
            spec = parse_part_spec(path)
        except Exception:
            continue
        pt_map = _build_pt_map(path)
        for vcrow in spec.get('vcs', []):
            marks, spec_names = set(), []
            for o in vcrow.get('opts', []):
                m = master.get(str(o.get('pel_code', '')).strip())
                if not m: continue
                sp = str(m.get('사양', '')).strip()
                if not sp: continue
                grp = str(m.get('옵션그룹', '')).strip() or 'OPTION'
                try: order = float(m.get('표시순서') or 9999)
                except Exception: order = 9999.0
                if sp not in col_defs or order < col_defs[sp]['order']:
                    col_defs[sp] = {'group': grp, 'order': order}
                marks.add(sp); spec_names.append(sp)
            all_rows.append({
                'vc': vcrow.get('vc', ''), 'region': vcrow.get('region', ''),
                'powertrain': pt_map.get(str(vcrow.get('vc', '')).strip(), ''),
                'factory': factory,
                'spec_text': '+'.join(spec_names),
                'marks': sorted(marks),
            })
    columns = [{'spec': sp, 'group': d['group'], 'order': d['order']}
               for sp, d in sorted(col_defs.items(), key=lambda kv: (kv[1]['order'], kv[0]))]
    groups = []
    for c in columns:
        if groups and groups[-1]['group'] == c['group']:
            groups[-1]['span'] += 1
        else:
            groups.append({'group': c['group'], 'span': 1})
    return {'columns': columns, 'groups': groups, 'rows': all_rows, 'vc_count': len(all_rows)}


def _transform_pel_spec(part_path: str, factory: str = '') -> dict:
    """단일 부품사양서 → 사양수현황 그리드."""
    return _transform_pel_spec_multi([{'path': part_path, 'factory': factory}])


# ── 게시판별 "사용 방법" 안내 이미지 (범용 — board 식별자로 여러 게시판 재사용 가능) ─────
BOARD_GUIDE_IMG_EXTS = ('.png', '.jpg', '.jpeg', '.gif', '.webp')
BOARD_GUIDE_IMG_MIME = {'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                        'gif': 'image/gif', 'webp': 'image/webp'}
BOARD_GUIDE_ALLOWED = {'pel_spec', 'bom_generate'}  # 다른 게시판에 적용할 땐 여기에 식별자만 추가
BOARD_GUIDE_DIR = os.path.join(DATA_DIR, 'board_guide_images')
os.makedirs(BOARD_GUIDE_DIR, exist_ok=True)


@app.post('/board-guide/{board}/upload')
async def board_guide_upload(request: Request, board: str, file: UploadFile = File(...)):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    if board not in BOARD_GUIDE_ALLOWED:
        return JSONResponse({'error': '잘못된 게시판입니다.'}, status_code=400)
    me = current_user(request)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in BOARD_GUIDE_IMG_EXTS:
        return JSONResponse({'error': 'png, jpg, jpeg, gif, webp 이미지 파일만 지원합니다.'}, status_code=400)
    save_path = os.path.join(BOARD_GUIDE_DIR, f'{board}_{uuid.uuid4().hex[:12]}{ext}')
    with open(save_path, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    old = get_board_guide_image(board)
    set_board_guide_image(board, filename=file.filename, file_path=save_path,
                          file_ext=ext.lstrip('.'), uploaded_by=me['username'])
    if old and old.get('file_path') and os.path.exists(old['file_path']):
        try: os.unlink(old['file_path'])
        except Exception: pass
    return JSONResponse({'ok': True})


@app.get('/board-guide/{board}/view')
async def board_guide_view(request: Request, board: str):
    redir = require_login(request)
    if redir: return redir
    info = get_board_guide_image(board)
    if not info or not os.path.exists(info['file_path']):
        return JSONResponse({'error': '등록된 이미지가 없습니다.'}, status_code=404)
    ext = info.get('file_ext', '').lower()
    if ext not in BOARD_GUIDE_IMG_MIME:
        return JSONResponse({'error': '이미지 파일이 아닙니다.'}, status_code=400)
    return FileResponse(info['file_path'], media_type=BOARD_GUIDE_IMG_MIME[ext])


@app.post('/board-guide/{board}/reset')
async def board_guide_reset(request: Request, board: str):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    if board not in BOARD_GUIDE_ALLOWED:
        return JSONResponse({'error': '잘못된 게시판입니다.'}, status_code=400)
    info = clear_board_guide_image(board)
    if info and info.get('file_path') and os.path.exists(info['file_path']):
        try: os.unlink(info['file_path'])
        except Exception: pass
    return JSONResponse({'ok': True})


@app.get('/pel-spec', response_class=HTMLResponse)
async def pel_spec_page(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    guide_image = get_board_guide_image('pel_spec')
    return templates.TemplateResponse(request=request, name='pel_spec.html', context={
        'me': me, 'vcodes': get_all_vehicle_codes(), 'sel_vehicle': vehicle,
        'guide_image': guide_image,
    })


@app.get('/pel-spec/list')
async def pel_spec_list(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not vehicle:
        return JSONResponse({'items': []})
    return JSONResponse({'items': get_pel_spec_list(vehicle)})


FACTORY_OPTIONS = ['공통', '광주', '화성']


@app.post('/pel-spec/detect-code')
def pel_spec_detect_code(request: Request, file: UploadFile = File(...)):
    """업로드 전 미리보기 — 파일에서 차종년식 코드 추출(공장 자동제안용) + 파일명 기반 열구분 추정."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    ext = os.path.splitext(file.filename)[1].lower()
    tmp = os.path.join(PEL_SPEC_DIR, f'_tmp_{uuid.uuid4().hex[:10]}{ext}')
    with open(tmp, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    try:
        code = _extract_my_code(tmp)
    finally:
        try: os.unlink(tmp)
        except Exception: pass
    return JSONResponse({'my_code': code, 'row_level': _detect_row_level(file.filename)})


@app.get('/pel-spec/row-levels')
async def pel_spec_row_levels(request: Request, vehicle: str = ''):
    """차종에 업로드된 열구분 목록 (통합 그리드 탭 구성용)."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not vehicle:
        return JSONResponse({'row_levels': []})
    return JSONResponse({'row_levels': get_pel_spec_row_levels(vehicle)})


@app.post('/pel-spec/upload')
def pel_spec_upload(
    request: Request,
    vehicle_code: str = Form(...),
    powertrain: str = Form('전체'),
    factory: str = Form('공통'),
    row_level: str = Form(''),
    revision: str = Form('VER.1'),
    title: str = Form(...),
    description: str = Form(''),
    file: UploadFile = File(...),
):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if not vehicle_code.strip() or not title.strip():
        return JSONResponse({'error': '차종과 제목은 필수입니다.'}, status_code=400)
    factory = factory.strip() or '공통'
    if factory not in FACTORY_OPTIONS:
        factory = '공통'
    row_level = row_level.strip()
    if row_level not in ROW_LEVEL_OPTIONS:
        row_level = ''
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return JSONResponse({'error': 'xlsx/xls 파일만 업로드 가능합니다.'}, status_code=400)
    file_id = uuid.uuid4().hex[:16]
    saved_path = os.path.join(PEL_SPEC_DIR, f'{file_id}{ext}')
    with open(saved_path, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    my_code = _extract_my_code(saved_path)
    if not row_level:
        row_level = _detect_row_level(file.filename)
    # 업로드 즉시 변환 시도 (오류나도 저장은 유지)
    try:
        grid = _transform_pel_spec(saved_path, factory)
        vc_count = grid['vc_count']; col_count = len(grid['columns'])
        msg = f'{vc_count}개 VC · {col_count}개 사양열 변환됨 · 공장={factory}'
    except Exception as ex:
        vc_count = 0; col_count = 0
        msg = f'파일 저장됨 (변환 오류: {ex})'
    new_id = add_pel_spec(vehicle_code.strip().upper(), powertrain.strip() or '전체',
                          revision.strip() or 'VER.1', title.strip(), description.strip(),
                          file.filename, file_id, saved_path, me['username'],
                          factory=factory, my_code=my_code, row_level=row_level)
    return JSONResponse({'ok': True, 'id': new_id, 'uploaded_by': me['username'],
                         'vc_count': vc_count, 'col_count': col_count,
                         'factory': factory, 'my_code': my_code, 'row_level': row_level, 'message': msg})


@app.get('/pel-spec/grid/{item_id}')
def pel_spec_grid(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    item = get_pel_spec(item_id)
    if not item or not item.get('file_path') or not os.path.exists(item['file_path']):
        return JSONResponse({'error': '원본 파일이 없습니다.'}, status_code=404)
    try:
        grid = _transform_pel_spec(item['file_path'], item.get('factory', ''))
        grid['meta'] = {'vehicle': item['vehicle_code'], 'powertrain': item['powertrain'],
                        'factory': item.get('factory', ''), 'my_code': item.get('my_code', ''),
                        'row_level': item.get('row_level', ''),
                        'revision': item['revision'], 'title': item['title'],
                        'filename': item['filename'], 'merged': False}
        return JSONResponse(grid)
    except Exception as ex:
        import traceback
        return JSONResponse({'error': f'변환 오류: {ex}', 'trace': traceback.format_exc()}, status_code=500)


@app.get('/pel-spec/grid-merged/{vehicle}')
def pel_spec_grid_merged(request: Request, vehicle: str, row_level: str = ''):
    """차종(및 열구분)의 공장별 최신 PEL을 하나의 사양수현황 그리드로 병합."""
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    latest = get_pel_spec_latest_by_factory(vehicle, row_level)
    sources = [{'path': it['file_path'], 'factory': it.get('factory', '') or '공통'}
               for it in latest if it.get('file_path') and os.path.exists(it['file_path'])]
    if not sources:
        return JSONResponse({'error': '병합할 PEL이 없습니다.'}, status_code=404)
    try:
        grid = _transform_pel_spec_multi(sources)
        parts = [f"{it.get('factory','공통')} {it['revision']}({it.get('my_code','') or '-'})" for it in latest]
        title = f'{vehicle} {row_level} 통합' if row_level else f'{vehicle} 통합'
        grid['meta'] = {'vehicle': vehicle, 'merged': True, 'row_level': row_level,
                        'sources': parts, 'factory_count': len(sources),
                        'title': title, 'revision': '통합'}
        return JSONResponse(grid)
    except Exception as ex:
        import traceback
        return JSONResponse({'error': f'변환 오류: {ex}', 'trace': traceback.format_exc()}, status_code=500)


def _extract_header_style(part_path):
    """원본 부품사양서의 고정틀 헤더 색을 추출 (UPG VC 헤더셀 기준).
       반환: {'fill', 'gfill'(그룹헤더용 약간 진하게), 'font', 'bold'}."""
    style = {'fill': '1A237E', 'gfill': '283593', 'font': 'FFFFFF', 'bold': True}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(part_path)
        ws = wb.active
        for r in range(1, 13):
            for c in range(1, 7):
                if str(ws.cell(r, c).value or '').strip() == 'UPG VC':
                    cell = ws.cell(r, c)
                    rgb = getattr(getattr(cell.fill, 'fgColor', None), 'rgb', None)
                    if isinstance(rgb, str) and len(rgb) >= 6 and rgb[-6:] != '000000':
                        style['fill'] = rgb[-6:]
                        style['gfill'] = _darken_hex(rgb[-6:], 0.85)
                    frgb = getattr(getattr(cell.font, 'color', None), 'rgb', None)
                    if isinstance(frgb, str) and len(frgb) >= 6:
                        style['font'] = frgb[-6:]
                    style['bold'] = bool(cell.font.bold)
                    wb.close()
                    return style
        wb.close()
    except Exception:
        pass
    return style


def _darken_hex(hex6, factor):
    try:
        r = int(int(hex6[0:2], 16) * factor)
        g = int(int(hex6[2:4], 16) * factor)
        b = int(int(hex6[4:6], 16) * factor)
        return f'{r:02X}{g:02X}{b:02X}'
    except Exception:
        return hex6


def _pel_row_val(r, col_id):
    if col_id == 'factory': return r.get('factory', '') or '(없음)'
    if col_id == 'powertrain': return r.get('powertrain', '') or '(없음)'
    if col_id == 'region': return r.get('region', '') or '(없음)'
    if col_id.startswith('opt:'): return '●' if col_id[4:] in set(r.get('marks', [])) else '(없음)'
    return ''


def _pel_filter_generic(rows, filters_json, q):
    """웹 오토필터와 동일 — filters: {colId: [허용값...]} + 텍스트검색."""
    try:
        filters = json.loads(filters_json) if filters_json else {}
    except Exception:
        filters = {}
    ql = (q or '').strip().lower()
    out = []
    for r in rows:
        ok = True
        for col_id, allowed in filters.items():
            if _pel_row_val(r, col_id) not in allowed:
                ok = False; break
        if not ok:
            continue
        if ql:
            s = f"{r.get('vc','')} {r.get('region','')} {r.get('spec_text','')}".lower()
            if ql not in s:
                continue
        out.append(r)
    return out


def _pel_grid_to_excel(grid, filename, style=None):
    """사양수현황 그리드 dict → 엑셀 (공장 열 포함, 그룹헤더 병합, ●).
       style: 원본 부품사양서에서 추출한 고정틀 색(없으면 시스템 기본)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    st = style or {'fill': '1A237E', 'gfill': '283593', 'font': 'FFFFFF', 'bold': True}
    wb = Workbook(); ws = wb.active; ws.title = '사양수현황'
    center = Alignment(horizontal='center', vertical='center')
    hfill = PatternFill('solid', start_color=st['fill'])
    hfont = Font(bold=st.get('bold', True), color=st['font'])
    gfill = PatternFill('solid', start_color=st['gfill'])
    fixed = ['NO', 'VC', '공장', '파워트레인', '지역', 'SPEC']
    ws.append(fixed + sum([[g['group']] + [''] * (g['span'] - 1) for g in grid['groups']], []))
    ws.append([''] * len(fixed) + [c['spec'] for c in grid['columns']])
    for cell in ws[1]:
        cell.fill = gfill; cell.font = hfont; cell.alignment = center
    for cell in ws[2]:
        cell.fill = hfill; cell.font = hfont; cell.alignment = center
    cidx = len(fixed) + 1
    for g in grid['groups']:
        if g['span'] > 1:
            ws.merge_cells(start_row=1, start_column=cidx, end_row=1, end_column=cidx + g['span'] - 1)
        cidx += g['span']
    for i in range(1, len(fixed) + 1):
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
        ws.cell(1, i).value = fixed[i - 1]; ws.cell(1, i).fill = hfill
    for i, r in enumerate(grid['rows'], 1):
        base = [i, r['vc'], r.get('factory', ''), r['powertrain'], r['region'], r['spec_text']]
        markset = set(r['marks'])
        line = base + ['●' if c['spec'] in markset else '' for c in grid['columns']]
        ws.append(line)
        for c in range(len(fixed) + 1, len(line) + 1):
            ws.cell(i + 2, c).alignment = center
    out_path = os.path.join(REPORTS_DIR, f'PELSPEC_{uuid.uuid4().hex[:10]}.xlsx')
    wb.save(out_path)
    return FileResponse(out_path, filename=filename,
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.get('/pel-spec/download/{item_id}')
def pel_spec_download(request: Request, item_id: int, mode: str = 'grid',
                            filters: str = '', q: str = ''):
    redir = require_login(request)
    if redir: return redir
    item = get_pel_spec(item_id)
    if not item or not item.get('file_path') or not os.path.exists(item['file_path']):
        return JSONResponse({'error': '원본 파일이 없습니다.'}, status_code=404)
    if mode == 'original':
        return FileResponse(item['file_path'], filename=item['filename'])
    try:
        grid = _transform_pel_spec(item['file_path'], item.get('factory', ''))
    except Exception as ex:
        return JSONResponse({'error': f'변환 오류: {ex}'}, status_code=500)
    grid = dict(grid)
    grid['rows'] = _pel_filter_generic(grid['rows'], filters, q)
    base = os.path.splitext(item['filename'])[0]
    return _pel_grid_to_excel(grid, f'{base}_사양수현황.xlsx',
                              style=_extract_header_style(item['file_path']))


@app.get('/pel-spec/download-merged/{vehicle}')
def pel_spec_download_merged(request: Request, vehicle: str, filters: str = '', q: str = '', row_level: str = ''):
    redir = require_login(request)
    if redir: return redir
    latest = get_pel_spec_latest_by_factory(vehicle, row_level)
    sources = [{'path': it['file_path'], 'factory': it.get('factory', '') or '공통'}
               for it in latest if it.get('file_path') and os.path.exists(it['file_path'])]
    if not sources:
        return JSONResponse({'error': '병합할 PEL이 없습니다.'}, status_code=404)
    grid = _transform_pel_spec_multi(sources)
    grid['rows'] = _pel_filter_generic(grid['rows'], filters, q)
    suffix = f'_{row_level}' if row_level else ''
    return _pel_grid_to_excel(grid, f'{vehicle}{suffix}_통합_사양수현황.xlsx',
                              style=_extract_header_style(sources[0]['path']))


@app.post('/pel-spec/delete/{item_id}')
async def pel_spec_delete(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    item = get_pel_spec(item_id)
    if not item:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and item['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    info = delete_pel_spec(item_id)
    if info and info.get('file_path') and os.path.exists(info['file_path']):
        try: os.unlink(info['file_path'])
        except Exception: pass
    return JSONResponse({'ok': True})


# ── 영업 단가 원본 파일 (차종별 리비전 게시판) ─────────────────────────────────────
SALES_FILE_DIR = os.path.join(DATA_DIR, 'sales_files')
os.makedirs(SALES_FILE_DIR, exist_ok=True)


@app.get('/sales/price/files', response_class=HTMLResponse)
async def sales_files_page(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    return templates.TemplateResponse(request=request, name='sales_files.html', context={
        'me': me, 'vcodes': get_all_vehicle_codes(), 'sel_vehicle': vehicle,
    })


@app.get('/sales/price/files/list')
async def sales_files_list(request: Request, vehicle: str = ''):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not vehicle:
        return JSONResponse({'items': []})
    return JSONResponse({'items': get_sales_file_list(vehicle)})


@app.post('/sales/price/files/upload')
def sales_files_upload(
    request: Request,
    vehicle_code: str = Form(...),
    powertrain: str = Form('전체'),
    revision: str = Form('VER.1'),
    title: str = Form(...),
    description: str = Form(''),
    file: UploadFile = File(...),
):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    if not vehicle_code.strip() or not title.strip():
        return JSONResponse({'error': '차종과 제목은 필수입니다.'}, status_code=400)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls', '.xlsm'):
        return JSONResponse({'error': 'xlsx/xls 파일만 업로드 가능합니다.'}, status_code=400)
    file_id = uuid.uuid4().hex[:16]
    saved_path = os.path.join(SALES_FILE_DIR, f'{file_id}{ext}')
    with open(saved_path, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    new_id = add_sales_file(vehicle_code.strip().upper(), powertrain.strip() or '전체',
                            revision.strip() or 'VER.1', title.strip(), description.strip(),
                            file.filename, file_id, saved_path, me['username'])
    return JSONResponse({'ok': True, 'id': new_id, 'uploaded_by': me['username'],
                         'message': '업로드 완료'})


# 영업단가 시트 열 매핑 (1-indexed) — KMC 품의 자료 서식
SALES_SHEET_COLS = {'no': 1, 'spec': 2, 'pno': 3, 'cmp': 4, 'color': 5,
                    'jeonga': 6, 'sagup': 7, 'hap': 8, 'after': 9,
                    'bigo': 10, 'sangak': 12, 'upcode': 15}
SALES_DATA_START = 8


def _sales_num(v):
    if v is None or str(v).strip() == '':
        return None
    try:
        f = float(str(v).replace(',', ''))
        return int(f) if f == int(f) else f
    except Exception:
        return None


def _parse_sales_sheet(path):
    """KMC 단가 집계표의 데이터 영역(8행~)을 파싱. A열이 숫자인 행만."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    C = SALES_SHEET_COLS
    rows = []
    for r in range(SALES_DATA_START, ws.max_row + 1):
        a = ws.cell(r, C['no']).value
        if not isinstance(a, (int, float)):
            continue
        def g(k):
            return ws.cell(r, C[k]).value
        rows.append({
            'r': r, 'no': a,
            'spec': g('spec') or '', 'pno': g('pno') or '', 'cmp': g('cmp') or '',
            'color': g('color') or '', 'jeonga': _sales_num(g('jeonga')),
            'sagup': _sales_num(g('sagup')), 'hap': _sales_num(g('hap')),
            'after': _sales_num(g('after')), 'bigo': g('bigo') or '',
            'sangak': _sales_num(g('sangak')), 'upcode': g('upcode') or '',
        })
    return rows


@app.get('/sales/price/files/sheet/{item_id}')
def sales_files_sheet(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    item = get_sales_file(item_id)
    if not item or not item.get('file_path') or not os.path.exists(item['file_path']):
        return JSONResponse({'error': '원본 파일이 없습니다.'}, status_code=404)
    try:
        rows = _parse_sales_sheet(item['file_path'])
    except Exception as ex:
        return JSONResponse({'error': f'파싱 오류: {ex}'}, status_code=500)
    try:
        edits = json.loads(item.get('edits_json') or '{}')
    except Exception:
        edits = {}
    ccc = get_ccc_codes_for_dropdown(item['vehicle_code'])
    return JSONResponse({'rows': rows, 'edits': edits, 'ccc_codes': ccc,
                         'meta': {'vehicle': item['vehicle_code'], 'revision': item['revision'],
                                  'title': item['title'], 'filename': item['filename']}})


@app.post('/sales/price/files/save-edits/{item_id}')
async def sales_files_save_edits(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    item = get_sales_file(item_id)
    if not item:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and item['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 수정할 수 있습니다.'}, status_code=403)
    body = await request.json()
    edits = body.get('edits', {})
    update_sales_file_edits(item_id, json.dumps(edits, ensure_ascii=False))
    return JSONResponse({'ok': True, 'count': len(edits)})


@app.get('/sales/price/files/download/{item_id}')
async def sales_files_download(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return redir
    item = get_sales_file(item_id)
    if not item or not item.get('file_path') or not os.path.exists(item['file_path']):
        return JSONResponse({'error': '원본 파일이 없습니다.'}, status_code=404)
    try:
        edits = json.loads(item.get('edits_json') or '{}')
    except Exception:
        edits = {}
    if not edits:
        # 편집 없음 → 원본 그대로
        return FileResponse(item['file_path'], filename=item['filename'])
    # 원본을 템플릿으로 열어 편집값만 덮어쓰기 (서식/양식 100% 보존)
    import openpyxl
    C = SALES_SHEET_COLS
    wb = openpyxl.load_workbook(item['file_path'])
    ws = wb.active
    for rstr, e in edits.items():
        try:
            r = int(rstr)
        except Exception:
            continue
        if e.get('color') is not None and str(e.get('color')) != '':
            ws.cell(r, C['color']).value = e['color']
        f = _sales_num(e.get('jeonga'))
        g = _sales_num(e.get('sagup'))
        if f is not None:
            ws.cell(r, C['jeonga']).value = f
        if g is not None:
            ws.cell(r, C['sagup']).value = g
            ws.cell(r, C['hap']).value = g            # 합계 = 사급
        if f is not None and g is not None:
            ws.cell(r, C['after']).value = f + g      # 변경후 = 종전가 + 사급
    out_path = os.path.join(REPORTS_DIR, f'SALES_{uuid.uuid4().hex[:10]}.xlsx')
    wb.save(out_path)
    return FileResponse(out_path, filename=item['filename'],
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.post('/sales/price/files/delete/{item_id}')
async def sales_files_delete(request: Request, item_id: int):
    redir = require_login(request)
    if redir: return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    item = get_sales_file(item_id)
    if not item:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    if me['role'] != 'admin' and item['uploaded_by'] != me['username']:
        return JSONResponse({'error': '본인 또는 관리자만 삭제할 수 있습니다.'}, status_code=403)
    info = delete_sales_file(item_id)
    if info and info.get('file_path') and os.path.exists(info['file_path']):
        try: os.unlink(info['file_path'])
        except Exception: pass
    return JSONResponse({'ok': True})


# ── 생산 대시보드 ─────────────────────────────────────────────────────────────
def _month_weeks(year: int, month: int) -> list:
    """월요일 시작 기준 «n월 n주차» 구간을 계산한다.
       1일이 포함된 첫 주는 월요일 전이라도 1주차로 센다(예: 7/1(수)~7/5(일)=1주차,
       7/6(월)~7/12(일)=2주차...) — 사용자 확정 규칙."""
    import calendar
    from datetime import date
    last_day = calendar.monthrange(year, month)[1]
    first_weekday = date(year, month, 1).weekday()  # 월=0..일=6
    first_week_end = min(1 + (6 - first_weekday), last_day)
    weeks = [{'week_no': 1, 'start': 1, 'end': first_week_end}]
    d = first_week_end + 1
    wn = 2
    while d <= last_day:
        end = min(d + 6, last_day)
        weeks.append({'week_no': wn, 'start': d, 'end': end})
        d = end + 1
        wn += 1
    return weeks


@app.get('/production-dashboard', response_class=HTMLResponse)
async def production_dashboard_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    from datetime import datetime
    today = datetime.now()
    return templates.TemplateResponse(request=request, name='production_dashboard.html',
                                      context={'me': me, 'vcodes': get_all_vehicle_codes(),
                                               'cur_year': today.year, 'cur_month': today.month})


@app.get('/production-dashboard/api/weeks')
async def production_dashboard_weeks(request: Request, year: int, month: int):
    if require_login(request):
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    if not (1 <= month <= 12):
        return JSONResponse({'error': '월이 올바르지 않습니다.'}, status_code=400)
    return JSONResponse({'weeks': _month_weeks(year, month)})


@app.get('/production-dashboard/api/summary')
async def production_dashboard_summary(request: Request, year: int, month: int):
    if require_login(request):
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    vmap = {v['code']: v['name'] for v in get_all_vehicle_codes()}
    items = []
    for s in get_production_summary(year, month):
        items.append({'code': s['vehicle_code'], 'name': vmap.get(s['vehicle_code'], s['vehicle_code']),
                     'plan': s['plan_sum'] or 0, 'actual': s['actual_sum'] or 0,
                     'revenue': s['revenue_sum'] or 0, 'profit': s['profit_sum'] or 0})
    items.sort(key=lambda x: -(x['plan'] + x['actual']))
    return JSONResponse({'items': items})


@app.get('/production-dashboard/api/rows')
async def production_dashboard_rows(request: Request, year: int = None, month: int = None):
    if require_login(request):
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    rows = get_production_qty_rows(year, month)
    vmap = {v['code']: v['name'] for v in get_all_vehicle_codes()}
    for r in rows:
        r['vehicle_name'] = vmap.get(r['vehicle_code'], r['vehicle_code'])
    return JSONResponse({'rows': rows})


@app.post('/production-dashboard/row')
async def production_dashboard_save(request: Request):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    me = current_user(request)
    try:
        item = await request.json()
    except Exception:
        return JSONResponse({'error': '잘못된 요청'}, status_code=400)
    vehicle_code = str(item.get('vehicle_code', '')).strip().upper()
    try:
        year = int(item.get('year'))
        month = int(item.get('month'))
        week_no = int(item.get('week_no'))
        plan_qty = int(item.get('plan_qty') or 0)
        actual_qty = int(item.get('actual_qty') or 0)
        revenue = int(float(item.get('revenue') or 0))
        profit = int(float(item.get('profit') or 0))
    except (TypeError, ValueError):
        return JSONResponse({'error': '연도/월/주차/수량·금액은 숫자여야 합니다.'}, status_code=400)
    if not vehicle_code or not (1 <= month <= 12) or not (1 <= week_no <= 6):
        return JSONResponse({'error': '입력값을 확인하세요.'}, status_code=400)
    r = upsert_production_qty(vehicle_code, year, month, week_no, plan_qty, actual_qty,
                              me['username'], revenue=revenue, profit=profit)
    if not r.get('ok'):
        return JSONResponse({'error': r.get('msg', '저장 실패')}, status_code=400)
    return JSONResponse({'ok': True})


@app.post('/production-dashboard/row/{row_id:int}/delete')
async def production_dashboard_delete(request: Request, row_id: int):
    redir = require_login(request)
    if redir:
        return JSONResponse({'error': '로그인 필요'}, status_code=401)
    delete_production_qty(row_id)
    return JSONResponse({'ok': True})


# ── 전체 프로세스 개요 (사이드바 최상단) ───────────────────────────────────────
PROCESS_DIAGRAM_EXTS = ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.pptx')
PROCESS_DIAGRAM_IMG_MIME = {'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                            'gif': 'image/gif', 'webp': 'image/webp'}


FLOWCHART_IMG_EXTS = ('.png', '.jpg', '.jpeg', '.gif', '.webp')


@app.get('/process-overview', response_class=HTMLResponse)
async def process_overview_page(request: Request):
    redir = require_login(request)
    if redir: return redir
    me = current_user(request)
    diagrams = get_all_process_diagrams()
    flowchart_override = get_flowchart_override()
    return templates.TemplateResponse(request=request, name='process_overview.html',
                                      context={'me': me, 'diagrams': diagrams,
                                               'img_exts': set(PROCESS_DIAGRAM_IMG_MIME),
                                               'flowchart_override': flowchart_override})


@app.post('/process-overview/flowchart/upload')
async def flowchart_upload(request: Request, file: UploadFile = File(...)):
    """관리자가 사진을 올리면 코드로 그린 SVG 흐름도 자리를 그 사진이 대체한다."""
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    me = current_user(request)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in FLOWCHART_IMG_EXTS:
        return JSONResponse({'error': 'png, jpg, jpeg, gif, webp 이미지 파일만 지원합니다.'}, status_code=400)
    save_path = os.path.join(PROCESS_DIAGRAM_DIR, f'flowchart_{uuid.uuid4().hex[:12]}{ext}')
    with open(save_path, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    old = get_flowchart_override()
    set_flowchart_override(filename=file.filename, file_path=save_path,
                           file_ext=ext.lstrip('.'), uploaded_by=me['username'])
    if old and old.get('file_path') and os.path.exists(old['file_path']):
        try: os.unlink(old['file_path'])
        except Exception: pass
    return JSONResponse({'ok': True})


@app.get('/process-overview/flowchart/view')
async def flowchart_view(request: Request):
    redir = require_login(request)
    if redir: return redir
    info = get_flowchart_override()
    if not info or not os.path.exists(info['file_path']):
        return JSONResponse({'error': '등록된 사진이 없습니다.'}, status_code=404)
    ext = info.get('file_ext', '').lower()
    if ext not in PROCESS_DIAGRAM_IMG_MIME:
        return JSONResponse({'error': '이미지 파일이 아닙니다.'}, status_code=400)
    return FileResponse(info['file_path'], media_type=PROCESS_DIAGRAM_IMG_MIME[ext])


@app.post('/process-overview/flowchart/reset')
async def flowchart_reset(request: Request):
    """사진을 지우고 코드로 그린 SVG 흐름도로 복원."""
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    info = clear_flowchart_override()
    if info and info.get('file_path') and os.path.exists(info['file_path']):
        try: os.unlink(info['file_path'])
        except Exception: pass
    return JSONResponse({'ok': True})


@app.post('/process-overview/upload')
async def process_overview_upload(request: Request, title: str = Form(...),
                                  description: str = Form(''), file: UploadFile = File(...)):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    me = current_user(request)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in PROCESS_DIAGRAM_EXTS:
        return JSONResponse({'error': 'png, jpg, jpeg, gif, webp, pptx 파일만 지원합니다.'}, status_code=400)
    if not title.strip():
        return JSONResponse({'error': '제목을 입력하세요.'}, status_code=400)
    save_path = os.path.join(PROCESS_DIAGRAM_DIR, f'{uuid.uuid4().hex[:12]}{ext}')
    with open(save_path, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    result = add_process_diagram(title=title, description=description, filename=file.filename,
                                 file_path=save_path, file_ext=ext.lstrip('.'),
                                 uploaded_by=me['username'])
    return JSONResponse({'ok': True, 'id': result['id']})


@app.post('/process-overview/{diagram_id:int}/replace')
async def process_overview_replace(request: Request, diagram_id: int, file: UploadFile = File(...)):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    me = current_user(request)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in PROCESS_DIAGRAM_EXTS:
        return JSONResponse({'error': 'png, jpg, jpeg, gif, webp, pptx 파일만 지원합니다.'}, status_code=400)
    save_path = os.path.join(PROCESS_DIAGRAM_DIR, f'{uuid.uuid4().hex[:12]}{ext}')
    with open(save_path, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    old = replace_process_diagram_file(diagram_id, filename=file.filename, file_path=save_path,
                                       file_ext=ext.lstrip('.'), uploaded_by=me['username'])
    if not old:
        try: os.unlink(save_path)
        except Exception: pass
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    old_path = old.get('file_path')
    if old_path and os.path.exists(old_path):
        try: os.unlink(old_path)
        except Exception: pass
    return JSONResponse({'ok': True})


@app.get('/process-overview/{diagram_id:int}/view')
async def process_overview_view(request: Request, diagram_id: int):
    redir = require_login(request)
    if redir: return redir
    info = get_process_diagram(diagram_id)
    if not info or not os.path.exists(info['file_path']):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    ext = info.get('file_ext', '').lower()
    if ext not in PROCESS_DIAGRAM_IMG_MIME:
        return JSONResponse({'error': '이미지 파일이 아닙니다.'}, status_code=400)
    return FileResponse(info['file_path'], media_type=PROCESS_DIAGRAM_IMG_MIME[ext])


@app.get('/process-overview/{diagram_id:int}/download')
async def process_overview_download(request: Request, diagram_id: int):
    redir = require_login(request)
    if redir: return redir
    info = get_process_diagram(diagram_id)
    if not info or not os.path.exists(info['file_path']):
        return JSONResponse({'error': '파일을 찾을 수 없습니다.'}, status_code=404)
    return FileResponse(info['file_path'], filename=info['filename'])


@app.post('/process-overview/{diagram_id:int}/delete')
async def process_overview_delete(request: Request, diagram_id: int):
    if require_admin(request):
        return JSONResponse({'error': '관리자 권한이 필요합니다.'}, status_code=403)
    info = delete_process_diagram(diagram_id)
    if not info:
        return JSONResponse({'error': '찾을 수 없습니다.'}, status_code=404)
    fp = info.get('file_path')
    if fp and os.path.exists(fp):
        try: os.unlink(fp)
        except Exception: pass
    return JSONResponse({'ok': True})
