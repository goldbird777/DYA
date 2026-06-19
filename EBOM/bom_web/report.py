"""
Excel 검증 리포트 생성
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from validators import SPEC_KEYWORDS, MEANINGS, SKIP_SPEC_KEYS


def _side():
    return Side(style='thin', color='CCCCCC')


def _brd():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left',   vertical='center', wrap_text=True)


def _hdr(cell, bg: str, fg: str = 'FFFFFF', sz: int = 10):
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.font      = Font(name='맑은 고딕', color=fg, bold=True, size=sz)
    cell.alignment = CTR
    cell.border    = _brd()


def _dat(cell, bg: str, bold: bool = False, color: str = '111111', sz: int = 10):
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.font      = Font(name='맑은 고딕', color=color, bold=bold, size=sz)
    cell.alignment = LFT
    cell.border    = _brd()


# 시트별 색상
COLORS = dict(
    navy='002C5F', blue='1565C0', dark='37474F',
    red_hd='B71C1C',   red_bg='FFCDD2',
    warn_hd='E65100',  warn_bg='FFF3E0',
    qty_hd='F57F17',   qty_bg='FFFDE7',
    dup_hd='1565C0',   dup_bg='E3F2FD',
    pno_hd='4527A0',   pno_bg='EDE7F6',
    struck_hd='546E7A', struck_bg='ECEFF1',
    hl_hd='00695C',    hl_bg='E0F2F1',
    green_bg='E8F5E9',
)

TYPE_SHEET = {
    '🔴 오사양 누락': ('🔴 오사양 누락',  COLORS['red_hd'],  COLORS['red_bg']),
    '⚠️ 사양 불일치': ('⚠️ 사양 불일치', COLORS['warn_hd'], COLORS['warn_bg']),
    '🟡 수량 오기입': ('🟡 수량 오기입',  COLORS['qty_hd'],  COLORS['qty_bg']),
    '❌ 레벨 중복':   ('❌ 레벨 중복',    COLORS['red_hd'],  COLORS['red_bg']),
    '🔵 P/NO 중복':   ('🔵 PNO 오류',    COLORS['pno_hd'],  COLORS['pno_bg']),
    '🟠 P/NO 누락':   ('🔵 PNO 오류',    COLORS['pno_hd'],  COLORS['pno_bg']),
}


def _title_row(ws, text: str, cols: int, bg: str, sz: int = 12):
    end = chr(ord('A') + cols - 1)
    ws.merge_cells(f'A1:{end}1')
    ws['A1'] = '  ' + text
    ws['A1'].fill      = PatternFill('solid', start_color=bg)
    ws['A1'].font      = Font(name='맑은 고딕', color='FFFFFF', bold=True, size=sz)
    ws['A1'].alignment = LFT
    ws.row_dimensions[1].height = 30


def _error_sheet(wb: Workbook, title: str, errors: list, hd_bg: str, data_bg: str):
    ws = wb.create_sheet(title)
    cols = ['Variant', 'Lv', 'P/NO', 'P/NAME', '부품 DESCRIPTION',
            '1레벨 DESCRIPTION', '상세 내용', 'Excel 행']
    widths = [12, 6, 32, 32, 40, 55, 50, 14]

    for col, w in zip('ABCDEFGH', widths):
        ws.column_dimensions[col].width = w

    _title_row(ws, title, len(cols), hd_bg)

    for ci, h in enumerate(cols, 1):
        _hdr(ws.cell(2, ci, h), COLORS['navy'])
    ws.row_dimensions[2].height = 20

    if errors:
        for ri, e in enumerate(errors, 3):
            row_data = [
                e['variant'], e['level'], e['pno'], e['pname'],
                e['desc'], e['lv1_desc'], e.get('detail', e['message']),
                str(e['excel_rows']),
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(ri, ci, val)
                _dat(cell, data_bg, bold=(ci == 1),
                     color='C62828' if ci == 1 and 'ERROR' in str(e.get('severity')) else '222222')
            ws.row_dimensions[ri].height = 22
    else:
        ws.merge_cells('A3:H3')
        ws['A3'] = '  ✅ 해당 오류 없음'
        ws['A3'].fill      = PatternFill('solid', start_color=COLORS['green_bg'])
        ws['A3'].font      = Font(name='맑은 고딕', color='1B5E20', bold=True, size=11)
        ws['A3'].alignment = LFT

    return ws


def make_report(filename: str, errors: list, lv1_by_vc: dict,
                variant_cols: dict, struck_parts: list,
                highlighted_parts: list, out_path: str) -> str:

    wb = Workbook()

    # ── Sheet 1: 검증 요약 ────────────────────────────────────────────────
    ws = wb.active
    ws.title = '📋 검증 요약'
    for col, w in zip('ABC', [30, 14, 60]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:C1')
    ws['A1'] = '  DYA 시트 BOM 자동 검증 리포트'
    ws['A1'].fill      = PatternFill('solid', start_color=COLORS['navy'])
    ws['A1'].font      = Font(name='맑은 고딕', color='FFFFFF', bold=True, size=14)
    ws['A1'].alignment = LFT
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:C2')
    ws['A2'] = f'  파일: {os.path.basename(filename)}'
    ws['A2'].fill      = PatternFill('solid', start_color='1565C0')
    ws['A2'].font      = Font(name='맑은 고딕', color='BBDEFB', size=10)
    ws['A2'].alignment = LFT
    ws.row_dimensions[2].height = 20

    err_types = {t: sum(1 for e in errors if e['type'] == t) for t in [
        '🔴 오사양 누락', '⚠️ 사양 불일치', '🟡 수량 오기입',
        '❌ 레벨 중복', '🔵 P/NO 중복', '🟠 P/NO 누락']}

    total_errors   = sum(1 for e in errors if e['severity'] == 'ERROR')
    total_warnings = sum(1 for e in errors if e['severity'] == 'WARNING')
    ok_count       = len(lv1_by_vc) - len({e['variant'] for e in errors if e['severity'] == 'ERROR'})

    stats = [
        ('총 Variant 수',        len(lv1_by_vc),          None),
        ('❌ ERROR 합계',         total_errors,             'ERROR'),
        ('⚠️  WARNING 합계',      total_warnings,           'WARN'),
        ('✅ 오류 없는 Variant',   max(ok_count, 0),         'OK'),
        ('🔕 취소선 폐지 부품',    len(struck_parts),        'STRUCK'),
        ('🌟 음영 변경 부품',      len(highlighted_parts),   'HL'),
        ('─── 유형별 ───',        '',                       None),
        ('🔴 오사양 누락',         err_types['🔴 오사양 누락'], 'ERROR'),
        ('⚠️  사양 불일치',         err_types['⚠️ 사양 불일치'], 'WARN'),
        ('🟡 수량 오기입',         err_types['🟡 수량 오기입'], 'WARN'),
        ('❌ 레벨 중복',           err_types['❌ 레벨 중복'],  'ERROR'),
        ('🔵/🟠 P/NO 오류',        err_types['🔵 P/NO 중복'] + err_types['🟠 P/NO 누락'], 'WARN'),
    ]

    sev_colors = {'ERROR': 'C62828', 'WARN': 'E65100',
                  'OK': '2E7D32', 'STRUCK': '546E7A', 'HL': '00695C', None: '37474F'}

    for i, (lbl, val, sev) in enumerate(stats):
        ri = i + 4
        ws.row_dimensions[ri].height = 24
        bg = 'F5F5F5' if i % 2 else 'FFFFFF'
        if lbl.startswith('─'):
            bg = 'E3F2FD'
        ws[f'A{ri}'] = lbl
        ws[f'B{ri}'] = val
        for col in 'ABC':
            _dat(ws[f'{col}{ri}'], bg)
        ws[f'A{ri}'].font = Font(name='맑은 고딕', bold=True, size=10, color='37474F')
        ws[f'B{ri}'].font = Font(name='맑은 고딕', bold=True, size=13,
                                  color=sev_colors.get(sev, '37474F'))
        ws[f'B{ri}'].alignment = CTR

    # Variant 목록
    r0 = len(stats) + 5
    ws.row_dimensions[r0].height = 10
    ws.merge_cells(f'A{r0+1}:C{r0+1}')
    ws[f'A{r0+1}'] = '  1레벨 Variant 목록'
    ws[f'A{r0+1}'].fill      = PatternFill('solid', start_color=COLORS['blue'])
    ws[f'A{r0+1}'].font      = Font(name='맑은 고딕', color='FFFFFF', bold=True, size=11)
    ws[f'A{r0+1}'].alignment = LFT
    ws.row_dimensions[r0+1].height = 24

    for ci, h in enumerate(['Variant', 'P/NO', 'DESCRIPTION (사양)'], 1):
        _hdr(ws.cell(r0+2, ci, h), COLORS['navy'])
    ws.row_dimensions[r0+2].height = 20

    for i, (vc_code, lv1) in enumerate(sorted(lv1_by_vc.items()), r0+3):
        has_err = any(e['variant'] == vc_code and e['severity'] == 'ERROR'  for e in errors)
        has_wrn = any(e['variant'] == vc_code and e['severity'] == 'WARNING' for e in errors)
        bg  = COLORS['red_bg']  if has_err else (COLORS['warn_bg'] if has_wrn else COLORS['green_bg'])
        clr = 'C62828'          if has_err else ('E65100'           if has_wrn else '1B5E20')
        for ci, val in enumerate([vc_code, lv1['pno'], lv1['desc']], 1):
            _dat(ws.cell(i, ci, val), bg, bold=(ci == 1), color=clr if ci == 1 else '222222')
        ws.row_dimensions[i].height = 18

    # ── Sheet 2~5: 오류 유형별 ────────────────────────────────────────────
    seen_sheets: dict[str, object] = {}
    for e_type, (sheet_name, hd_bg, data_bg) in TYPE_SHEET.items():
        if sheet_name not in seen_sheets:
            e_list = [e for e in errors if e['type'] in
                      [k for k, v in TYPE_SHEET.items() if v[0] == sheet_name]]
            ws_e = _error_sheet(wb, sheet_name, e_list, hd_bg, data_bg)
            seen_sheets[sheet_name] = ws_e

    # ── Sheet: 취소선 부품 ────────────────────────────────────────────────
    ws_s = wb.create_sheet('🔕 취소선 폐지 부품')
    for col, w in zip('ABCD', [10, 32, 32, 55]):
        ws_s.column_dimensions[col].width = w
    _title_row(ws_s, '취소선(strikethrough) 감지 부품 — 검증에서 자동 제외됨',
               4, COLORS['struck_hd'])
    for ci, h in enumerate(['Excel 행', 'P/NO', 'P/NAME', 'DESCRIPTION'], 1):
        _hdr(ws_s.cell(2, ci, h), COLORS['dark'])
    ws_s.row_dimensions[2].height = 20

    if struck_parts:
        for ri, p in enumerate(struck_parts, 3):
            for ci, val in enumerate([p['row_idx'], p['pno'], p['pname'], p['desc']], 1):
                cell = ws_s.cell(ri, ci, val)
                _dat(cell, COLORS['struck_bg'], color='757575')
                cell.font = Font(name='맑은 고딕', color='9E9E9E', size=10, strike=True)
            ws_s.row_dimensions[ri].height = 18
    else:
        ws_s.merge_cells('A3:D3')
        ws_s['A3'] = '  취소선 부품 없음'
        ws_s['A3'].fill      = PatternFill('solid', start_color=COLORS['green_bg'])
        ws_s['A3'].font      = Font(name='맑은 고딕', color='1B5E20', bold=True)
        ws_s['A3'].alignment = LFT

    # ── Sheet: 음영 변경 부품 ─────────────────────────────────────────────
    ws_h = wb.create_sheet('🌟 음영 변경 부품')
    for col, w in zip('ABCDE', [10, 32, 32, 55, 12]):
        ws_h.column_dimensions[col].width = w
    _title_row(ws_h, '음영(배경색) 감지 부품 — 변경·수정된 항목', 5, COLORS['hl_hd'])
    for ci, h in enumerate(['Excel 행', 'P/NO', 'P/NAME', 'DESCRIPTION', '색상(RGB)'], 1):
        _hdr(ws_h.cell(2, ci, h), COLORS['dark'])
    ws_h.row_dimensions[2].height = 20

    if highlighted_parts:
        for ri, p in enumerate(highlighted_parts, 3):
            for ci, val in enumerate([p['row_idx'], p['pno'], p['pname'], p['desc'], p.get('color', '')], 1):
                _dat(ws_h.cell(ri, ci, val), COLORS['hl_bg'])
            ws_h.row_dimensions[ri].height = 18
    else:
        ws_h.merge_cells('A3:E3')
        ws_h['A3'] = '  음영 부품 없음'
        ws_h['A3'].fill      = PatternFill('solid', start_color=COLORS['green_bg'])
        ws_h['A3'].font      = Font(name='맑은 고딕', color='1B5E20', bold=True)
        ws_h['A3'].alignment = LFT

    # ── Sheet: 키워드 기준표 ──────────────────────────────────────────────
    ws_k = wb.create_sheet('📖 키워드 기준표')
    for col, w in zip('ABCD', [15, 55, 30, 14]):
        ws_k.column_dimensions[col].width = w
    _title_row(ws_k, '사양 키워드 매핑 기준표', 4, COLORS['navy'])
    for ci, h in enumerate(['키워드', '인식 패턴', '의미', '검증 제외'], 1):
        _hdr(ws_k.cell(2, ci, h), COLORS['blue'])
    ws_k.row_dimensions[2].height = 20

    for ri, (key, kws) in enumerate(SPEC_KEYWORDS.items(), 3):
        bg   = 'F5F5F5' if ri % 2 else 'FFFFFF'
        skip = '제외' if key in SKIP_SPEC_KEYS else ''
        for ci, val in enumerate([key, ', '.join(kws), MEANINGS.get(key, ''), skip], 1):
            cell = ws_k.cell(ri, ci, val)
            _dat(cell, bg)
            if ci == 4 and skip:
                cell.font = Font(name='맑은 고딕', color='9E9E9E', size=10, italic=True)
        ws_k.row_dimensions[ri].height = 18

    wb.save(out_path)
    return out_path
