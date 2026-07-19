# -*- coding: utf-8 -*-
"""DYA ALC-2 변환 엔진 (DYA_ALC2_LOCAL/app.py 웹 포팅).

핵심: HKMC Q파트 종합 + ALC 원본 5종(FRT LH/RH, RR BACK LH/CUSH/RH) + 통합 ALC2 마스터
 → 각 생산행(KEY02~06 = KMC20)을 판정:
    · 원본누락 : KEY가 해당 시트 ALC 코드집에 없음
    · 기존매칭 : KMC20이 통합 마스터에 있음 → 기존 DYA ALC-2 재사용
    · 신규승인필요 : 마스터에 없음 → 임시코드(AAxxJ) 부여, 수동승인 대상

app.py는 config.json의 열 위치(A,B,C)를 썼지만, 여기서는 **헤더 이름 자동인식**으로
열을 찾으므로 HKMC 양식의 열이 바뀌어도 동작한다.
"""
import os
import re
import openpyxl

# ALC 원본 슬롯 순서 = KEY02~KEY06 순서와 1:1 대응
ALC_SLOTS = ['FRT LH', 'FRT RH', 'RR BACK LH', 'RR CUSH', 'RR BACK RH']


def _cell(ws, r, c):
    if c is None:
        return ''
    v = ws.cell(r, c).value
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _find_header(ws, keys, maxr=20):
    """지정 키워드가 모두 들어있는 헤더 행 번호를 찾는다(대소문자 무시, 부분일치)."""
    keys = [k.upper() for k in keys]
    for r in range(1, min(maxr, ws.max_row) + 1):
        rowvals = [_cell(ws, r, c).upper() for c in range(1, ws.max_column + 1)]
        if all(any(k in v for v in rowvals) for k in keys):
            return r
    return None


def read_qpart(path):
    """Q파트 종합 → [{vehicle, line, key01(국가), keys[가변], kmc20(가변길이), excel_row}].
       KEY02부터 값 있는 마지막 KEY열까지 조합. X 표기 좌석은 '****'로 치환(행 유지).
       예) [507S,607S,604S,X,704S] → '507S607S604S****704S' (20). NQ5 7그룹이면 28자리."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    hr = _find_header(ws, ['차종', 'KEY01', 'KEY02'])
    if hr is None:
        raise ValueError('Q파트에서 헤더(차종/KEY01~)를 찾지 못했습니다.')
    hdr = {_cell(ws, hr, c).upper(): c for c in range(1, ws.max_column + 1)}
    cvehicle, cline, ckey01 = hdr.get('차종'), hdr.get('라인'), hdr.get('KEY01')
    key_cols = [hdr.get('KEY%02d' % i) for i in range(2, 10)]
    key_cols = [c for c in key_cols if c]
    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        vehicle = _cell(ws, r, cvehicle)
        if not vehicle:
            continue
        raw = [_cell(ws, r, c) for c in key_cols]
        while raw and raw[-1] == '':          # 후미 빈칸 제거 → 차종별 그룹 수
            raw.pop()
        if len(raw) < 2:
            continue
        groups, ok = [], True
        for v in raw:
            vu = v.strip().upper()
            if vu == 'X':
                groups.append('****')          # X = 해당 좌석 미해당 → ****
            elif re.fullmatch(r'[A-Za-z0-9]{4}', vu):
                groups.append(vu)
            else:
                ok = False; break              # 예상외 값(중간 빈칸 등) → 제외
        if not ok:
            continue
        rows.append({'vehicle': vehicle, 'line': _cell(ws, r, cline),
                     'key01': _cell(ws, r, ckey01), 'keys': groups,
                     'kmc20': ''.join(groups), 'excel_row': r})
    return rows


def read_alc(path):
    """ALC 코드집 → 4자리 CODE 집합."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    hr = _find_header(ws, ['CODE', 'PART NO'])
    if hr is None:
        raise ValueError('ALC 코드집에서 헤더(CODE/PART NO)를 찾지 못했습니다.')
    hd = {_cell(ws, hr, c).upper(): c for c in range(1, ws.max_column + 1)}
    cc = hd.get('CODE')
    codes = set()
    for r in range(hr + 1, ws.max_row + 1):
        cd = _cell(ws, r, cc).upper()
        if re.fullmatch(r'[A-Z0-9]{4}', cd):
            codes.add(cd)
    return codes


_MASTER_CACHE = {}   # path -> (mtime, size, master)


def read_master(path):
    """통합 ALC2 대장 → {KMC코드: {alc2(5자리)}}. 헤더로 DYA/KMC 열 자동탐지.
       대장(3천행×283열)은 매 실행마다 바뀌지 않으므로 mtime 기준으로 캐시한다."""
    try:
        st = os.stat(path)
        ck = (st.st_mtime, st.st_size)
        hit = _MASTER_CACHE.get(path)
        if hit and hit[0] == ck:
            return hit[1]
    except OSError:
        ck = None
    rows = _load_rows(path)   # read_only 스트리밍
    hi = _find_hdr_idx(rows, ['ALC-2 CODE', 'KMC ALC-2'])
    master = {}
    if hi is None:
        return master
    hd = {_t(v).upper(): i for i, v in enumerate(rows[hi])}
    dcol = next((i for k, i in hd.items() if 'ALC-2 CODE' in k and 'KMC' not in k), None)
    kcol = next((i for k, i in hd.items() if 'KMC ALC-2' in k), None)
    if dcol is None or kcol is None:
        return master
    for r in range(hi + 1, len(rows)):
        row = rows[r]
        if kcol >= len(row) or dcol >= len(row):
            continue
        a = _t(row[dcol]).upper()
        k = re.sub(r'\s', '', _t(row[kcol])).upper()
        # KMC 코드는 차종별 가변길이(20·28자)이고 미적용 좌석은 '*'로 표기된다
        if re.fullmatch(r'[A-Z0-9]{5}', a) and re.fullmatch(r'[A-Z0-9*]{12,40}', k):
            master[k] = {'alc2': a, 'row': r + 1}
    if ck:
        _MASTER_CACHE[path] = (ck, master)
    return master


def _load_rows(path):
    """read_only + values_only 로 전 행을 리스트로 (대용량 ALC 가속)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _t(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _find_hdr_idx(rows, keys, maxr=20):
    keys = [k.upper() for k in keys]
    for i in range(min(maxr, len(rows))):
        vals = [_t(x).upper() for x in rows[i]]
        if all(any(k in v for v in vals) for k in keys):
            return i
    return None


def _combi_col_range(rows, hi):
    """CODE/PART NO 헤더(0-based hi) 아래 속성 그룹 행에서 «COMBI» 열 구간만 찾는다.
       그 오른쪽의 EXCLU/EO-IN/EO-OUT은 실제 적용 옵션이 아니라 제외/예외 정보라서
       PEL 코드로 읽으면 안 된다 — COMBI 밖의 6자리 코드를 옵션으로 오인하면 안 됨.
       반환: (시작열, 끝열+1) 0-based 반열림 구간. 못 찾으면 None(전체 스캔으로 폴백)."""
    for i in range(hi, min(hi + 4, len(rows))):
        row = rows[i] or []
        start = end = None
        for j, v in enumerate(row):
            if _t(v).upper() == 'COMBI':
                if start is None:
                    start = j
                end = j
        if start is not None:
            return start, end + 1
    return None


def _pel_codes_in_row(row, combi_range):
    """행에서 PEL 코드(6자리)를 뽑는다 — COMBI 구간을 찾았으면 그 안에서만,
       못 찾았으면(구조가 다른 파일 대비) 전체 행에서 폴백 스캔."""
    cells = row[combi_range[0]:combi_range[1]] if combi_range else row
    return [_t(x).upper() for x in cells if re.fullmatch(r'[0-9A-Z]{6}', _t(x).upper())]


def read_alc_full(path):
    """ALC 코드집 1회 로드 → {CODE(4자리): [PEL코드(6자리)...]}. (codes = result.keys())
       PEL 코드는 COMBI 열 구간에서만 읽는다(EXCLU/EO-IN/EO-OUT 제외)."""
    rows = _load_rows(path)
    hi = _find_hdr_idx(rows, ['CODE', 'PART NO'])
    if hi is None:
        raise ValueError('ALC 코드집 헤더(CODE/PART NO)를 찾지 못했습니다.')
    hdr = {_t(x).upper(): j for j, x in enumerate(rows[hi])}
    cc = hdr.get('CODE')
    combi = _combi_col_range(rows, hi)
    result = {}
    for i in range(hi + 1, len(rows)):
        r = rows[i]
        code = _t(r[cc]).upper() if (cc is not None and cc < len(r)) else ''
        if not re.fullmatch(r'[A-Z0-9]{4}', code):
            continue
        result[code] = _pel_codes_in_row(r, combi)
    return result


def _find_dt_cols(rows, hi):
    """CODE/PART NO 헤더 행(0-based hi) 부근에서 «DT»(Driver Type) 단어를 찾아 그 아래
       1~2행의 LHD/RHD 열 인덱스를 반환한다. 실물 파일에서 DT는 보통 CODE와 같은 행(hi)에
       있고 LHD/RHD는 hi+1행, L/R은 hi+2행에 있다 — T/U열·10행처럼 위치를 고정하지 않고
       매번 탐색한다. 없으면 (None, None)."""
    for i in range(max(0, hi - 2), hi + 1):
        row = rows[i] if i < len(rows) else []
        for j, v in enumerate(row or []):
            if _t(v).upper() != 'DT':
                continue
            for k in range(i + 1, min(i + 4, len(rows))):
                sub = rows[k] or []
                lhd = rhd = None
                for jj in range(max(0, j - 2), min(len(sub), j + 10)):
                    t = _t(sub[jj]).upper()
                    if t == 'LHD':
                        lhd = jj
                    elif t == 'RHD':
                        rhd = jj
                if lhd is not None or rhd is not None:
                    return lhd, rhd
    return None, None


def read_alc_full_ex(path):
    """전석(FRT LH/RH) 검증용 확장 리더 — 1회 로드로 pel_map + CODE별 DT 방향 +
       메타정보('컬럼명' 행, 예: '1열시트-DRV')를 함께 얻는다.
       DT 표식은 실제로는 '*'이지만 특정 문자에 의존하지 않고 '비어있지 않음'으로 판정한다.
       반환: {'pel': {CODE:[pel...]}, 'dt': {CODE:'LHD'|'RHD'|'BOTH'|'NONE'},
              'dt_found': bool, 'meta': str}."""
    rows = _load_rows(path)
    hi = _find_hdr_idx(rows, ['CODE', 'PART NO'])
    if hi is None:
        raise ValueError('ALC 코드집 헤더(CODE/PART NO)를 찾지 못했습니다.')
    hdr = {_t(x).upper(): j for j, x in enumerate(rows[hi])}
    cc = hdr.get('CODE')
    lhd_c, rhd_c = _find_dt_cols(rows, hi)
    combi = _combi_col_range(rows, hi)
    meta = ''
    for i in range(min(15, len(rows))):
        row = rows[i] or []
        if row and _t(row[0]) == '컬럼명':
            for v in row[1:]:
                t = _t(v)
                if t:
                    meta = t
                    break
            break
    pel, dt = {}, {}
    for i in range(hi + 1, len(rows)):
        r = rows[i]
        code = _t(r[cc]).upper() if (cc is not None and cc < len(r)) else ''
        if not re.fullmatch(r'[A-Z0-9]{4}', code):
            continue
        pel[code] = _pel_codes_in_row(r, combi)
        if lhd_c is not None or rhd_c is not None:
            has_l = lhd_c is not None and lhd_c < len(r) and _t(r[lhd_c]) != ''
            has_r = rhd_c is not None and rhd_c < len(r) and _t(r[rhd_c]) != ''
            dt[code] = 'BOTH' if (has_l and has_r) else ('LHD' if has_l else ('RHD' if has_r else 'NONE'))
    return {'pel': pel, 'dt': dt, 'dt_found': lhd_c is not None or rhd_c is not None, 'meta': meta}


def check_frt_dt(qpart_path, alc_paths, hkmc_country_map=None):
    """전석(FRT LH/RH) DT 표식·메타정보·KEY01 국가코드를 검증해 경고 목록을 만든다.
       역할 배정은 항상 고정(FRT LH→DRIVER, FRT RH→PASSENGER)이며 이 함수는 그 배정을
       절대 바꾸지 않는다 — DT는 방향 속성 및 교차검증 전용이지 역할 교환용이 아니다.
       hkmc_country_map: {KEY01값: country_codes 행} — 없으면(None) 국가코드 검증을 건너뛴다.
       반환: {'warnings': [str,...], 'dt_found': {slot: bool}}."""
    exts, warnings, dt_found = {}, [], {}
    role_word = {'FRT LH': 'DRV', 'FRT RH': 'PASS'}
    for slot in ('FRT LH', 'FRT RH'):
        p = alc_paths.get(slot)
        if not p:
            warnings.append(f'{slot} 파일이 없어 DT 검증을 건너뜁니다.')
            continue
        ex = read_alc_full_ex(p)
        exts[slot] = ex
        dt_found[slot] = ex['dt_found']
        if not ex['dt_found']:
            warnings.append(f'{slot} 파일에서 DT(LHD/RHD) 헤더를 찾지 못했습니다 — 방향 검증을 건너뜁니다.')
        if role_word[slot] not in ex['meta'].upper():
            warnings.append(f"{slot} 파일의 메타정보('{ex['meta']}')가 예상 역할({role_word[slot]})과 일치하지 않습니다.")

    qrows = read_qpart(qpart_path)
    bad_dt = {'FRT LH': [], 'FRT RH': []}
    mismatch, unreg = [], set()
    for q in qrows:
        if hkmc_country_map is not None:
            k1 = q.get('key01', '')
            if k1 and k1 not in hkmc_country_map:
                unreg.add(k1)
        dirs = {}
        for slot, idx in (('FRT LH', 0), ('FRT RH', 1)):
            ex = exts.get(slot)
            if not ex or not ex['dt_found']:
                continue
            key = q['keys'][idx] if idx < len(q['keys']) else None
            if not key or key == '****':
                continue
            d = ex['dt'].get(key)
            if d in ('BOTH', 'NONE') and len(bad_dt[slot]) < 20:
                bad_dt[slot].append((q['kmc20'], key, d))
            elif d in ('LHD', 'RHD'):
                dirs[slot] = d
        if 'FRT LH' in dirs and 'FRT RH' in dirs and dirs['FRT LH'] != dirs['FRT RH'] and len(mismatch) < 20:
            mismatch.append((q['kmc20'], dirs['FRT LH'], dirs['FRT RH']))

    for slot, items in bad_dt.items():
        for kmc20, code, d in items[:5]:
            label = '중복(LHD+RHD 동시 표식)' if d == 'BOTH' else '누락(LHD/RHD 모두 공백)'
            warnings.append(f'{slot} {code}({kmc20}) DT 표식 {label}')
        if len(items) > 5:
            warnings.append(f'{slot} DT 누락/중복 총 {len(items)}건 (상위 5건만 표시)')
    for kmc20, lh, rh in mismatch[:5]:
        warnings.append(f'{kmc20}: 운전석 DT={lh}, 조수석 DT={rh} — 같은 조합인데 방향이 다릅니다.')
    if len(mismatch) > 5:
        warnings.append(f'전석 좌우 DT 불일치 총 {len(mismatch)}건 (상위 5건만 표시)')
    if unreg:
        warnings.append('국가코드 게시판에 등록되지 않은 KEY01 값: ' + ', '.join(sorted(unreg)))
    return {'warnings': warnings, 'dt_found': dt_found}


def analyze(qpart_path, alc_paths, master_path, master_pel):
    """6파일 1회씩만 읽어 판정 + O·X 를 함께 산출 (최적화 경로)."""
    qrows = read_qpart(qpart_path)
    alc_full = {}
    for slot in ALC_SLOTS:
        p = alc_paths.get(slot)
        alc_full[slot] = read_alc_full(p) if p else {}
    master = read_master(master_path) if master_path else {}
    used = {v['alc2'] for v in master.values()}
    seq = 1
    col_defs, jrows, oxrows = {}, [], []
    unknown_pel = {}      # PEL 코드 -> {발견된 슬롯}
    for q in qrows:
        # 판정
        missing = []
        for i in range(min(len(ALC_SLOTS), len(q['keys']))):
            k = q['keys'][i]
            if k == '****':
                continue
            if alc_full[ALC_SLOTS[i]] and k not in alc_full[ALC_SLOTS[i]]:
                missing.append(k)
        hit = master.get(q['kmc20'])
        if missing:
            status, alc2, detail = '원본누락', '', ', '.join(missing)
        elif hit:
            status, alc2, detail = '기존매칭', hit['alc2'], '정상'
        else:
            while ('AA%02dJ' % seq) in used:
                seq += 1
            alc2 = 'AA%02dJ' % seq
            used.add(alc2); seq += 1
            status, detail = '신규승인필요', '마스터 신규 조합(원단/사양 확인 후 확정)'
        jrows.append({'vehicle': q['vehicle'], 'key01': q.get('key01', ''), 'keys': q['keys'],
                      'kmc20': q['kmc20'], 'alc2': alc2, 'status': status, 'detail': detail})
        # O·X (재사용된 PEL 코드로)
        marks = set()
        for i in range(min(len(ALC_SLOTS), len(q['keys']))):
            k = q['keys'][i]
            if k == '****':
                continue
            for pc in alc_full[ALC_SLOTS[i]].get(k, []):
                m = master_pel.get(pc)
                if not m:
                    # PEL CODE 마스터에 없는 코드 → O 표기가 누락되므로 경고 대상
                    unknown_pel.setdefault(pc, set()).add(ALC_SLOTS[i])
                    continue
                sp = _spec_label(m)
                if not sp:
                    continue
                grp = str(m.get('옵션그룹', '')).strip() or 'OPTION'
                try:
                    order = float(m.get('표시순서') or 9999)
                except Exception:
                    order = 9999.0
                if sp not in col_defs or order < col_defs[sp]['order']:
                    col_defs[sp] = {'group': grp, 'order': order}
                marks.add(sp)
        oxrows.append({'vehicle': q['vehicle'], 'key01': q.get('key01', ''),
                       'kmc20': q['kmc20'], 'marks': sorted(marks)})
    columns = [{'spec': sp, 'group': d['group'], 'order': d['order']}
               for sp, d in sorted(col_defs.items(), key=lambda kv: (kv[1]['order'], kv[0]))]
    groups = []
    for c in columns:
        if groups and groups[-1]['group'] == c['group']:
            groups[-1]['span'] += 1
        else:
            groups.append({'group': c['group'], 'span': 1})
    stats = {'total': len(jrows),
             'matched': sum(r['status'] == '기존매칭' for r in jrows),
             'new': sum(r['status'] == '신규승인필요' for r in jrows),
             'missing': sum(r['status'] == '원본누락' for r in jrows)}
    return {'rows': jrows, 'stats': stats,
            'unknown_pel': [{'code': pc, 'slots': sorted(sl)}
                            for pc, sl in sorted(unknown_pel.items())],
            'ox': {'columns': columns, 'groups': groups, 'rows': oxrows}}


def read_alc_pel(path):
    """ALC 코드집 → {CODE(4자리): [PEL코드(6자리) list]}.
       라벨 없는 확산 열에 흩어진 6자리 PEL 코드를 각 CODE 행에서 수집한다.
       COMBI 열 구간에서만 읽는다(EXCLU/EO-IN/EO-OUT은 제외/예외 정보라 제외)."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    hr = _find_header(ws, ['CODE', 'PART NO'])
    if hr is None:
        raise ValueError('ALC 코드집에서 헤더(CODE/PART NO)를 찾지 못했습니다.')
    hd = {_cell(ws, hr, c).upper(): c for c in range(1, ws.max_column + 1)}
    ccode = hd.get('CODE')
    combi_start = combi_end = None
    for r in range(hr, min(hr + 4, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            if _cell(ws, r, c).upper() == 'COMBI':
                if combi_start is None:
                    combi_start = c
                combi_end = c
        if combi_start is not None:
            break
    col_range = range(combi_start, combi_end + 1) if combi_start is not None else range(1, ws.max_column + 1)
    result = {}
    for r in range(hr + 1, ws.max_row + 1):
        code = _cell(ws, r, ccode).upper()
        if not re.fullmatch(r'[A-Z0-9]{4}', code):
            continue
        pels = []
        for c in col_range:
            v = _cell(ws, r, c).upper()
            if re.fullmatch(r'[0-9A-Z]{6}', v):
                pels.append(v)
        result[code] = pels
    return result


def _norm(s):
    """비교용 정규화 — 대문자화 + 영숫자만 남김(공백·기호 무시)."""
    return re.sub(r'[^A-Z0-9]', '', str(s or '').upper())


# ALC_SLOTS → «★통합 ALC2 코드» 서식의 좌석위치(top) 블록.
# 후석 3개(2열 LH/CTR·CUSH/RH)는 좌우가 물리적 위치라 핸들방향과 무관하게 고정.
# 전석은 실제 고객사 파일을 확인한 결과(ALC2_PEL_OX_REFERENCE.md 2.1절) FRT LH 파일의
# 메타정보가 항상 '1열시트-DRV', FRT RH가 항상 '1열시트-PASS'라서 LHD/RHD와 무관하게
# FRT LH→DRIVER, FRT RH→PASSENGER로 고정한다. 파일 안의 DT(LHD/RHD) 표식은 역할을
# 서로 바꾸는 데 쓰지 않고 check_frt_dt()의 방향 검증·경고에만 사용한다.
SLOT_TOP_MAP = {
    'FRT LH': 'DRIVER',
    'FRT RH': 'PASSENGER',
    'RR BACK LH': 'Rr 2ND LH',
    'RR CUSH': 'Rr 2ND CTR or CUSH',
    'RR BACK RH': 'Rr 2ND RH',
}


def match_option_columns(option_cols, m, top_filter=None):
    """PEL 마스터 항목(m: 사양/설명)이 서식의 어느 옵션 열(«★통합 ALC2 코드» 템플릿의
       ERGO/LUMBAR SUPPORT/THORAX... 같은 고정 열)에 해당하는지 col_letter 집합으로 반환.
       option_cols: alc2_ledger.find_option_columns()의 결과 {col_letter: {'top','group','label'}}.
       top_filter가 있으면 그 좌석위치(top) 열만 후보로 본다 — 같은 리프 라벨(LUMBAR SUPPORT,
       THORAX...)이 좌석위치마다 반복되므로 좌석을 모르는 채로는 매칭하면 안 된다.
       매칭은 정규화 후 완전일치만 인정한다(오탐 방지) — «사양» 또는 «설명»의 콤마 구분
       용어 중 하나가 열 리프 라벨과 정확히 같아야 한다. 같은 좌석위치 안에서 리프 라벨이
       중복되면(POWER/MANUAL 둘 다 «LUMBAR SUPPORT») 그룹명 첫 단어도 용어 중 어딘가에
       포함되어 있어야 확정한다 — 없으면 모호하므로 매칭하지 않는다."""
    phrases = [m.get('사양', '')] + str(m.get('설명', '')).split(',')
    phrase_norms = [_norm(p) for p in phrases if str(p).strip()]
    phrase_norms = [p for p in phrase_norms if p]
    if not phrase_norms:
        return set()
    cand = option_cols
    if top_filter is not None:
        tf = _norm(top_filter)
        cand = {c: info for c, info in option_cols.items() if _norm(info.get('top', '')) == tf}
    label_cols = {}
    for col, info in cand.items():
        label_cols.setdefault(_norm(info['label']), []).append(col)
    hit = set()
    for col, info in cand.items():
        ln = _norm(info['label'])
        if not ln or ln not in phrase_norms:
            continue
        if len(label_cols[ln]) > 1:
            gword = _norm(info['group'].split()[0]) if info.get('group') else ''
            if not gword or not any(gword in p for p in phrase_norms):
                continue
        hit.add(col)
    return hit


def _seat_type_columns(option_cols):
    """좌석위치별 «시트종류» POWER/MANUAL 열을 찾는다. 반환: {top: {'POWER':col,'MANUAL':col}}."""
    out = {}
    for col, info in option_cols.items():
        if _norm(info.get('group', '')) != _norm('시트종류'):
            continue
        label = _norm(info.get('label', ''))
        if label not in ('POWER', 'MANUAL'):
            continue
        out.setdefault(info.get('top', ''), {})[label] = col
    return out


def _is_power_seat_pel(m):
    """PEL 마스터 항목이 «시트종류: 파워» 전체를 뜻하는 코드인지 판정한다.
       럼버서포트 등 개별 옵션의 파워 여부가 아니라 시트 자체의 구동 방식을 뜻하는
       옵션그룹=OPTION + «사양»이 PWR 계열(PWR, PWR(IMS))인 항목만 인정한다."""
    sp = _norm(m.get('사양', ''))
    grp = str(m.get('옵션그룹', '')).strip().upper()
    return grp == 'OPTION' and sp.startswith('PWR')


def build_option_marks(qpart_path, alc_paths, master_pel, option_cols):
    """«★통합 ALC2 코드» 서식의 고정 옵션 열(ERGO/LUMBAR SUPPORT/THORAX...)에 대해
       각 생산조합(kmc20)이 어느 열에 O가 찍혀야 하는지 계산한다.
       SLOT_TOP_MAP에 없는 슬롯(현재 FRT LH/FRT RH)은 좌석위치를 단정할 수 없어 건너뛴다.
       «시트종류»(POWER/MANUAL) 열은 텍스트 매칭이 아니라, 그 좌석에 파워시트 코드
       (_is_power_seat_pel)가 있으면 POWER, 없으면 MANUAL로 정확히 하나만 표기한다 —
       옵션 열 텍스트가 항상 바뀌어도(예: 옛 'Y열' 고정 위치) PEL 코드 존재 여부로 판정.
       반환: {kmc20: set(col_letter)}."""
    if not option_cols:
        return {}
    seat_type_cols = _seat_type_columns(option_cols)
    qrows = read_qpart(qpart_path)
    alc_full = {}
    for slot in ALC_SLOTS:
        p = alc_paths.get(slot)
        alc_full[slot] = read_alc_full(p) if p else {}
    col_cache = {}
    out = {}
    for q in qrows:
        hit_cols = set()
        power_seen, checked_tops = set(), set()
        for i in range(min(len(ALC_SLOTS), len(q['keys']))):
            slot = ALC_SLOTS[i]
            top = SLOT_TOP_MAP.get(slot)
            if not top:
                continue
            k = q['keys'][i]
            if k == '****':
                continue
            checked_tops.add(top)
            for pc in alc_full[slot].get(k, []):
                m = master_pel.get(pc)
                if not m:
                    continue
                ck = (pc, top)
                if ck not in col_cache:
                    col_cache[ck] = match_option_columns(option_cols, m, top_filter=top)
                hit_cols |= col_cache[ck]
                if _is_power_seat_pel(m):
                    power_seen.add(top)
        for top in checked_tops:
            cols = seat_type_cols.get(top)
            if not cols:
                continue
            want = 'POWER' if top in power_seen else 'MANUAL'
            if want in cols:
                hit_cols.add(cols[want])
        out[q['kmc20']] = hit_cols
    return out


def _spec_label(m):
    """PEL 마스터 항목의 O/X 열 이름. «사양»이 비어 있으면 «설명»의 콤마 구분 용어 중
       첫 번째로 대체 — 설계 사양명이 없어 O 표기가 통째로 누락되던 것을 방지."""
    sp = str(m.get('사양', '')).strip()
    if sp:
        return sp
    for part in str(m.get('설명', '')).split(','):
        part = part.strip()
        if part:
            return part
    return ''


def build_ox(qpart_path, alc_paths, master_pel):
    """6파일 + PEL CODE 마스터 → O/X 통합코드집 (PEL 사양변경 패턴).
       행=생산조합(KMC ALC-2), 열=옵션(사양), 값=O(●)/X.
       master_pel: {PEL코드: {'사양','설명','옵션그룹','표시순서'}} (load_pel_master.data).
       열 이름은 _spec_label()로 결정 — «사양»이 비어 있으면 «설명»의 콤마 구분 용어로 대체."""
    qrows = read_qpart(qpart_path)
    alc_pel = {}
    for slot in ALC_SLOTS:
        p = alc_paths.get(slot)
        alc_pel[slot] = read_alc_pel(p) if p else {}
    col_defs, rows = {}, []
    for q in qrows:
        marks = set()
        for i, slot in enumerate(ALC_SLOTS):
            if i >= len(q['keys']):
                break
            key = q['keys'][i]
            if key == '****':
                continue
            for pc in alc_pel[slot].get(key, []):
                m = master_pel.get(pc)
                if not m:
                    continue
                sp = _spec_label(m)
                if not sp:
                    continue
                grp = str(m.get('옵션그룹', '')).strip() or 'OPTION'
                try:
                    order = float(m.get('표시순서') or 9999)
                except Exception:
                    order = 9999.0
                if sp not in col_defs or order < col_defs[sp]['order']:
                    col_defs[sp] = {'group': grp, 'order': order}
                marks.add(sp)
        rows.append({'vehicle': q['vehicle'], 'key01': q.get('key01', ''),
                     'kmc20': q['kmc20'], 'marks': sorted(marks)})
    columns = [{'spec': sp, 'group': d['group'], 'order': d['order']}
               for sp, d in sorted(col_defs.items(), key=lambda kv: (kv[1]['order'], kv[0]))]
    groups = []
    for c in columns:
        if groups and groups[-1]['group'] == c['group']:
            groups[-1]['span'] += 1
        else:
            groups.append({'group': c['group'], 'span': 1})
    return {'columns': columns, 'groups': groups, 'rows': rows, 'vc_count': len(rows)}


def convert(qpart_path, alc_paths, master_path):
    """alc_paths: {slot명: path} (ALC_SLOTS 5종). master_path: 통합 마스터.
       반환: {rows:[...판정...], stats:{...}}."""
    qrows = read_qpart(qpart_path)
    maps = []
    for slot in ALC_SLOTS:
        p = alc_paths.get(slot)
        maps.append(read_alc(p) if p else set())
    master = read_master(master_path) if master_path else {}
    used = {v['alc2'] for v in master.values()}
    seq = 1
    results = []
    for q in qrows:
        # 앞 5개 KEY만 5개 ALC 코드집과 대응(**** 좌석·ALC없는 뒤 KEY는 검사 제외)
        missing = []
        for i in range(min(len(ALC_SLOTS), len(q['keys']))):
            k = q['keys'][i]
            if k == '****':
                continue
            if maps[i] and k not in maps[i]:
                missing.append(k)
        hit = master.get(q['kmc20'])
        if missing:
            status, alc2, detail = '원본누락', '', ', '.join(missing)
        elif hit:
            status, alc2, detail = '기존매칭', hit['alc2'], '정상'
        else:
            while ('AA%02dJ' % seq) in used:
                seq += 1
            alc2 = 'AA%02dJ' % seq
            used.add(alc2); seq += 1
            status, detail = '신규승인필요', '마스터 신규 조합(원단/사양 확인 후 확정)'
        results.append({'vehicle': q['vehicle'], 'line': q['line'], 'key01': q.get('key01', ''),
                        'keys': q['keys'], 'kmc20': q['kmc20'], 'alc2': alc2,
                        'status': status, 'detail': detail})
    stats = {
        'total': len(results),
        'matched': sum(r['status'] == '기존매칭' for r in results),
        'new': sum(r['status'] == '신규승인필요' for r in results),
        'missing': sum(r['status'] == '원본누락' for r in results),
    }
    return {'rows': results, 'stats': stats}
