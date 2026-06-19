"""
Excel 파일을 파싱해서 웹 렌더링용 JSON 데이터로 변환
"""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def _get_rgb(color):
    try:
        if color is None:
            return None
        if color.type == 'rgb':
            val = color.rgb
        elif color.type == 'theme':
            return None
        else:
            return None
        if not val or val in ('00000000', 'FFFFFFFF', 'FF000000', 'FFFFFFFF'):
            return None
        return str(val)[-6:]
    except Exception:
        return None


def _rgb(color):
    rgb = _get_rgb(color)
    if not rgb or rgb.upper() in ('000000', 'FFFFFF'):
        return None
    return '#' + rgb


def _font_color(color):
    rgb = _get_rgb(color)
    if not rgb:
        return None
    return '#' + rgb


def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []

    for ws in wb.worksheets:
        # 병합 셀 맵 만들기
        merge_map = {}   # (row, col) -> {'rowspan':n, 'colspan':n}
        merge_skip = set()  # 병합된 슬레이브 셀 (숨길 셀)

        for merge in ws.merged_cells.ranges:
            min_row, min_col = merge.min_row, merge.min_col
            max_row, max_col = merge.max_row, merge.max_col
            rowspan = max_row - min_row + 1
            colspan = max_col - min_col + 1
            merge_map[(min_row, min_col)] = {'rowspan': rowspan, 'colspan': colspan}
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) != (min_row, min_col):
                        merge_skip.add((r, c))

        # 컬럼 너비 (픽셀 환산: 1 Excel unit ≈ 8px)
        col_widths = {}
        for col_letter, dim in ws.column_dimensions.items():
            col_idx = column_index_from_string(col_letter)
            w = dim.width or 8
            col_widths[col_idx] = max(int(w * 8), 30)

        # 행 높이 (픽셀 환산: 1pt ≈ 1.33px)
        row_heights = {}
        for row_idx, dim in ws.row_dimensions.items():
            h = dim.height or 15
            row_heights[row_idx] = max(int(h * 1.33), 18)

        # 실제 데이터 범위 (최대 500행, 50열로 제한)
        max_row = min(ws.max_row or 1, 500)
        max_col = min(ws.max_column or 1, 60)

        rows = []
        for r in range(1, max_row + 1):
            cells = []
            for c in range(1, max_col + 1):
                if (r, c) in merge_skip:
                    cells.append(None)  # 숨길 셀
                    continue

                cell = ws.cell(row=r, column=c)
                val = cell.value
                text = '' if val is None else str(val)

                style = {}

                # 폰트
                if cell.font:
                    if cell.font.bold:
                        style['fontWeight'] = 'bold'
                    if cell.font.italic:
                        style['fontStyle'] = 'italic'
                    if cell.font.strike:
                        style['textDecoration'] = 'line-through'
                    if cell.font.size:
                        style['fontSize'] = f'{cell.font.size}pt'
                    fc = _font_color(cell.font.color) if cell.font.color else None
                    if fc and fc != '#000000':
                        style['color'] = fc

                # 배경색
                if cell.fill and cell.fill.fill_type not in (None, 'none'):
                    bg = _rgb(cell.fill.fgColor) if cell.fill.fgColor else None
                    if bg:
                        style['backgroundColor'] = bg

                # 정렬
                if cell.alignment:
                    if cell.alignment.horizontal:
                        style['textAlign'] = cell.alignment.horizontal
                    if cell.alignment.vertical:
                        style['verticalAlign'] = cell.alignment.vertical
                    if cell.alignment.wrap_text:
                        style['whiteSpace'] = 'pre-wrap'

                span = merge_map.get((r, c), {})

                cells.append({
                    'text': text,
                    'style': style,
                    'rowspan': span.get('rowspan', 1),
                    'colspan': span.get('colspan', 1),
                })
            rows.append(cells)

        sheets.append({
            'name': ws.title,
            'rows': rows,
            'col_widths': col_widths,
            'row_heights': row_heights,
            'max_col': max_col,
        })

    return sheets
