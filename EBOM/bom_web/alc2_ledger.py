"""★통합 ALC2 코드(REV) 서식을 유지한 채 데이터 영역만 변환 결과로 교체한다.

헤더의 색상·병합·열 구조(282열)는 원본 그대로 두고, 첫 데이터 행(8행)부터
이번 변환 결과만 채운다. 기존 대장 데이터를 남기면 이전 파일과 비교가 안 된다.

openpyxl로 load_workbook→save 하면 3,111행×283열(88만 셀)을 전부 파싱·재직렬화하느라
15초가 걸린다. 이 모듈은 원본 xlsx(zip)를 복사하면서 대장 시트의 XML만 손보므로
0.5초 안에 끝난다. 서식은 원본 첫 데이터 행의 스타일 인덱스(s=)를 견본으로 상속한다.

행을 지우면 함께 정리해야 하는 것들(안 하면 Excel 복구 경고·유령 표식이 남는다):
  calcChain / 데이터 영역 병합 / autoFilter·조건부서식 범위 / 메모(코멘트)와 VML 도형
"""
import re
import shutil
import zipfile
from xml.sax.saxutils import escape

LEDGER_SHEET = '통합 ALC2 코드'
# 헤더명 → 논리 필드 (열 순서가 바뀌어도 헤더로 찾는다)
HEADER_MAP = {'NO': 'no', '차종': 'vehicle', 'ALC-2 CODE': 'alc2', 'KMC ALC-2 CODE': 'kmc'}


def _col_letter(ref):
    return re.match(r'([A-Z]+)', ref).group(1)


def _sheet_part(zf, sheet_name):
    """시트 이름 → xl/worksheets/sheetN.xml 경로."""
    wbx = zf.read('xl/workbook.xml').decode('utf-8')
    rid = None
    for m in re.finditer(r'<sheet\b[^>]*?name="([^"]+)"[^>]*?r:id="([^"]+)"[^>]*/>', wbx):
        if m.group(1) == sheet_name:
            rid = m.group(2)
            break
    if rid is None:  # 이름이 바뀐 경우 첫 시트
        m = re.search(r'<sheet\b[^>]*?r:id="([^"]+)"[^>]*/>', wbx)
        if not m:
            return None
        rid = m.group(1)
    rels = zf.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    m = re.search(r'Id="%s"[^>]*Target="([^"]+)"' % re.escape(rid), rels)
    if not m:
        return None
    tgt = m.group(1).lstrip('/')
    return tgt if tgt.startswith('xl/') else 'xl/' + tgt


def find_columns(path, sheet_name=LEDGER_SHEET):
    """헤더 행을 찾아 {필드: 열문자} 와 첫 데이터 행 번호를 돌려준다."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
    cols, hdr_row, ncol = {}, 0, 0
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=12, max_col=30, values_only=True), 1):
        vals = {str(v).strip().upper(): i for i, v in enumerate(row, 1) if v is not None}
        if 'KMC ALC-2 CODE' in vals:
            hdr_row = r
            for want, key in HEADER_MAP.items():
                if want in vals:
                    cols[key] = get_column_letter(vals[want])
                    if key == 'no':
                        ncol = vals[want]
            break
    # 첫 데이터 행 = 헤더 아래에서 KMC 코드가 처음 채워진 행
    first_data = 0
    if hdr_row and 'kmc' in cols:
        kc = column_index_from_string(cols['kmc'])
        for r, row in enumerate(ws.iter_rows(min_row=hdr_row + 1, min_col=kc, max_col=kc,
                                             values_only=True), hdr_row + 1):
            if row[0] is not None and str(row[0]).strip():
                first_data = r
                break
    wb.close()
    return cols, first_data or (hdr_row + 1)


def _cell_xml(ref, style, value):
    s = ' s="%s"' % style if style else ''
    if value is None or value == '':
        return '<c r="%s"%s/>' % (ref, s)
    if isinstance(value, (int, float)):
        return '<c r="%s"%s><v>%s</v></c>' % (ref, s, value)
    return '<c r="%s"%s t="inlineStr"><is><t>%s</t></is></c>' % (ref, s, escape(str(value)))


def replace_rows(src, dst, values_by_row, first_data_row, sheet_name=LEDGER_SHEET):
    """헤더(서식·색상·열)는 그대로 두고, first_data_row 이후의 기존 데이터를 전부 지운 뒤
       values_by_row를 first_data_row부터 다시 채운다.
       values_by_row: [{열문자: 값}, ...]. 반환: 기록된 행 수."""
    zin = zipfile.ZipFile(src)
    part = _sheet_part(zin, sheet_name)
    if part is None:
        zin.close(); shutil.copy2(src, dst); return 0
    xml = zin.read(part).decode('utf-8')

    ds = xml.find('<sheetData')
    de = xml.find('</sheetData>')
    if ds < 0 or de < 0:
        zin.close(); shutil.copy2(src, dst); return 0
    body = xml[xml.index('>', ds) + 1:de]

    keep, template = [], None
    for m in re.finditer(r'<row\b[^>]*?(?:/>|>.*?</row>)', body, re.S):
        blk = m.group(0)
        rn = re.search(r'\br="(\d+)"', blk)
        if rn and int(rn.group(1)) >= first_data_row:
            if template is None:
                template = blk          # 첫 데이터 행 = 서식 견본
            continue                     # 기존 데이터 행은 버린다
        keep.append(blk)

    if template is None:                 # 데이터가 없던 서식이면 마지막 헤더 행을 견본으로
        template = keep[-1] if keep else '<row r="%d"></row>' % first_data_row

    attrs = re.match(r'<row\b([^>]*?)/?>', template).group(1)
    attrs_tpl = re.sub(r'\br="\d+"', 'r="{R}"', attrs)
    styles = {}
    for cm in re.finditer(r'<c\b([^>]*?)/?>', template):
        a = cm.group(1)
        rm = re.search(r'r="([A-Z]+\d+)"', a)
        if not rm:
            continue
        sm = re.search(r's="(\d+)"', a)
        styles[_col_letter(rm.group(1))] = sm.group(1) if sm else ''
    order = sorted(styles, key=lambda c: (len(c), c))

    out = []
    for i, vals in enumerate(values_by_row):
        rn = first_data_row + i
        cells = [_cell_xml('%s%d' % (c, rn), styles.get(c, ''), vals.get(c)) for c in order]
        for c in vals:                   # 견본에 없던 열도 채운다
            if c not in styles:
                cells.append(_cell_xml('%s%d' % (c, rn), '', vals[c]))
        out.append('<row%s>%s</row>' % (attrs_tpl.format(R=rn), ''.join(cells)))

    xml = xml[:xml.index('>', ds) + 1] + ''.join(keep) + ''.join(out) + xml[de:]
    last_row = first_data_row + len(values_by_row) - 1
    xml = re.sub(r'(<dimension ref="[A-Z]+\d+:[A-Z]+)\d+"',
                 lambda mm: '%s%d"' % (mm.group(1), max(last_row, first_data_row)), xml, count=1)
    xml = re.sub(r'<mergeCells[^>]*>.*?</mergeCells>',
                 lambda mm: _trim_merges(mm.group(0), first_data_row), xml, count=1, flags=re.S)
    # 자동필터·조건부서식이 삭제된 행 범위를 계속 가리키면 빈 행까지 필터가 걸린다
    xml = re.sub(r'(<autoFilter ref="[A-Z]+\d+:[A-Z]+)\d+"',
                 lambda mm: '%s%d"' % (mm.group(1), last_row), xml, count=1)
    xml = re.sub(r'<conditionalFormatting sqref="([^"]+)">(.*?)</conditionalFormatting>',
                 lambda mm: _clamp_cf(mm.group(0), mm.group(1), last_row), xml, flags=re.S)

    # 데이터 행과 함께 수식도 사라지므로 calcChain은 버린다(남기면 Excel 복구 경고).
    #   → 참조하는 [Content_Types].xml 항목과 workbook 관계도 같이 정리해야 한다.
    drop = {n for n in zin.namelist() if n.endswith('calcChain.xml')}
    # 삭제된 행에 달려 있던 메모(코멘트)와 그 VML 도형도 같이 걷어낸다.
    # 그대로 두면 빈 셀에 메모 표식이 남는다. 헤더 행 메모(열 설명)는 보존.
    cparts = _comment_parts(zin, part)
    with zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            if it.filename in drop:
                continue
            if it.filename == part:
                data = xml.encode('utf-8')
            elif it.filename in cparts:
                data = _trim_comment_part(it.filename, zin.read(it.filename), first_data_row - 1)
            elif it.filename == '[Content_Types].xml':
                ct = zin.read(it.filename).decode('utf-8')
                data = re.sub(r'<Override[^>]*calcChain\.xml"[^>]*/>', '', ct).encode('utf-8')
            elif it.filename == 'xl/_rels/workbook.xml.rels':
                rl = zin.read(it.filename).decode('utf-8')
                data = re.sub(r'<Relationship[^>]*calcChain\.xml"[^>]*/>', '', rl).encode('utf-8')
            else:
                data = zin.read(it.filename)
            zout.writestr(it, data)
    zin.close()
    return len(values_by_row)


def _comment_parts(zf, sheet_part):
    """해당 시트에 연결된 comments/vml/threadedComments 파트 경로."""
    rels = 'xl/worksheets/_rels/%s.rels' % sheet_part.rsplit('/', 1)[-1]
    out = set()
    try:
        txt = zf.read(rels).decode('utf-8')
    except KeyError:
        return out
    for tgt in re.findall(r'Target="([^"]+)"', txt):
        n = tgt.replace('../', 'xl/').lstrip('/')
        if 'comments' in n.lower() or n.lower().endswith('.vml'):
            out.add(n)
    for n in list(out):                       # threadedComments는 comments가 참조
        if n.endswith('.vml'):
            continue
        r2 = n.rsplit('/', 1)
        try:
            t = zf.read('%s/_rels/%s.rels' % (r2[0], r2[-1])).decode('utf-8')
            for tgt in re.findall(r'Target="([^"]+)"', t):
                out.add(tgt.replace('../', 'xl/').lstrip('/'))
        except KeyError:
            pass
    return {n for n in out if n in zf.namelist()}


def _trim_comment_part(name, raw, keep_upto):
    """헤더 영역(keep_upto 행 이하) 메모만 남긴다.
       옛 데이터 행에 달렸던 메모가 새 코드 위에 남으면 오해를 부른다.
       comments/threadedComments는 ref 행으로, VML은 <x:Row>(0-base)로 걸러낸다."""
    txt = raw.decode('utf-8')
    if name.lower().endswith('.vml'):
        def keep_shape(m):
            rm = re.search(r'<x:Row>(\d+)</x:Row>', m.group(0))
            return '' if rm and int(rm.group(1)) + 1 > keep_upto else m.group(0)
        return re.sub(r'<v:shape\b.*?</v:shape>', keep_shape, txt, flags=re.S).encode('utf-8')
    tag = 'threadedComment' if 'threaded' in name.lower() else 'comment'

    def keep(m):
        rm = re.search(r'ref="[A-Z]+(\d+)"', m.group(0))
        return '' if rm and int(rm.group(1)) > keep_upto else m.group(0)

    pat = r'<%s\b[^>]*?(?:/>|>.*?</%s>)' % (tag, tag)
    return re.sub(pat, keep, txt, flags=re.S).encode('utf-8')


def _clamp_cf(block, sqref, last_row):
    """조건부서식 범위를 새 마지막 행까지로 자른다. 전부 벗어나면 규칙 자체를 뺀다."""
    out = []
    for rng in sqref.split():
        parts = rng.split(':')
        starts = re.match(r'([A-Z]+)(\d+)$', parts[0])
        if not starts:
            out.append(rng)
            continue
        if len(parts) == 1:
            if int(starts.group(2)) <= last_row:
                out.append(rng)
            continue
        ends = re.match(r'([A-Z]+)(\d+)$', parts[1])
        if not ends:
            out.append(rng)
            continue
        if int(starts.group(2)) > last_row:
            continue                                   # 범위 전체가 삭제 영역
        e = min(int(ends.group(2)), last_row)
        out.append('%s:%s%d' % (parts[0], ends.group(1), e))
    if not out:
        return ''
    return block.replace('sqref="%s"' % sqref, 'sqref="%s"' % ' '.join(out), 1)


def _trim_merges(block, first_data_row):
    """데이터 영역(삭제 대상)에 걸린 병합만 제거하고 헤더 병합은 보존."""
    kept = [m for m in re.findall(r'<mergeCell ref="([A-Z]+\d+:[A-Z]+\d+)"/>', block)
            if int(re.search(r'(\d+)$', m.split(':')[0]).group(1)) < first_data_row]
    if not kept:
        return ''
    return '<mergeCells count="%d">%s</mergeCells>' % (
        len(kept), ''.join('<mergeCell ref="%s"/>' % r for r in kept))
