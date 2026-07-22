"""
부품사양서 → BOM 자동 생성기 (표준화 양식 기반)

흐름:
  1) 부품사양서 파싱
     - 사양군 (C~H, K2~K7) ○ → 국가/지역 코드  (예: "일반+중동")
     - 옵션제약 COMB OPT1~OPT21 PEL CODE → DESCRIPTION (PEL 마스터 룩업, +로 연결)
     - B2 옆 "Level1 P/NO" 사용자 입력값 → 시트 어셈블리 1레벨 품번
  2) 표준화 양식 로드 (활성 리비전)
  3) 양식 셀에 데이터 채워서 저장
     - 각 VC당 1행 (rowN = 8 + VC인덱스)
     - 매트릭스: W열부터 가로로 VC당 1컬럼

양식이 자주 바뀌므로 컬럼은 헤더 텍스트로 동적 탐색.
"""
import os, shutil
from copy import copy as _copy_style
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
from datetime import datetime


# ════════════════════════════════════════════════════════════════════════════
# 1. 부품사양서 파싱
# ════════════════════════════════════════════════════════════════════════════
def _is_mark(v) -> bool:
    """○, ●, * 등 마킹 문자 판정."""
    if v is None: return False
    s = str(v).strip()
    return bool(s) and s in ('○', '●', '◯', '*', '✓', 'O', 'o', 'V', 'v')


def parse_part_spec(filepath: str) -> dict:
    """부품사양서 → VC 리스트.

    각 VC: {
      'vc': '001',
      'region': '일반+중동',
      'opts': [{'opt': 'OPT1', 'pel_code': '5693A3'}, ...],
    }
    + 헤더 메타: opt_labels, level1_pno, vehicle_info, ...
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # ── 1) 헤더 행 자동 탐색 (A열에 "UPG VC" 있는 행)
    group_row = None
    for r in range(1, min(ws.max_row + 1, 30)):
        v = ws.cell(r, 1).value
        if v and 'UPG VC' in str(v):
            group_row = r
            break
    if not group_row:
        raise ValueError('헤더를 찾을 수 없습니다 (1열에 "UPG VC" 없음)')

    sub_row    = group_row + 1   # COMB / EXCL
    detail_row = group_row + 2   # 라벨 (일반, 중동, OPT1, OPT2 ...)
    data_start = group_row + 3

    # ── 2) "사양군" 그룹 열 범위 탐색
    spec_group_start = None
    spec_group_end = None
    opt_start_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(group_row, c).value
        if not v: continue
        s = str(v).strip()
        if s == '사양군':
            spec_group_start = c
        elif spec_group_start and spec_group_end is None and s and s != '사양군':
            spec_group_end = c - 1
        if '옵션제약' in s:
            opt_start_col = c
            if spec_group_end is None and spec_group_start:
                spec_group_end = c - 1
            break
    if not opt_start_col:
        raise ValueError('"옵션제약" 그룹을 찾을 수 없습니다')

    # ── 3) 옵션제약 안에서 COMB 서브섹션만 추출
    initial_sub = ws.cell(sub_row, opt_start_col).value
    initial_sub = str(initial_sub).strip() if initial_sub else ''

    opt_cols = []
    for c in range(opt_start_col, ws.max_column + 1):
        grp = ws.cell(group_row, c).value
        if c > opt_start_col and grp and str(grp).strip():
            break
        sub = ws.cell(sub_row, c).value
        if c > opt_start_col and sub and str(sub).strip() and str(sub).strip() != initial_sub:
            break
        lbl = ws.cell(detail_row, c).value
        if not lbl or not str(lbl).upper().startswith('OPT'):
            break
        opt_cols.append((c, str(lbl).strip()))
    if not opt_cols:
        raise ValueError('OPT 컬럼을 찾을 수 없습니다')

    # ── 4) 사양군 라벨 (행 7) 미리 추출
    region_labels = {}
    if spec_group_start:
        # spec_group_end가 None이면 옵션제약 직전까지
        end = spec_group_end or (opt_start_col - 1)
        for c in range(spec_group_start, end + 1):
            lbl = ws.cell(detail_row, c).value
            if lbl: region_labels[c] = str(lbl).strip()

    # ── 5) B2 옆에서 "Level1 P/NO" 라벨 + 값 탐색
    level1_pno = ''
    for r in range(1, 6):
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if v and 'Level1 P/NO' in str(v):
                # 옆 셀 (오른쪽) 또는 같은 셀 안에 ":" 뒤 값
                s = str(v).strip()
                if ':' in s and s.split(':', 1)[1].strip():
                    level1_pno = s.split(':', 1)[1].strip()
                else:
                    nxt = ws.cell(r, c + 1).value
                    if nxt: level1_pno = str(nxt).strip()
                break
        if level1_pno: break

    # ── 6) 데이터 행: VC별로 사양군 ○ 라벨 모으기 + OPT PEL CODE 모으기
    rows = []
    for r in range(data_start, ws.max_row + 1):
        vc_raw = ws.cell(r, 1).value
        if vc_raw is None: continue
        vc = str(vc_raw).strip()
        if not vc: continue

        # 사양군 영역 ○ 라벨
        region_parts = []
        if spec_group_start and region_labels:
            end = spec_group_end or (opt_start_col - 1)
            for c in range(spec_group_start, end + 1):
                if _is_mark(ws.cell(r, c).value):
                    lab = region_labels.get(c, '')
                    if lab: region_parts.append(lab)
        region = '+'.join(region_parts)

        # OPT PEL CODE
        opts = []
        for col_idx, opt_label in opt_cols:
            pel = ws.cell(r, col_idx).value
            if pel is not None and str(pel).strip():
                opts.append({'opt': opt_label, 'pel_code': str(pel).strip()})

        if opts:
            rows.append({'vc': vc, 'region': region, 'opts': opts})

    return {
        'vehicle_info': str(ws.cell(2, 1).value or ''),
        'opt_count': len(opt_cols),
        'opt_labels': [lbl for _, lbl in opt_cols],
        'level1_pno': level1_pno,
        'spec_group_range': (spec_group_start, spec_group_end),
        'vcs': rows,
    }


# ════════════════════════════════════════════════════════════════════════════
# 2. PEL 마스터 로드
# ════════════════════════════════════════════════════════════════════════════
def load_pel_master(pel_path: str) -> dict:
    if not os.path.exists(pel_path):
        return {'data': {}, 'columns': [], 'code_col': None}
    df = pd.read_excel(pel_path, sheet_name=0).fillna('')
    cols = [str(c) for c in df.columns]

    code_col = None
    for c in cols:
        if str(c).strip().upper() == 'CODE':
            code_col = c
            break
    if not code_col:
        code_col = cols[1] if len(cols) >= 2 else (cols[0] if cols else None)

    master = {}
    for _, row in df.iterrows():
        code = str(row[code_col]).strip() if code_col else ''
        if not code: continue
        master[code] = {c: str(row[c]) for c in cols}
    return {'data': master, 'columns': cols, 'code_col': code_col}


# ════════════════════════════════════════════════════════════════════════════
# 3. 표준화 양식 채우기
# ════════════════════════════════════════════════════════════════════════════
DATA_START_ROW   = 8   # A8부터 VC 데이터 시작
# rev_002(2026-07-22, '생산 공장' 열 추가)로 열이 한 칸씩 밀림:
#   U(21)=지역, V(22)=생산 공장, W(23)=MATERIAL, X(24~)=VC 매트릭스
PLANT_COL        = 22  # V열 = 생산 공장 (BRE에서 자동 채움)
MATERIAL_COL     = 23  # W열 = MATERIAL (기존 V→W 이동)
MATRIX_START_COL = 24  # X열 = 24 (기존 W→X 이동)
MATRIX_VC_ROW    = 4   # X4 = VC 번호
MATRIX_REGION_ROW = 5  # X5 = 지역
MATRIX_LV1_ROW   = 6   # X6 = 1레벨 P/NO


# ── 고객 BRE(BOM Report Excel) 파싱 ──────────────────────────────────────────
BRE_PLANT_MAP = {'BS': '광주', 'DE': '화성'}


def _norm_vc(s) -> str:
    """VC 코드 비교용 정규화 (앞의 0 제거: '001'→'1')."""
    t = str(s).strip()
    if not t:
        return ''
    return t.lstrip('0') or '0'


def parse_bre(path: str) -> dict:
    """고객 BRE(.xlsm) → 생산공장 + VC별 1레벨 P/NO 매핑.
    - 생산공장: 시트명 접두 2글자로 판정 (BS=광주, DE=화성) — 파일명이 같아도 시트로 구분 가능.
    - VC→1레벨 P/NO: 'SEAT ASSY' 를 품명에 포함한 Level1 행이 각 VC열에 1.0 표기된 것으로 매핑.
      (BRE는 VC마다 SEAT ASSY 품번이 다르므로 VC 중복이어도 공장별로 정확히 구분됨)
    """
    import re as _re
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb.worksheets[0]
        sheet = ws.title or ''
        code = sheet[:2].upper()
        plant = BRE_PLANT_MAP.get(code, '')
        maxc = min(ws.max_column, 400)   # VC열은 J부터 연속 — 과도한 열 스캔 방지

        # 헤더행 탐색: A열이 'Level' 인 행 (없으면 5행 가정)
        header_row = None
        for r in range(1, 12):
            if str(ws.cell(r, 1).value or '').strip().lower() == 'level':
                header_row = r
                break
        if header_row is None:
            header_row = 5
        vc_row = header_row + 1        # VC 코드가 적힌 행
        data_start = header_row + 2    # 데이터 시작

        # VC 열 찾기 (2~3자리 숫자코드: '001','032' 등)
        vc_cols = {}
        for c in range(1, maxc + 1):
            v = ws.cell(vc_row, c).value
            s = str(v).strip() if v is not None else ''
            if _re.fullmatch(r'\d{2,3}', s):
                vc_cols[c] = s

        # 데이터: Level==1 & 품명에 'SEAT ASSY' 포함 → 표기된 VC열마다 P/NO 매핑
        vc_level1 = {}
        for r in range(data_start, ws.max_row + 1):
            lv = str(ws.cell(r, 1).value or '').strip()
            if lv not in ('1', '1.0'):
                continue
            pno = str(ws.cell(r, 4).value or '').strip()    # D: NEW PART-NO
            pname = str(ws.cell(r, 6).value or '').strip()  # F: NEW PART NAME
            if not pno or 'SEAT ASSY' not in pname.upper():
                continue
            for c, vc in vc_cols.items():
                mk = ws.cell(r, c).value
                if mk is not None and str(mk).strip() not in ('', '0'):
                    vc_level1[_norm_vc(vc)] = pno
        return {'plant': plant, 'plant_code': code, 'sheet': sheet,
                'vc_level1': vc_level1, 'vc_count': len(vc_cols),
                'matched_vc': len(vc_level1)}
    finally:
        wb.close()


def generate_bom_from_sources(sources: list, pel_path: str,
                               template_path: str, output_path: str) -> dict:
    """차종(운전석 등) 기준 — 공장별 소스(부품사양서 + 선택 BRE)를 순서대로 이어붙여
    하나의 표준 BOM으로 생성. 공장이 1개만 있어도 그대로 동작(호환).

    sources: [{'spec_path': str, 'bre_info': dict|None, 'plant_label': str(폴백 공장명)}, ...]
    """
    if not sources:
        raise ValueError('부품사양서가 1개 이상 필요합니다.')

    master_info = load_pel_master(pel_path)
    master = master_info['data']
    master_cols = master_info['columns']

    def pick_col(cands):
        for c in master_cols:
            cu = str(c).strip().upper()
            for cand in cands:
                if cand in cu or cand in str(c):
                    return c
        return None

    spec_col = pick_col(['사양', '명칭', 'NAME', 'SPEC']) or (master_cols[2] if len(master_cols) > 2 else None)

    def pel_to_name(code: str) -> str:
        """PEL 코드 → 사양(spec) 텍스트.
        사양이 비어 있으면 원본 코드 그대로 유지 (설명으로 fallback 하지 않음 — 사용자 명시)."""
        entry = master.get(code)
        if not entry: return code
        if spec_col:
            sp = str(entry.get(spec_col, '')).strip()
            if sp: return sp
        return code

    # ── 소스별(공장별) 부품사양서 파싱 → VC 블록을 순서대로 이어붙임
    #    (예: 광주 57개 + 화성 101개 → 158행짜리 하나의 표준 BOM)
    parsed_sources = []
    for src in sources:
        spec_i = parse_part_spec(src['spec_path'])
        bre_info_i = src.get('bre_info') or {}
        plant_i = bre_info_i.get('plant') or src.get('plant_label', '')
        parsed_sources.append({
            'spec': spec_i, 'plant': plant_i,
            'lv1_map': bre_info_i.get('vc_level1', {}),
            'bre_info': bre_info_i,
        })

    combined = []  # (vc_block, plant_name, lv1_map, level1_pno_fallback)
    for ps in parsed_sources:
        for vb in ps['spec']['vcs']:
            combined.append((vb, ps['plant'], ps['lv1_map'], ps['spec']['level1_pno']))
    if not combined:
        raise ValueError('업로드한 부품사양서에서 VC 데이터를 찾지 못했습니다.')

    first_spec = parsed_sources[0]['spec']

    # ── 템플릿 복사해서 작업본 만들기 (서식 보존)
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # ── 행 8 (첫 데이터 행)의 서식 + 행 높이 캐싱 → 데이터 행 추가 시 복제용
    ref_row = DATA_START_ROW
    ref_height = ws.row_dimensions[ref_row].height
    ref_styles = {}  # {col_idx: (font, fill, border, alignment, number_format)}
    last_styled_col = max(ws.max_column, MATRIX_START_COL + len(combined) + 10)
    for c in range(1, last_styled_col + 1):
        cell = ws.cell(ref_row, c)
        if cell.has_style:
            ref_styles[c] = (
                _copy_style(cell.font),
                _copy_style(cell.fill),
                _copy_style(cell.border),
                _copy_style(cell.alignment),
                cell.number_format,
            )
    # L열(P/NAME) 기본값 — 양식의 placeholder 그대로 사용
    default_pname = str(ws.cell(ref_row, 12).value or 'SEAT ASSY-FR, LH')

    def _apply_ref_style(row_idx: int):
        """ref_row 의 서식을 row_idx 에 통째로 복제 (행 높이 포함)."""
        if row_idx == ref_row: return
        if ref_height is not None:
            ws.row_dimensions[row_idx].height = ref_height
        for c, (font, fill, border, align, fmt) in ref_styles.items():
            cell = ws.cell(row_idx, c)
            cell.font = _copy_style(font)
            cell.fill = _copy_style(fill)
            cell.border = _copy_style(border)
            cell.alignment = _copy_style(align)
            cell.number_format = fmt

    stats = {
        'vc_count': len(combined),
        'opt_count': first_spec['opt_count'],
        'pel_total': 0, 'matched': 0, 'unmatched': 0,
        'unmatched_codes': set(),
        'fully_matched_vc': 0, 'partial_vc': 0,
        'level1_pno': first_spec['level1_pno'],
        'level1_pno_missing': True,   # 채워지는 VC 있으면 아래서 False 처리
        'no_region_vc': 0,
        'bre_lv1_filled': 0,
    }

    bad_fill = PatternFill('solid', fgColor='FFCDD2')
    warn_fill = PatternFill('solid', fgColor='FFF8E1')

    # ── VC별로 양식 채우기 (공장별 소스를 이어붙인 combined 순서대로)
    for idx, (vc_block, plant_name, vc_level1_map, level1_fallback) in enumerate(combined):
        row = DATA_START_ROW + idx
        col = MATRIX_START_COL + idx
        vc = vc_block['vc']
        region = vc_block['region']

        # DESCRIPTION 만들기
        names, unmatched_in_vc = [], []
        for op in vc_block['opts']:
            pel = op['pel_code']
            entry = master.get(pel)
            if entry:
                stats['matched'] += 1
                names.append(pel_to_name(pel))
            else:
                stats['unmatched'] += 1
                stats['unmatched_codes'].add(pel)
                unmatched_in_vc.append(pel)
                names.append(f'?{pel}?')
            stats['pel_total'] += 1
        description = ' + '.join(names) if names else '(빈 VC)'

        if not region: stats['no_region_vc'] += 1
        if unmatched_in_vc: stats['partial_vc'] += 1
        else: stats['fully_matched_vc'] += 1

        # 1레벨 P/NO: BRE에 해당 VC 매칭이 있으면 우선, 없으면 그 공장 부품사양서의 전역값
        lv1 = vc_level1_map.get(_norm_vc(vc)) or level1_fallback
        if lv1:
            stats['level1_pno_missing'] = False
        if vc_level1_map.get(_norm_vc(vc)):
            stats['bre_lv1_filled'] += 1

        # 좌측 데이터 행
        _apply_ref_style(row)                                    # 서식 + 행 높이 복제
        ws.cell(row, 1).value = vc                              # A: VC 번호
        ws.cell(row, 2).value = 1                                # B: LEVEL = 1
        if lv1:
            ws.cell(row, 10).value = lv1                         # J: 1레벨 P/NO
        ws.cell(row, 12).value = default_pname                   # L: P/NAME (양식 placeholder)
        ws.cell(row, 14).value = description                     # N: DESCRIPTION
        if region:
            ws.cell(row, 21).value = region                      # U: 지역
        if plant_name:
            ws.cell(row, PLANT_COL).value = plant_name           # V: 생산 공장 (BRE)
        ws.cell(row, MATERIAL_COL).value = 'ASSY'                # W: MATERIAL

        if unmatched_in_vc:
            ws.cell(row, 14).fill = warn_fill

        # 매트릭스 상단 (X4, X5, X6)
        ws.cell(MATRIX_VC_ROW,     col).value = vc
        if region:
            ws.cell(MATRIX_REGION_ROW, col).value = region
        if lv1:
            ws.cell(MATRIX_LV1_ROW,    col).value = lv1

        # 교차점 QTY 마커 (대각선)
        ws.cell(row, col).value = 1

    # ── 템플릿이 갖고 있던 placeholder 행 정리 (실제 VC 수보다 많을 때)
    last_used = DATA_START_ROW + len(combined) - 1
    for r in range(last_used + 1, ws.max_row + 1):
        v = ws.cell(r, 12).value  # L열 placeholder 확인
        if v is None: continue
        # 데이터 값만 비우고 서식은 유지 (V=22 생산공장, W=23 MATERIAL 포함)
        for c in (1, 2, 10, 12, 14, 21, 22, 23):
            ws.cell(r, c).value = None

    wb.save(output_path)

    return {
        'total': stats['pel_total'],
        'matched': stats['matched'],
        'unmatched': stats['unmatched'],
        'unmatched_codes': sorted(stats['unmatched_codes']),
        'unmatched_unique_count': len(stats['unmatched_codes']),
        'vc_count': stats['vc_count'],
        'opt_count': stats['opt_count'],
        'fully_matched_vc': stats['fully_matched_vc'],
        'partial_vc': stats['partial_vc'],
        'opt_labels': first_spec['opt_labels'],
        'vehicle_info': first_spec['vehicle_info'],
        'level1_pno': stats['level1_pno'],
        'level1_pno_missing': stats['level1_pno_missing'],
        'no_region_vc': stats['no_region_vc'],
        'bre_plant': ' + '.join(sorted({p['plant'] for p in parsed_sources if p['plant']})),
        'bre_lv1_filled': stats['bre_lv1_filled'],
        'bre_vc_count': sum(p['bre_info'].get('matched_vc', 0) for p in parsed_sources),
        'plants_used': [{'plant': p['plant'] or ps_src.get('plant_label', ''), 'vc_count': len(p['spec']['vcs'])}
                        for p, ps_src in zip(parsed_sources, sources)],
    }


# ════════════════════════════════════════════════════════════════════════════
# 4. 호환용 진입점 — main.py가 부르는 generate_bom(...)
# ════════════════════════════════════════════════════════════════════════════
def generate_bom(sources: list, pel_path: str, output_path: str,
                 template_path: str = None) -> dict:
    """활성 표준화 양식이 있으면 그걸 기반으로 채움.
    없으면 명시적 에러 (사용자가 admin에게 양식 등록 요청해야 함).

    sources: [{'spec_path': str, 'bre_info': dict|None, 'plant_label': str}, ...]
    차종(운전석 등) 기준으로 공장별(광주/화성) 소스를 여러 개 넘기면 하나의 표준 BOM으로 합쳐진다.
    """
    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError(
            '활성 표준화 BOM 양식이 등록되지 않았습니다. '
            '관리자에게 [📚 리비전 관리] 메뉴에서 양식 등록을 요청하세요.'
        )
    return generate_bom_from_sources(sources, pel_path, template_path, output_path)
